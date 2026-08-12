# %% [markdown]
# # 2026 ISP inputs and assumptions workbook — semantic tables
#
# This notebook reads the semantic source tables identified by the discovery notebook.
# Source ranges are explicit Excel coordinates. Semantic headers and data rows are reconstructed from
# workbook structure, while genuinely missing cells remain missing unless a table-specific rule applies.

# %%
from __future__ import annotations

import os
import warnings
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries

try:
    from IPython import get_ipython
    from IPython.display import display
except ImportError:
    def get_ipython():
        return None
    display = print

_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_NS_REL_DOC = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_NS_REL_PKG = 'http://schemas.openxmlformats.org/package/2006/relationships'
warnings.filterwarnings('ignore', message='Data Validation extension is not supported and will be removed', category=UserWarning, module='openpyxl.worksheet._reader')

def _repo_root() -> Path:
    try:
        return Path(__file__).resolve().parent.parent
    except NameError:
        cwd = Path.cwd().resolve()
        for candidate in (cwd, *cwd.parents):
            if (candidate / 'Project.toml').is_file() and (candidate / 'scripts').is_dir():
                return candidate
        raise RuntimeError('Could not locate the ParseISP.jl repository root.')

_PROJECT = _repo_root()
_DEFAULT_WORKBOOK = _PROJECT / 'scrapped' / 'ISP_reference_files' / '2026-isp-inputs-and-assumptions-workbook.xlsm'
WORKBOOK_PATH = Path(os.environ.get('PARSEISP_2026_INPUTS_WORKBOOK', _DEFAULT_WORKBOOK)).expanduser().resolve()
if not WORKBOOK_PATH.is_file():
    raise FileNotFoundError(f'ISP 2026 inputs workbook not found: {WORKBOOK_PATH}')

SHEET_RANGES = {'Disclaimer': [],
 'Change Log': ['B2:E657'],
 'Assumptions Summary': ['B6:D35', 'B37:E119', 'B121:C147'],
 'Scenarios': ['B5:E29'],
 'Existing Gen Data Summary': ['B10:AT738'],
 'New Entrant Data Summary': ['B9:BB535'],
 'New Electrolyser Data Summary': ['B5:AQ67'],
 'Fuel Price Summary': ['B8:S8', 'B9:S9', 'B11:AK738', 'B743:AK1268'],
 'Regional Build Costs Summary': ['B7:C10', 'B12:AV75'],
 'Energy Policy Targets': ['C15:E30',
                           'C32:F62',
                           'C65:G93',
                           'C96:E110',
                           'C114:N154',
                           'C157:E159',
                           'C161:I185',
                           'C187:F193',
                           'C195:E197',
                           'C200:E202',
                           'C206:E209',
                           'C213:E220',
                           'C223:E235',
                           'C239:F258',
                           'C262:E267',
                           'C271:E289',
                           'D291:E295',
                           'C297:E310',
                           'C313:E327'],
 'Carbon Budgets': ['B5:E9', 'B15:G21', 'B25:D31'],
 'Economic Growth Forecasts': ['B5:E8', 'B12:AG19', 'B21:AG28', 'B30:AG37', 'B41:AG48', 'B50:AG57', 'B59:AG66'],
 'Demand and Energy Forecasts': [],
 'End use fuel consumption data': ['B6:AF15', 'B17:AF26', 'B28:AF37'],
 'Appliance Uptake Forecasts': ['B13:AG20', 'B22:AG29', 'B31:AG38'],
 'Elec. Retail Price Indices': ['B8:AG12'],
 'Connections Forecasts': ['B8:E11', 'B15:AH22', 'B24:AH31', 'B33:AH40'],
 'Energy Efficiency': ['B8:G10',
                       'B14:AG30',
                       'B32:AG48',
                       'B50:AG66',
                       'B68:AG84',
                       'B86:AG102',
                       'B106:AG122',
                       'B124:AG140',
                       'B142:AG158',
                       'B160:AG176',
                       'B178:AG194'],
 'Rooftop PV': ['B8:E10', 'B12:AH63', 'B65:AH116'],
 'PVNSG': ['B8:E10', 'B12:AH63', 'B65:AH116'],
 'ONSG': ['B8:AH55', 'B57:AG74'],
 'Battery & Plug-in EVs': ['B7:E9', 'B11:AH62', 'B64:AH115'],
 'Fuel cell EVs': ['B7:E9', 'B11:AH62'],
 'EV V2G': ['B8:E10', 'B12:AH62', 'B64:AH115'],
 'Data Centre Forecasts': ['B6:E8', 'B10:AF16', 'B18:AF24', 'B26:AF32'],
 'DSP': ['B7:AI84', 'B87:AI164'],
 'Electrification': ['B7:E10', 'B12:AF19', 'B21:AF28', 'B30:AF37'],
 'Embedded energy storages': ['B7:E9', 'B11:AH62', 'B65:AH116'],
 'Aggregated energy storages': ['B7:E9', 'B11:AH62', 'B65:AH116'],
 'Network representation': ['B2:E22', 'B24:D42', 'B44:D53', 'B55:D63', 'B65:D82'],
 'Renewable energy zones': ['B6:E53'],
 'Network capability': ['B6:K25', 'B34:K42', 'B51:N60', 'B75:V84', 'B89:E94', 'B99:D115', 'B122:C134', 'B139:C148'],
 'Network losses': ['B5:J28', 'B30:J34', 'B36:J88'],
 'Transmission Reliability': ['B7:E13'],
 'Distribution network': ['B11:G38', 'B40:H57', 'B59:AZ1433'],
 'Connection cost': ['B6:J61', 'B62:R73'],
 'Connection cost forecasts': ['B8:AJ144', 'B147:AJ388'],
 'Flow path augmentation options': ['B11:Q127'],
 'Flow path cost forecasts': ['B10:AI111', 'B115:AI216', 'B220:AI321'],
 'REZ augmentations options': ['B10:O37', 'B39:O77', 'B79:O96', 'B98:O110', 'B112:O137'],
 'REZ cost forecasts': ['B11:AJ117', 'B118:AJ224', 'B225:AJ331'],
 'Distribution cost forecasts': ['B5:AJ84'],
 'Maximum capacity': ['B9:J750', 'L9:O31'],
 'Hybrid site limits': ['B9:G67'],
 'Seasonal ratings': ['B9:E36', 'B42:AI770'],
 'Generator Reliability Settings': ['B9:M16', 'B21:M60', 'B62:H90'],
 'Maintenance': ['B5:D29', 'G5:I32'],
 'Retirement': ['B8:F738', 'H8:I50'],
 'Hydro Scheme Inflows': ['B4:T79', 'B81:T121', 'B123:T141', 'B143:T162'],
 'Capacity Factors ': ['B2:V214'],
 'Heat rates': ['B7:E740', 'H7:I31'],
 'Auxiliary': ['B5:E736', 'G5:H29'],
 'Storage properties': ['B2:J19', 'B21:E35', 'G21:J27', 'B38:C45'],
 'Emissions intensity': ['B4:E744', 'G4:H29'],
 'Build costs': ['B2:AJ77'],
 'Fixed OPEX': ['B5:E739', 'G5:I32'],
 'Variable OPEX': ['B5:E738', 'G5:H32'],
 'Marginal Loss Factors': ['B10:G748', 'I10:M536', 'O10:S161'],
 'Locational Cost Factors': ['B9:H80', 'B83:I132', 'B134:G158', 'B161:X227'],
 'Build limits - REZs': ['B2:Q62',
                         'B64:N119',
                         'B121:F132',
                         'B136:K265',
                         'B267:K317',
                         'B319:K335',
                         'B337:E356',
                         'B358:G368'],
 'Build limits - PHES': ['B2:W27'],
 'First-of-a-kind premium': ['B2:D11'],
 'Lead time and project life': ['B2:H35'],
 'Financial parameters': ['B2:F7', 'B10:F41', 'B43:G51', 'B54:C90'],
 'Affine Heat rates': ['B6:F192', 'H6:K29'],
 'Max Ramp Rates': ['B7:F191', 'H7:J30'],
 'Coal Min Stable Level': ['B2:G63'],
 'GPG Min Stable Level': ['B10:E150', 'G10:H35'],
 'Coal and Biomass price': ['B8:AG54', 'B57:AG61'],
 'Gas, Liquid fuel, H2 price': ['B7:AG129',
                                'B132:AG224',
                                'B228:AG249',
                                'B253:AG274',
                                'B278:AG302',
                                'B305:AG429',
                                'B433:AG438',
                                'B440:AG452'],
 'Gas System Properties': ['B7:F49', 'B51:G105', 'B108:H122', 'B127:F130', 'B132:E144', 'B146:E169', 'B171:E185'],
 'GPG emissions reduction - BioM': ['B2:AF12'],
 'Power System Security': ['B4:D49', 'B52:AE56', 'B58:G72', 'B74:G94'],
 'Reserves': ['B2:C14'],
 'Hydrogen demand - Domestic': ['B2:AH53'],
 'Hydrogen monthly profiles': ['B2:AG44'],
 'Hydrogen demand-Export&Commod': ['B2:AH52', 'B54:AH105', 'B107:AH156'],
 'Hydrogen consumption locations': ['B7:C9', 'B15:C18', 'B24:F40', 'B42:B44', 'B46:D57'],
 'Water for Hydrogen': ['B2:AH52'],
 'Desalination demand for H2': ['B2:AH52'],
 'H2 as fuel for GPG Limit': ['B2:AG21'],
 'Build Cost - Hydrogen pipeline': ['B2:AJ156'],
 'Other hydrogen assumptions': ['B2:C5', 'B7:AF11', 'B13:AF17', 'B19:AF23', 'B26:AF30', 'B32:C35'],
 'Summary Mapping': ['B2:AF733', 'B734:AF786', 'B790:AF1316', 'B1319:AF1381']}

workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True, keep_links=False)

_SHEET_CACHE: dict[str, pd.DataFrame] = {}

def _sheet_source_frame(sheet_name: str) -> pd.DataFrame:
    if sheet_name in _SHEET_CACHE:
        return _SHEET_CACHE[sheet_name]
    ranges = SHEET_RANGES[sheet_name]
    if not ranges:
        return pd.DataFrame()
    bounds = [range_boundaries(cell_range) for cell_range in ranges]
    min_col = min(bound[0] for bound in bounds); min_row = min(bound[1] for bound in bounds)
    max_col = max(bound[2] for bound in bounds); max_row = max(bound[3] for bound in bounds)
    ws = workbook[sheet_name]
    values = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    frame = pd.DataFrame(values, index=pd.Index(range(min_row, max_row + 1), name="excel_row"), columns=[get_column_letter(c) for c in range(min_col, max_col + 1)])
    _SHEET_CACHE[sheet_name] = frame
    return frame

def read_source_range(sheet_name: str, cell_range: str) -> pd.DataFrame:
    """Read one exact Excel source range without inventing or filling values."""
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    columns = [get_column_letter(c) for c in range(min_col, max_col + 1)]
    return _sheet_source_frame(sheet_name).loc[min_row:max_row, columns].copy()

def _worksheet_xml_path(path: Path, sheet_name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        book = ET.fromstring(archive.read('xl/workbook.xml'))
        rels = ET.fromstring(archive.read('xl/_rels/workbook.xml.rels'))
        rel_map = {rel.attrib['Id']: rel.attrib['Target'] for rel in rels.findall(f'{{{_NS_REL_PKG}}}Relationship')}
        for sheet in book.find(f'{{{_NS}}}sheets'):
            if sheet.attrib['name'] == sheet_name:
                target = rel_map[sheet.attrib[f'{{{_NS_REL_DOC}}}id']].lstrip('/')
                return target if target.startswith('xl/') else 'xl/' + target
    raise KeyError(sheet_name)

def _merged_ranges(path: Path, sheet_name: str) -> list[str]:
    xml_path = _worksheet_xml_path(path, sheet_name)
    result = []
    with zipfile.ZipFile(path) as archive, archive.open(xml_path) as handle:
        for event, elem in ET.iterparse(handle, events=('end',)):
            if elem.tag == f'{{{_NS}}}mergeCell':
                result.append(elem.attrib['ref'])
            elem.clear()
    return result

def _is_note_row(values: list[object]) -> bool:
    nonempty = [value for value in values if not pd.isna(value) and str(value).strip()]
    if not nonempty:
        return False
    first = str(nonempty[0]).strip().lower()
    return first.startswith(('source:', 'note:', 'notes:', '*')) or (
        len(nonempty) == 1 and first.startswith(('http://', 'https://'))
    )


def _unique_headers(values: list[object]) -> list[str]:
    result = []
    counts: dict[str, int] = {}
    for i, value in enumerate(values, 1):
        base = str(value).strip() if not pd.isna(value) and str(value).strip() else f'Value {i}'
        count = counts.get(base, 0) + 1
        counts[base] = count
        result.append(base if count == 1 else f'{base} ({count})')
    return result


def _merged_header_values(sheet_name: str, frame: pd.DataFrame, row: int) -> list[object]:
    values = list(frame.loc[row])
    columns = list(frame.columns)
    col_to_i = {column: i for i, column in enumerate(columns)}
    for merged in _merged_ranges(WORKBOOK_PATH, sheet_name):
        min_col, min_row, max_col, max_row = range_boundaries(merged)
        if not (min_row <= row <= max_row):
            continue
        anchor_col = get_column_letter(min_col)
        if anchor_col not in col_to_i or min_row not in frame.index:
            continue
        anchor = frame.at[min_row, anchor_col]
        if pd.isna(anchor):
            continue
        for col_num in range(min_col, max_col + 1):
            col = get_column_letter(col_num)
            if col in col_to_i and pd.isna(values[col_to_i[col]]):
                values[col_to_i[col]] = anchor
    return values


def _combine_header_rows(sheet_name: str, frame: pd.DataFrame, rows: list[int]) -> list[str]:
    parts = [_merged_header_values(sheet_name, frame, row) for row in rows]
    headers = []
    for col_i in range(len(frame.columns)):
        seen = []
        for row_values in parts:
            value = row_values[col_i]
            if pd.isna(value) or not str(value).strip():
                continue
            text = str(value).strip()
            if not seen or seen[-1] != text:
                seen.append(text)
        headers.append(' — '.join(seen) if seen else None)
    return _unique_headers(headers)


def _detect_header_row(frame: pd.DataFrame) -> int | None:
    rows = [row for row in frame.index if frame.loc[row].notna().any()]
    if not rows:
        return None
    candidates = rows[: min(12, len(rows))]
    best = None
    best_score = float('-inf')
    for row in candidates:
        vals = [v for v in frame.loc[row].tolist() if not pd.isna(v) and str(v).strip()]
        if not vals:
            continue
        text = sum(isinstance(v, str) for v in vals)
        n = len(vals)
        # Titles/prose are usually single-cell; ordinary headers span columns.
        score = n * 3 + text * 2 - (8 if n == 1 else 0)
        # Prefer a row followed by populated records.
        later = [r for r in rows if r > row][:2]
        if later:
            score += sum(int(frame.loc[r].notna().sum()) for r in later) / 4
        if score > best_score:
            best_score, best = score, row
    return best


def parse_semantic_range(
    sheet_name: str,
    cell_range: str,
    *,
    header_rows: list[int] | None = None,
    data_rows: list[int] | range | None = None,
    min_values: int = 1,
    column_names: list[str] | None = None,
) -> pd.DataFrame:
    source = read_source_range(sheet_name, cell_range)
    # Drop columns that contain no values anywhere in the declared source block.
    frame = source.dropna(axis=1, how='all').copy()
    if data_rows is not None:
        rows = [row for row in data_rows if row in frame.index]
    else:
        if header_rows is None:
            detected = _detect_header_row(frame)
            header_rows = [detected] if detected is not None else []
        start = max(header_rows) + 1 if header_rows else int(frame.index.min())
        rows = [row for row in frame.index if row >= start]
    # Keep actual records; source/note prose is provenance rather than table data.
    rows = [
        row for row in rows
        if frame.loc[row].notna().sum() >= min_values and not _is_note_row(frame.loc[row].tolist())
    ]
    result = frame.loc[rows].copy()
    # Columns that are empty across semantic data are not output fields.
    if len(result):
        result = result.dropna(axis=1, how='all')
    if column_names is not None:
        if len(column_names) != len(result.columns):
            raise AssertionError((sheet_name, cell_range, len(column_names), len(result.columns)))
        result.columns = column_names
    elif header_rows:
        header_frame = frame.loc[:, result.columns] if len(result.columns) else frame
        result.columns = _combine_header_rows(sheet_name, header_frame, header_rows)
    elif len(result.columns) == 2:
        result.columns = ['Field', 'Assumption']
    elif len(result.columns) == 1:
        result.columns = ['Statement']
    else:
        result.columns = _unique_headers([None] * len(result.columns))
    result = result.reset_index(drop=True)
    result.attrs.update(source_sheet=sheet_name, source_range=cell_range, header_rows=header_rows or [])
    return result


def parse_flow_path_augmentation_options() -> pd.DataFrame:
    frame = read_source_range('Flow path augmentation options', 'B11:Q127')
    # Rows 12–13 form the semantic header; row 14 onward contains option rows.
    option_rows = frame.index[frame.get('E').notna() & (frame.get('E') != 'Option name')]
    result = frame.loc[option_rows].copy()
    for merged in _merged_ranges(WORKBOOK_PATH, 'Flow path augmentation options'):
        min_col, min_row, max_col, max_row = range_boundaries(merged)
        if min_col <= 2 <= max_col:
            anchor = frame.at[min_row, 'B'] if min_row in frame.index and 'B' in frame.columns else None
            if anchor is not None:
                for row in result.index:
                    if min_row <= row <= max_row and pd.isna(result.at[row, 'B']):
                        result.at[row, 'B'] = anchor
    result.columns = _combine_header_rows('Flow path augmentation options', frame.loc[:, result.columns], [12, 13])
    result = result.reset_index(drop=False).rename(columns={'excel_row': 'source_row'})
    assert len(result) == 62, f'Expected 62 semantic option rows, found {len(result)}'
    return result


def parse_run_of_river_hydro() -> pd.DataFrame:
    frame = read_source_range('Hydro Scheme Inflows', 'B81:T121')
    pieces = []
    for scheme, header, rows in [('Kareeya', 87, range(88, 103)), ('Barron Gorge', 105, range(106, 121))]:
        part = frame.loc[list(rows), [c for c in frame.columns if c <= 'O']].copy()
        part.columns = _combine_header_rows('Hydro Scheme Inflows', frame.loc[:, part.columns], [header])
        part.insert(0, 'Scheme', scheme)
        pieces.append(part)
    result = pd.concat(pieces, ignore_index=True)
    assert len(result) == 30
    return result


def parse_summary_mapping(cell_range: str) -> pd.DataFrame:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    source = read_source_range('Summary Mapping', cell_range)
    header = read_source_range('Summary Mapping', 'B4:AF6')
    columns = [get_column_letter(c) for c in range(min_col, max_col + 1)]
    data_start = 7 if min_row == 2 else min_row + 3
    result = source.loc[data_start:max_row, columns].dropna(axis=0, how='all').copy()
    result = result[result[columns[0]].notna()].copy()
    result.columns = _combine_header_rows('Summary Mapping', header.loc[:, columns], [4, 5, 6])
    result = result.reset_index(drop=True)
    assert result.columns[0] == 'RowID'
    return result


TABLE_RULES = {
    ('Existing Gen Data Summary', 'Existing generation data summary'): {'header_rows': [10, 11, 12], 'data_rows': range(13, 739)},
    ('New Entrant Data Summary', 'New entrant data summary'): {'header_rows': [9, 10, 11], 'data_rows': range(12, 536)},
    ('New Electrolyser Data Summary', 'New electrolyser data summary'): {'header_rows': [5, 6, 7], 'data_rows': range(8, 68)},
    ('Connections Forecasts', 'Consultant forecast mapping'): {'header_rows': [9], 'data_rows': [10]},
    ('Electrification', 'Consultant forecast mapping'): {'header_rows': [8], 'data_rows': [9]},
    ('Fuel cell EVs', 'Fuel-cell EV forecasts'): {'header_rows': [14]},
    ('Network representation', 'Network nodes'): {'header_rows': [5]},
    ('Network capability', 'Flow-path transfer capability'): {'header_rows': [6, 7], 'data_rows': range(8, 26)},
    ('Seasonal ratings', 'Seasonal generator ratings'): {'header_rows': [43, 44], 'data_rows': range(45, 771)},
    ('Transmission Reliability', 'Transmission unplanned outage rates'): {'header_rows': [7], 'data_rows': range(8, 14)},
    ('Marginal Loss Factors', 'Existing generator marginal loss factors'): {'header_rows': [12], 'data_rows': range(13, 749)},
    ('Coal Min Stable Level', 'Coal minimum stable level'): {'header_rows': [12, 13], 'data_rows': range(14, 64)},
    ('Financial parameters', 'Value of customer reliability'): {'header_rows': [47], 'data_rows': [48]},
    ('Other hydrogen assumptions', 'Electrolyser electricity consumption rate'): {'header_rows': [9], 'data_rows': [10, 11]},
    ('Fuel Price Summary', 'Gas price scenario selection'): {'data_rows': [8], 'column_names': ['Mapping field', 'Workbook selection', 'Consultant mapping field', 'Step Change', 'Slower Growth', 'Accelerated Transition']},
    ('Fuel Price Summary', 'Coal and biomass price scenario selection'): {'data_rows': [9], 'column_names': ['Mapping field', 'Workbook selection', 'Consultant mapping field', 'Step Change', 'Slower Growth', 'Accelerated Transition']},
    ('Gas System Properties', 'Gas reserves and resources summary'): {'header_rows': [128, 129], 'data_rows': [130]},
    ('Gas System Properties', 'Gas reserves and resources by basin'): {'header_rows': [133], 'data_rows': range(134, 145)},
    ('Hydrogen consumption locations', 'Regional hydrogen consumption allocation rule'): {'header_rows': [7], 'data_rows': range(8, 10)},
    ('Hydrogen consumption locations', 'Subregional hydrogen consumption allocation rule'): {'header_rows': [15], 'data_rows': range(16, 19)},
    ('Hydrogen consumption locations', 'Hydrogen consumption allocation'): {'header_rows': [24, 25], 'data_rows': range(26, 41)},
    ('Embedded energy storages', 'Forecast mapping'): {'header_rows': [8], 'data_rows': [9]},
    ('Aggregated energy storages', 'Forecast mapping'): {'header_rows': [8], 'data_rows': [9]},
}


def validate_table(frame: pd.DataFrame, spec: dict) -> None:
    if 'expected_rows' in spec:
        assert len(frame) == spec['expected_rows'], (spec['name'], len(frame), spec['expected_rows'])
    if 'expected_cols' in spec:
        assert len(frame.columns) == spec['expected_cols'], (spec['name'], len(frame.columns), spec['expected_cols'])


def parse_spec(sheet_name: str, spec: dict) -> pd.DataFrame:
    if spec.get('parser') == 'flow_path_options':
        frame = parse_flow_path_augmentation_options()
    elif spec.get('parser') == 'run_of_river_hydro':
        frame = parse_run_of_river_hydro()
    elif spec.get('parser') == 'summary_mapping':
        frame = parse_summary_mapping(spec['range'])
    else:
        rule = dict(TABLE_RULES.get((sheet_name, spec['name']), {}))
        if spec['name'] in {'Consultant forecast mapping', 'Forecast mapping'} and not rule:
            source = read_source_range(sheet_name, spec['range'])
            record_rows = [
                row for row in source.index
                if any('scenario' in str(value).lower() and 'name' in str(value).lower() for value in source.loc[row].dropna())
            ]
            if record_rows:
                record = record_rows[0]
                rule = {'header_rows': [record - 1], 'data_rows': [record]}
        frame = parse_semantic_range(sheet_name, spec['range'], **rule)
        first_column_names = {
            ('Connections Forecasts', 'Consultant forecast mapping'): 'Mapping field',
            ('Electrification', 'Consultant forecast mapping'): 'Mapping field',
            ('Financial parameters', 'Value of customer reliability'): 'Metric',
            ('Embedded energy storages', 'Forecast mapping'): 'Mapping field',
            ('Aggregated energy storages', 'Forecast mapping'): 'Mapping field',
        }
        first_name = first_column_names.get((sheet_name, spec['name']))
        if first_name and len(frame.columns):
            columns = list(frame.columns); columns[0] = first_name; frame.columns = columns
        if frame.empty:
            source = read_source_range(sheet_name, spec['range']).dropna(axis=1, how='all')
            rows = [row for row in source.index if source.loc[row].notna().any() and not _is_note_row(source.loc[row].tolist())]
            frame = source.loc[rows].reset_index(drop=True)
            base = ['Field', 'Assumption'] if len(frame.columns) == 2 else ['Field', 'Statement', 'Detail']
            frame.columns = _unique_headers(base[:len(frame.columns)] + [None] * max(0, len(frame.columns) - len(base)))
    validate_table(frame, spec)
    return frame


def show_table(frame: pd.DataFrame) -> None:
    if get_ipython() is None:
        print(f"{frame.shape[0]} rows × {frame.shape[1]} columns")
    else:
        display(frame)

