# %% [markdown]
# # 2026 ISP inputs and assumptions workbook — table discovery
#
# This notebook inspects the workbook structure and records candidate semantic tables.
# A worksheet is not assumed to contain one table, and tables are not assumed to start at `A1`.
# `ws.max_row` and `ws.max_column` are not authoritative table boundaries because workbook formatting
# can extend far beyond meaningful data. For example, the reported dimensions reach `XFC` and `XDX`
# on worksheets whose populated cells occupy far fewer columns.
#
# Discovery therefore uses populated OOXML cells, formulas, merged ranges, headings, and local table
# boundaries. The catalogue is coverage-oriented: it keeps complex multi-row headers intact and does
# not fill missing values merely because neighbouring rows contain values.

# %%
from __future__ import annotations

import os
import re
import warnings
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries

_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_NS_REL_DOC = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_NS_REL_PKG = 'http://schemas.openxmlformats.org/package/2006/relationships'
_CELL_RE = re.compile(r'([A-Z]+)(\d+)')
warnings.filterwarnings('ignore', message='Data Validation extension is not supported and will be removed', category=UserWarning, module='openpyxl.worksheet._reader')

def _repo_root() -> Path:
    try:
        here = Path(__file__).resolve().parent
        return here.parent
    except NameError:
        cwd = Path.cwd().resolve()
        for candidate in (cwd, *cwd.parents):
            if (candidate / 'Project.toml').is_file() and (candidate / 'scripts').is_dir():
                return candidate
        raise RuntimeError('Could not locate the ParseISP.jl repository root.')

_PROJECT = _repo_root()
_DEFAULT_WORKBOOK = _PROJECT / 'scrapped' / 'ISP_reference_files' / '2026-isp-inputs-and-assumptions-workbook.xlsm'
WORKBOOK_PATH = Path(os.environ.get('PARSEISP_2026_INPUTS_WORKBOOK', _DEFAULT_WORKBOOK)).expanduser().resolve()
if not WORKBOOK_PATH.is_file():
    raise FileNotFoundError(f'ISP 2026 inputs workbook not found: {WORKBOOK_PATH}')

SHEET_RANGES = {'Disclaimer': [],
 'Change Log': ['B2:E657'],
 'Assumptions Summary': ['B6:D35', 'B37:E119', 'B121:C147'],
 'Scenarios': ['B5:E29'],
 'Existing Gen Data Summary': ['B10:AT738'],
 'New Entrant Data Summary': ['B9:BB535'],
 'New Electrolyser Data Summary': ['B5:AQ67'],
 'Fuel Price Summary': ['B7:S9', 'B11:AK738', 'B743:AK1268'],
 'Regional Build Costs Summary': ['B7:C10', 'B12:AV75'],
 'Energy Policy Targets': ['C15:E30',
                           'C32:F62',
                           'C65:G93',
                           'C96:E110',
                           'C114:N154',
                           'C157:E159',
                           'C161:I185',
                           'C187:F193',
                           'C195:E197',
                           'C200:E202',
                           'C206:E209',
                           'C213:E220',
                           'C223:E235',
                           'C239:F258',
                           'C262:E267',
                           'C271:E289',
                           'D291:E295',
                           'C297:E310',
                           'C313:E327'],
 'Carbon Budgets': ['B5:E9', 'B15:G21', 'B25:D31'],
 'Economic Growth Forecasts': ['B5:E8', 'B12:AG19', 'B21:AG28', 'B30:AG37', 'B41:AG48', 'B50:AG57', 'B59:AG66'],
 'Demand and Energy Forecasts': [],
 'End use fuel consumption data': ['B6:AF15', 'B17:AF26', 'B28:AF37'],
 'Appliance Uptake Forecasts': ['B13:AG20', 'B22:AG29', 'B31:AG38'],
 'Elec. Retail Price Indices': ['B8:AG12'],
 'Connections Forecasts': ['B8:E11', 'B15:AH22', 'B24:AH31', 'B33:AH40'],
 'Energy Efficiency': ['B8:G10',
                       'B14:AG30',
                       'B32:AG48',
                       'B50:AG66',
                       'B68:AG84',
                       'B86:AG102',
                       'B106:AG122',
                       'B124:AG140',
                       'B142:AG158',
                       'B160:AG176',
                       'B178:AG194'],
 'Rooftop PV': ['B8:E10', 'B12:AH63', 'B65:AH116'],
 'PVNSG': ['B8:E10', 'B12:AH63', 'B65:AH116'],
 'ONSG': ['B8:AH55', 'B57:AG74'],
 'Battery & Plug-in EVs': ['B7:E9', 'B11:AH62', 'B64:AH115'],
 'Fuel cell EVs': ['B7:E9', 'B11:AH62'],
 'EV V2G': ['B8:E10', 'B12:AH62', 'B64:AH115'],
 'Data Centre Forecasts': ['B6:E8', 'B10:AF16', 'B18:AF24', 'B26:AF32'],
 'DSP': ['B7:AI84', 'B87:AI164'],
 'Electrification': ['B7:E10', 'B12:AF19', 'B21:AF28', 'B30:AF37'],
 'Embedded energy storages': ['B7:E9', 'B11:AH62', 'B65:AH116'],
 'Aggregated energy storages': ['B7:E9', 'B11:AH62', 'B65:AH116'],
 'Network representation': ['B2:E22', 'B24:D42', 'B44:D53', 'B55:D63', 'B65:D82'],
 'Renewable energy zones': ['B6:E53'],
 'Network capability': ['B8:K25', 'B34:K42', 'B51:N60', 'B75:V84', 'B89:E94', 'B99:D115', 'B122:C134', 'B139:C148'],
 'Network losses': ['B5:J28', 'B30:J34', 'B36:J88'],
 'Transmission Reliability': ['B7:E13'],
 'Distribution network': ['B11:G38', 'B40:H57', 'B59:AZ1433'],
 'Connection cost': ['B6:J61', 'B62:R73'],
 'Connection cost forecasts': ['B8:AJ144', 'B147:AJ388'],
 'Flow path augmentation options': ['B11:Q127'],
 'Flow path cost forecasts': ['B10:AI111', 'B115:AI216', 'B220:AI321'],
 'REZ augmentations options': ['B10:O37', 'B39:O77', 'B79:O96', 'B98:O110', 'B112:O137'],
 'REZ cost forecasts': ['B11:AJ117', 'B118:AJ224', 'B225:AJ331'],
 'Distribution cost forecasts': ['B5:AJ84'],
 'Maximum capacity': ['B9:J750', 'L9:O31'],
 'Hybrid site limits': ['B9:G67'],
 'Seasonal ratings': ['B9:E36', 'B42:AI770'],
 'Generator Reliability Settings': ['B9:M16', 'B21:M60', 'B62:H90'],
 'Maintenance': ['B5:D29', 'G5:I32'],
 'Retirement': ['B8:F738', 'H8:I50'],
 'Hydro Scheme Inflows': ['B4:T79', 'B81:T121', 'B123:T141', 'B143:T162'],
 'Capacity Factors ': ['B2:V214'],
 'Heat rates': ['B7:E740', 'H7:I31'],
 'Auxiliary': ['B5:E736', 'G5:H29'],
 'Storage properties': ['B2:J19', 'B21:E35', 'G21:J27', 'B38:C45'],
 'Emissions intensity': ['B4:E744', 'G4:H29'],
 'Build costs': ['B2:AJ77'],
 'Fixed OPEX': ['B5:E739', 'G5:I32'],
 'Variable OPEX': ['B5:E738', 'G5:H32'],
 'Marginal Loss Factors': ['B10:F748', 'I10:M536', 'O10:S161'],
 'Locational Cost Factors': ['B9:H80', 'B83:I132', 'B134:G158', 'B161:X227'],
 'Build limits - REZs': ['B2:Q62',
                         'B64:N119',
                         'B121:F132',
                         'B136:K265',
                         'B267:K317',
                         'B319:K335',
                         'B337:E356',
                         'B358:G368'],
 'Build limits - PHES': ['B2:W27'],
 'First-of-a-kind premium': ['B2:D11'],
 'Lead time and project life': ['B2:H35'],
 'Financial parameters': ['B2:F7', 'B10:F41', 'B43:G51', 'B54:C90'],
 'Affine Heat rates': ['B6:F192', 'H6:K29'],
 'Max Ramp Rates': ['B7:F191', 'H7:J30'],
 'Coal Min Stable Level': ['B2:G63'],
 'GPG Min Stable Level': ['B10:E150', 'G10:H35'],
 'Coal and Biomass price': ['B8:AG54', 'B57:AG61'],
 'Gas, Liquid fuel, H2 price': ['B7:AG129',
                                'B132:AG224',
                                'B228:AG249',
                                'B253:AG274',
                                'B278:AG302',
                                'B305:AG429',
                                'B433:AG438',
                                'B440:AG452'],
 'Gas System Properties': ['B7:F49', 'B51:G105', 'B108:H122', 'B124:F144', 'B146:E169', 'B171:E185'],
 'GPG emissions reduction - BioM': ['B2:AF12'],
 'Power System Security': ['B4:D49', 'B52:AE56', 'B58:G72', 'B74:G94'],
 'Reserves': ['B2:C14'],
 'Hydrogen demand - Domestic': ['B2:AH53'],
 'Hydrogen monthly profiles': ['B2:AG44'],
 'Hydrogen demand-Export&Commod': ['B2:AH52', 'B54:AH105', 'B107:AH156'],
 'Hydrogen consumption locations': ['B5:F40', 'B42:B44', 'B46:D57'],
 'Water for Hydrogen': ['B2:AH52'],
 'Desalination demand for H2': ['B2:AH52'],
 'H2 as fuel for GPG Limit': ['B2:AG21'],
 'Build Cost - Hydrogen pipeline': ['B2:AJ156'],
 'Other hydrogen assumptions': ['B2:C5', 'B7:AF11', 'B13:AF17', 'B19:AF23', 'B26:AF30', 'B32:C35'],
 'Summary Mapping': ['C2:AF733', 'C734:AF786', 'C790:AF1316', 'C1319:AF1381']}

# %% [markdown]
# ## Discovery helpers
#
# The OOXML scan counts only cells that contain a value, inline string, or formula. Style-only cells
# are excluded so formatted worksheet dimensions do not become table boundaries.

# %%
def _column_number(label: str) -> int:
    value = 0
    for char in label:
        value = value * 26 + ord(char) - 64
    return value

def _cell_position(ref: str) -> tuple[int, int]:
    match = _CELL_RE.fullmatch(ref)
    if match is None:
        raise ValueError(f'Invalid cell reference: {ref}')
    return int(match.group(2)), _column_number(match.group(1))

def scan_workbook_structure(path: Path) -> dict[str, dict]:
    with zipfile.ZipFile(path) as archive:
        workbook = ET.fromstring(archive.read('xl/workbook.xml'))
        rels = ET.fromstring(archive.read('xl/_rels/workbook.xml.rels'))
        rel_map = {rel.attrib['Id']: rel.attrib['Target'] for rel in rels.findall(f'{{{_NS_REL_PKG}}}Relationship')}
        sheets = workbook.find(f'{{{_NS}}}sheets')
        result = {}
        for sheet in sheets:
            name = sheet.attrib['name']
            rid = sheet.attrib[f'{{{_NS_REL_DOC}}}id']
            target = rel_map[rid].lstrip('/')
            xml_path = target if target.startswith('xl/') else 'xl/' + target
            rows, cols, formulas, merges = [], [], 0, []
            reported = None
            with archive.open(xml_path) as handle:
                for event, elem in ET.iterparse(handle, events=('start', 'end')):
                    if event == 'start' and elem.tag == f'{{{_NS}}}dimension':
                        reported = elem.attrib.get('ref')
                    elif event == 'end' and elem.tag == f'{{{_NS}}}mergeCell':
                        merges.append(elem.attrib['ref'])
                        elem.clear()
                    elif event == 'end' and elem.tag == f'{{{_NS}}}c':
                        ref = elem.attrib.get('r')
                        has_value = elem.find(f'{{{_NS}}}v') is not None or elem.find(f'{{{_NS}}}is') is not None
                        has_formula = elem.find(f'{{{_NS}}}f') is not None
                        if ref and (has_value or has_formula):
                            row, col = _cell_position(ref)
                            rows.append(row); cols.append(col); formulas += int(has_formula)
                        elem.clear()
            bbox = None
            if rows:
                bbox = f'{get_column_letter(min(cols))}{min(rows)}:{get_column_letter(max(cols))}{max(rows)}'
            result[name] = {'reported_dimension': reported, 'content_bbox': bbox, 'nonempty_cells': len(rows), 'formula_cells': formulas, 'merged_ranges': merges}
    return result

DISCOVERY = scan_workbook_structure(WORKBOOK_PATH)
workbook_formula = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False, keep_links=False)

_SHEET_CACHE: dict[str, pd.DataFrame] = {}

def _sheet_source_frame(sheet_name: str) -> pd.DataFrame:
    if sheet_name in _SHEET_CACHE:
        return _SHEET_CACHE[sheet_name]
    ranges = SHEET_RANGES[sheet_name]
    if not ranges:
        content_bbox = DISCOVERY[sheet_name]["content_bbox"]
        ranges = [content_bbox] if content_bbox else []
    if not ranges:
        return pd.DataFrame()
    bounds = [range_boundaries(cell_range) for cell_range in ranges]
    min_col = min(bound[0] for bound in bounds); min_row = min(bound[1] for bound in bounds)
    max_col = max(bound[2] for bound in bounds); max_row = max(bound[3] for bound in bounds)
    ws = workbook_formula[sheet_name]
    values = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    frame = pd.DataFrame(values, index=range(min_row, max_row + 1), columns=[get_column_letter(c) for c in range(min_col, max_col + 1)])
    _SHEET_CACHE[sheet_name] = frame
    return frame

def inspect_candidate(sheet_name: str, cell_range: str, *, max_rows: int = 12) -> pd.DataFrame:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    columns = [get_column_letter(c) for c in range(min_col, max_col + 1)]
    frame = _sheet_source_frame(sheet_name).loc[min_row:max_row, columns].copy()
    frame = frame.dropna(axis=0, how="all").dropna(axis=1, how="all")
    return frame.head(max_rows)

def show_sheet_discovery(sheet_name: str) -> None:
    info = DISCOVERY[sheet_name]
    print(f"reported dimension: {info['reported_dimension']}")
    print(f"populated-cell extent: {info['content_bbox']}")
    print(f"populated cells: {info['nonempty_cells']}; formula cells: {info['formula_cells']}; merged ranges: {len(info['merged_ranges'])}")

