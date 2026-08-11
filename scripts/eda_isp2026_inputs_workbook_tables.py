# %% [markdown]
# # 2026 ISP inputs and assumptions workbook — table discovery
#
# This notebook inspects the workbook structure and records candidate semantic tables.
# A worksheet is not assumed to contain one table, and tables are not assumed to start at `A1`.
# `ws.max_row` and `ws.max_column` are not authoritative table boundaries because workbook formatting
# can extend far beyond meaningful data. For example, the reported dimensions reach `XFC` and `XDX`
# on worksheets whose populated cells occupy far fewer columns.
#
# Discovery therefore uses populated OOXML cells, formulas, merged ranges, headings, and local table
# boundaries. The catalogue is coverage-oriented: it keeps complex multi-row headers intact and does
# not fill missing values merely because neighbouring rows contain values.

# %%
from __future__ import annotations

import os
import re
import warnings
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries

_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_NS_REL_DOC = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_NS_REL_PKG = 'http://schemas.openxmlformats.org/package/2006/relationships'
_CELL_RE = re.compile(r'([A-Z]+)(\d+)')
warnings.filterwarnings('ignore', message='Data Validation extension is not supported and will be removed', category=UserWarning, module='openpyxl.worksheet._reader')

def _repo_root() -> Path:
    try:
        here = Path(__file__).resolve().parent
        return here.parent
    except NameError:
        cwd = Path.cwd().resolve()
        for candidate in (cwd, *cwd.parents):
            if (candidate / 'Project.toml').is_file() and (candidate / 'scripts').is_dir():
                return candidate
        raise RuntimeError('Could not locate the ParseISP.jl repository root.')

_PROJECT = _repo_root()
_DEFAULT_WORKBOOK = _PROJECT / 'scrapped' / 'ISP_reference_files' / '2026-isp-inputs-and-assumptions-workbook.xlsm'
WORKBOOK_PATH = Path(os.environ.get('PARSEISP_2026_INPUTS_WORKBOOK', _DEFAULT_WORKBOOK)).expanduser().resolve()
if not WORKBOOK_PATH.is_file():
    raise FileNotFoundError(f'ISP 2026 inputs workbook not found: {WORKBOOK_PATH}')

SHEET_RANGES = {'Disclaimer': [],
 'Change Log': ['B2:E657'],
 'Assumptions Summary': ['B6:D35', 'B37:E119', 'B121:C147'],
 'Scenarios': ['B5:E29'],
 'Existing Gen Data Summary': ['B10:AT738'],
 'New Entrant Data Summary': ['B9:BB535'],
 'New Electrolyser Data Summary': ['B5:AQ67'],
 'Fuel Price Summary': ['B7:S9', 'B11:AK738', 'B743:AK1268'],
 'Regional Build Costs Summary': ['B7:C10', 'B12:AV75'],
 'Energy Policy Targets': ['C15:E30',
                           'C32:F62',
                           'C65:G93',
                           'C96:E110',
                           'C114:N154',
                           'C157:E159',
                           'C161:I185',
                           'C187:F193',
                           'C195:E197',
                           'C200:E202',
                           'C206:E209',
                           'C213:E220',
                           'C223:E235',
                           'C239:F258',
                           'C262:E267',
                           'C271:E289',
                           'D291:E295',
                           'C297:E310',
                           'C313:E327'],
 'Carbon Budgets': ['B5:E9', 'B15:G21', 'B25:D31'],
 'Economic Growth Forecasts': ['B5:E8', 'B12:AG19', 'B21:AG28', 'B30:AG37', 'B41:AG48', 'B50:AG57', 'B59:AG66'],
 'Demand and Energy Forecasts': [],
 'End use fuel consumption data': ['B6:AF15', 'B17:AF26', 'B28:AF37'],
 'Appliance Uptake Forecasts': ['B13:AG20', 'B22:AG29', 'B31:AG38'],
 'Elec. Retail Price Indices': ['B8:AG12'],
 'Connections Forecasts': ['B8:E11', 'B15:AH22', 'B24:AH31', 'B33:AH40'],
 'Energy Efficiency': ['B8:G10',
                       'B14:AG30',
                       'B32:AG48',
                       'B50:AG66',
                       'B68:AG84',
                       'B86:AG102',
                       'B106:AG122',
                       'B124:AG140',
                       'B142:AG158',
                       'B160:AG176',
                       'B178:AG194'],
 'Rooftop PV': ['B8:E10', 'B12:AH63', 'B65:AH116'],
 'PVNSG': ['B8:E10', 'B12:AH63', 'B65:AH116'],
 'ONSG': ['B8:AH55', 'B57:AG74'],
 'Battery & Plug-in EVs': ['B7:E9', 'B11:AH62', 'B64:AH115'],
 'Fuel cell EVs': ['B7:E9', 'B11:AH62'],
 'EV V2G': ['B8:E10', 'B12:AH62', 'B64:AH115'],
 'Data Centre Forecasts': ['B6:E8', 'B10:AF16', 'B18:AF24', 'B26:AF32'],
 'DSP': ['B7:AI84', 'B87:AI164'],
 'Electrification': ['B7:E10', 'B12:AF19', 'B21:AF28', 'B30:AF37'],
 'Embedded energy storages': ['B7:E9', 'B11:AH62', 'B65:AH116'],
 'Aggregated energy storages': ['B7:E9', 'B11:AH62', 'B65:AH116'],
 'Network representation': ['B2:E22', 'B24:D42', 'B44:D53', 'B55:D63', 'B65:D82'],
 'Renewable energy zones': ['B6:E53'],
 'Network capability': ['B8:K25', 'B34:K42', 'B51:N60', 'B75:V84', 'B89:E94', 'B99:D115', 'B122:C134', 'B139:C148'],
 'Network losses': ['B5:J28', 'B30:J34', 'B36:J88'],
 'Transmission Reliability': ['B7:E13'],
 'Distribution network': ['B11:G38', 'B40:H57', 'B59:AZ1433'],
 'Connection cost': ['B6:J61', 'B62:R73'],
 'Connection cost forecasts': ['B8:AJ144', 'B147:AJ388'],
 'Flow path augmentation options': ['B11:Q127'],
 'Flow path cost forecasts': ['B10:AI111', 'B115:AI216', 'B220:AI321'],
 'REZ augmentations options': ['B10:O37', 'B39:O77', 'B79:O96', 'B98:O110', 'B112:O137'],
 'REZ cost forecasts': ['B11:AJ117', 'B118:AJ224', 'B225:AJ331'],
 'Distribution cost forecasts': ['B5:AJ84'],
 'Maximum capacity': ['B9:J750', 'L9:O31'],
 'Hybrid site limits': ['B9:G67'],
 'Seasonal ratings': ['B9:E36', 'B42:AI770'],
 'Generator Reliability Settings': ['B9:M16', 'B21:M60', 'B62:H90'],
 'Maintenance': ['B5:D29', 'G5:I32'],
 'Retirement': ['B8:F738', 'H8:I50'],
 'Hydro Scheme Inflows': ['B4:T79', 'B81:T121', 'B123:T141', 'B143:T162'],
 'Capacity Factors ': ['B2:V214'],
 'Heat rates': ['B7:E740', 'H7:I31'],
 'Auxiliary': ['B5:E736', 'G5:H29'],
 'Storage properties': ['B2:J19', 'B21:E35', 'G21:J27', 'B38:C45'],
 'Emissions intensity': ['B4:E744', 'G4:H29'],
 'Build costs': ['B2:AJ77'],
 'Fixed OPEX': ['B5:E739', 'G5:I32'],
 'Variable OPEX': ['B5:E738', 'G5:H32'],
 'Marginal Loss Factors': ['B10:F748', 'I10:M536', 'O10:S161'],
 'Locational Cost Factors': ['B9:H80', 'B83:I132', 'B134:G158', 'B161:X227'],
 'Build limits - REZs': ['B2:Q62',
                         'B64:N119',
                         'B121:F132',
                         'B136:K265',
                         'B267:K317',
                         'B319:K335',
                         'B337:E356',
                         'B358:G368'],
 'Build limits - PHES': ['B2:W27'],
 'First-of-a-kind premium': ['B2:D11'],
 'Lead time and project life': ['B2:H35'],
 'Financial parameters': ['B2:F7', 'B10:F41', 'B43:G51', 'B54:C90'],
 'Affine Heat rates': ['B6:F192', 'H6:K29'],
 'Max Ramp Rates': ['B7:F191', 'H7:J30'],
 'Coal Min Stable Level': ['B2:G63'],
 'GPG Min Stable Level': ['B10:E150', 'G10:H35'],
 'Coal and Biomass price': ['B8:AG54', 'B57:AG61'],
 'Gas, Liquid fuel, H2 price': ['B7:AG129',
                                'B132:AG224',
                                'B228:AG249',
                                'B253:AG274',
                                'B278:AG302',
                                'B305:AG429',
                                'B433:AG438',
                                'B440:AG452'],
 'Gas System Properties': ['B7:F49', 'B51:G105', 'B108:H122', 'B124:F144', 'B146:E169', 'B171:E185'],
 'GPG emissions reduction - BioM': ['B2:AF12'],
 'Power System Security': ['B4:D49', 'B52:AE56', 'B58:G72', 'B74:G94'],
 'Reserves': ['B2:C14'],
 'Hydrogen demand - Domestic': ['B2:AH53'],
 'Hydrogen monthly profiles': ['B2:AG44'],
 'Hydrogen demand-Export&Commod': ['B2:AH52', 'B54:AH105', 'B107:AH156'],
 'Hydrogen consumption locations': ['B5:F40', 'B42:B44', 'B46:D57'],
 'Water for Hydrogen': ['B2:AH52'],
 'Desalination demand for H2': ['B2:AH52'],
 'H2 as fuel for GPG Limit': ['B2:AG21'],
 'Build Cost - Hydrogen pipeline': ['B2:AJ156'],
 'Other hydrogen assumptions': ['B2:C5', 'B7:AF11', 'B13:AF17', 'B19:AF23', 'B26:AF30', 'B32:C35'],
 'Summary Mapping': ['C2:AF733', 'C734:AF786', 'C790:AF1316', 'C1319:AF1381']}

# %% [markdown]
# ## Discovery helpers
#
# The OOXML scan counts only cells that contain a value, inline string, or formula. Style-only cells
# are excluded so formatted worksheet dimensions do not become table boundaries.

# %%
def _column_number(label: str) -> int:
    value = 0
    for char in label:
        value = value * 26 + ord(char) - 64
    return value

def _cell_position(ref: str) -> tuple[int, int]:
    match = _CELL_RE.fullmatch(ref)
    if match is None:
        raise ValueError(f'Invalid cell reference: {ref}')
    return int(match.group(2)), _column_number(match.group(1))

def scan_workbook_structure(path: Path) -> dict[str, dict]:
    with zipfile.ZipFile(path) as archive:
        workbook = ET.fromstring(archive.read('xl/workbook.xml'))
        rels = ET.fromstring(archive.read('xl/_rels/workbook.xml.rels'))
        rel_map = {rel.attrib['Id']: rel.attrib['Target'] for rel in rels.findall(f'{{{_NS_REL_PKG}}}Relationship')}
        sheets = workbook.find(f'{{{_NS}}}sheets')
        result = {}
        for sheet in sheets:
            name = sheet.attrib['name']
            rid = sheet.attrib[f'{{{_NS_REL_DOC}}}id']
            target = rel_map[rid].lstrip('/')
            xml_path = target if target.startswith('xl/') else 'xl/' + target
            rows, cols, formulas, merges = [], [], 0, []
            reported = None
            with archive.open(xml_path) as handle:
                for event, elem in ET.iterparse(handle, events=('start', 'end')):
                    if event == 'start' and elem.tag == f'{{{_NS}}}dimension':
                        reported = elem.attrib.get('ref')
                    elif event == 'end' and elem.tag == f'{{{_NS}}}mergeCell':
                        merges.append(elem.attrib['ref'])
                        elem.clear()
                    elif event == 'end' and elem.tag == f'{{{_NS}}}c':
                        ref = elem.attrib.get('r')
                        has_value = elem.find(f'{{{_NS}}}v') is not None or elem.find(f'{{{_NS}}}is') is not None
                        has_formula = elem.find(f'{{{_NS}}}f') is not None
                        if ref and (has_value or has_formula):
                            row, col = _cell_position(ref)
                            rows.append(row); cols.append(col); formulas += int(has_formula)
                        elem.clear()
            bbox = None
            if rows:
                bbox = f'{get_column_letter(min(cols))}{min(rows)}:{get_column_letter(max(cols))}{max(rows)}'
            result[name] = {'reported_dimension': reported, 'content_bbox': bbox, 'nonempty_cells': len(rows), 'formula_cells': formulas, 'merged_ranges': merges}
    return result

DISCOVERY = scan_workbook_structure(WORKBOOK_PATH)
workbook_formula = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False, keep_links=False)

_SHEET_CACHE: dict[str, pd.DataFrame] = {}

def _sheet_source_frame(sheet_name: str) -> pd.DataFrame:
    if sheet_name in _SHEET_CACHE:
        return _SHEET_CACHE[sheet_name]
    ranges = SHEET_RANGES[sheet_name]
    if not ranges:
        content_bbox = DISCOVERY[sheet_name]["content_bbox"]
        ranges = [content_bbox] if content_bbox else []
    if not ranges:
        return pd.DataFrame()
    bounds = [range_boundaries(cell_range) for cell_range in ranges]
    min_col = min(bound[0] for bound in bounds); min_row = min(bound[1] for bound in bounds)
    max_col = max(bound[2] for bound in bounds); max_row = max(bound[3] for bound in bounds)
    ws = workbook_formula[sheet_name]
    values = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    frame = pd.DataFrame(values, index=range(min_row, max_row + 1), columns=[get_column_letter(c) for c in range(min_col, max_col + 1)])
    _SHEET_CACHE[sheet_name] = frame
    return frame

def inspect_candidate(sheet_name: str, cell_range: str, *, max_rows: int = 12) -> pd.DataFrame:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    columns = [get_column_letter(c) for c in range(min_col, max_col + 1)]
    frame = _sheet_source_frame(sheet_name).loc[min_row:max_row, columns].copy()
    frame = frame.dropna(axis=0, how="all").dropna(axis=1, how="all")
    return frame.head(max_rows)

def show_sheet_discovery(sheet_name: str) -> None:
    info = DISCOVERY[sheet_name]
    print(f"reported dimension: {info['reported_dimension']}")
    print(f"populated-cell extent: {info['content_bbox']}")
    print(f"populated cells: {info['nonempty_cells']}; formula cells: {info['formula_cells']}; merged ranges: {len(info['merged_ranges'])}")

# %% [markdown]
# ## Disclaimer
#
# States the workbook purpose, limitations, and conditions of use.

# %%
show_sheet_discovery('Disclaimer')

# %% [markdown]
#
# No embedded semantic data table was identified. The populated cells are notice, purpose, and disclaimer text.

# %% [markdown]
# ## Change Log
#
# Records changes made across workbook releases.

# %%
show_sheet_discovery('Change Log')

# %% [markdown]
# ### Workbook change log
#
# Records workbook versions, dates, changes, and supporting detail.
#
# Candidate source block: `B2:E657` (656 rows × 4 columns).

# %%
inspect_candidate('Change Log', 'B2:E657')

# %% [markdown]
# ## Assumptions Summary
#
# Provides workbook metadata, worksheet descriptions, and supporting materials.

# %%
show_sheet_discovery('Assumptions Summary')

# %% [markdown]
# ### Version history
#
# Records workbook version numbers, dates, and descriptions.
#
# Candidate source block: `B6:D35` (30 rows × 3 columns).

# %%
inspect_candidate('Assumptions Summary', 'B6:D35')

# %% [markdown]
# ### Worksheet descriptions
#
# Maps worksheets to assumption groups, descriptions, and sources.
#
# Candidate source block: `B37:E119` (83 rows × 4 columns).

# %%
inspect_candidate('Assumptions Summary', 'B37:E119')

# %% [markdown]
# ### Supporting materials
#
# Lists supporting source materials referenced by the workbook.
#
# Candidate source block: `B121:C147` (27 rows × 2 columns).

# %%
inspect_candidate('Assumptions Summary', 'B121:C147')

# %% [markdown]
# ## Scenarios
#
# Summary of scenario dimensions and parameters.

# %%
show_sheet_discovery('Scenarios')

# %% [markdown]
# ### Scenario parameters
#
# Compares the main parameter settings across the three ISP scenarios.
#
# Candidate source block: `B5:E29` (25 rows × 4 columns).

# %%
inspect_candidate('Scenarios', 'B5:E29')

# %% [markdown]
# ## Existing Gen Data Summary
#
# Summary (calculated) of the generator technical data.

# %%
show_sheet_discovery('Existing Gen Data Summary')

# %% [markdown]
# ### Existing generation data summary
#
# Summarises technical data for existing, committed, anticipated, and additional generators.
#
# Candidate source block: `B10:AT738` (729 rows × 45 columns).

# %%
inspect_candidate('Existing Gen Data Summary', 'B10:AT738')

# %% [markdown]
# ## New Entrant Data Summary
#
# Summary (calculated) of the new entrant generation and storage technical data.

# %%
show_sheet_discovery('New Entrant Data Summary')

# %% [markdown]
# ### New entrant data summary
#
# Summarises technical and cost data for new entrant generation and storage technologies.
#
# Candidate source block: `B9:BB535` (527 rows × 53 columns).

# %%
inspect_candidate('New Entrant Data Summary', 'B9:BB535')

# %% [markdown]
# ## New Electrolyser Data Summary
#
# Summary (calculated) of the new entrant electrolyser technical data.

# %%
show_sheet_discovery('New Electrolyser Data Summary')

# %% [markdown]
# ### New electrolyser data summary
#
# Summarises technical and cost data for new entrant electrolysers.
#
# Candidate source block: `B5:AQ67` (63 rows × 42 columns).

# %%
inspect_candidate('New Electrolyser Data Summary', 'B5:AQ67')

# %% [markdown]
# ## Fuel Price Summary
#
# Summary (calculated) of generator fuel costs.

# %%
show_sheet_discovery('Fuel Price Summary')

# %% [markdown]
# ### Fuel-price scenario selection
#
# Maps the selected ISP scenario to the fuel-price summary calculations.
#
# Candidate source block: `B7:S9` (3 rows × 18 columns).

# %%
inspect_candidate('Fuel Price Summary', 'B7:S9')

# %% [markdown]
# ### Existing generator fuel prices
#
# Summarises fuel prices for existing, committed, anticipated, and additional generators.
#
# Candidate source block: `B11:AK738` (728 rows × 36 columns).

# %%
inspect_candidate('Fuel Price Summary', 'B11:AK738')

# %% [markdown]
# ### New entrant fuel prices
#
# Summarises fuel prices for new entrant generation technologies.
#
# Candidate source block: `B743:AK1268` (526 rows × 36 columns).

# %%
inspect_candidate('Fuel Price Summary', 'B743:AK1268')

# %% [markdown]
# ## Regional Build Costs Summary
#
# Summary (calculated) of regional build costs for a selectable scenario / technology.

# %%
show_sheet_discovery('Regional Build Costs Summary')

# %% [markdown]
# ### Build-cost selection
#
# Records the scenario and technology controls used by the regional build-cost summary.
#
# Candidate source block: `B7:C10` (4 rows × 2 columns).

# %%
inspect_candidate('Regional Build Costs Summary', 'B7:C10')

# %% [markdown]
# ### Regional build costs
#
# Summarises regional build costs after locational cost factors are applied.
#
# Candidate source block: `B12:AV75` (64 rows × 47 columns).

# %%
inspect_candidate('Regional Build Costs Summary', 'B12:AV75')

# %% [markdown]
# ## Energy Policy Targets
#
# Target renewable settings for NEM-wide, Queensland, New South Wales, Victoria, South Australia and Tasmania energy policy targets.

# %%
show_sheet_discovery('Energy Policy Targets')

# %% [markdown]
# ### Powering Australia Plan 2030 target
#
# Candidate source block: `C15:E30` (16 rows × 3 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C15:E30')

# %% [markdown]
# ### Capacity Investment Scheme generation target
#
# Candidate source block: `C32:F62` (31 rows × 4 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C32:F62')

# %% [markdown]
# ### Capacity Investment Scheme clean dispatchable capacity target
#
# Candidate source block: `C65:G93` (29 rows × 5 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C65:G93')

# %% [markdown]
# ### Large-scale Renewable Energy Target
#
# Candidate source block: `C96:E110` (15 rows × 3 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C96:E110')

# %% [markdown]
# ### NSW Electricity Infrastructure Roadmap
#
# Candidate source block: `C114:N154` (41 rows × 12 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C114:N154')

# %% [markdown]
# ### Long-term energy services agreements
#
# Candidate source block: `C157:E159` (3 rows × 3 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C157:E159')

# %% [markdown]
# ### REZ Access Scheme
#
# Candidate source block: `C161:I185` (25 rows × 7 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C161:I185')

# %% [markdown]
# ### NSW Roadmap Tender 7 firming
#
# Candidate source block: `C187:F193` (7 rows × 4 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C187:F193')

# %% [markdown]
# ### NSW Renewable Fuels Scheme
#
# Candidate source block: `C195:E197` (3 rows × 3 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C195:E197')

# %% [markdown]
# ### NSW electricity landholder payment scheme
#
# Candidate source block: `C200:E202` (3 rows × 3 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C200:E202')

# %% [markdown]
# ### Queensland landholder payment scheme
#
# Candidate source block: `C206:E209` (4 rows × 3 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C206:E209')

# %% [markdown]
# ### South Australia net renewable energy generation target
#
# Candidate source block: `C213:E220` (8 rows × 3 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C213:E220')

# %% [markdown]
# ### Firm Energy Reliability Mechanism
#
# Candidate source block: `C223:E235` (13 rows × 3 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C223:E235')

# %% [markdown]
# ### Tasmania Renewable Energy Target
#
# Candidate source block: `C239:F258` (20 rows × 4 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C239:F258')

# %% [markdown]
# ### Tasmanian landholder payment scheme
#
# Candidate source block: `C262:E267` (6 rows × 3 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C262:E267')

# %% [markdown]
# ### Victorian Renewable Energy Target
#
# Candidate source block: `C271:E289` (19 rows × 3 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C271:E289')

# %% [markdown]
# ### VRET auctions
#
# Candidate source block: `D291:E295` (5 rows × 2 columns).

# %%
inspect_candidate('Energy Policy Targets', 'D291:E295')

# %% [markdown]
# ### Victorian Energy Storage Target
#
# Candidate source block: `C297:E310` (14 rows × 3 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C297:E310')

# %% [markdown]
# ### Victorian Offshore Wind Target
#
# Candidate source block: `C313:E327` (15 rows × 3 columns).

# %%
inspect_candidate('Energy Policy Targets', 'C313:E327')

# %% [markdown]
# ## Carbon Budgets
#
# Global mean temperature increases by 2100 aligned with each scenario, and cumulative carbon budgets over the period to 2050.

# %%
show_sheet_discovery('Carbon Budgets')

# %% [markdown]
# ### NEM-wide carbon budgets
#
# Sets cumulative NEM-wide carbon budgets by scenario.
#
# Candidate source block: `B5:E9` (5 rows × 4 columns).

# %%
inspect_candidate('Carbon Budgets', 'B5:E9')

# %% [markdown]
# ### State carbon targets
#
# Records jurisdictional carbon targets used by the workbook.
#
# Candidate source block: `B15:G21` (7 rows × 6 columns).

# %%
inspect_candidate('Carbon Budgets', 'B15:G21')

# %% [markdown]
# ### Converted state carbon budgets
#
# Expresses state carbon targets in the workbook carbon-budget form.
#
# Candidate source block: `B25:D31` (7 rows × 3 columns).

# %%
inspect_candidate('Carbon Budgets', 'B25:D31')

# %% [markdown]
# ## Economic Growth Forecasts
#
# Forecasts of Gross State Product (GSP) and Household Disposable Income (HDI).

# %%
show_sheet_discovery('Economic Growth Forecasts')

# %% [markdown]
# ### Consultant forecast mapping
#
# Maps ISP scenarios to the consultant economic-growth scenarios.
#
# Candidate source block: `B5:E8` (4 rows × 4 columns).

# %%
inspect_candidate('Economic Growth Forecasts', 'B5:E8')

# %% [markdown]
# ### Gross State Product — Slower Growth
#
# Candidate source block: `B12:AG19` (8 rows × 32 columns).

# %%
inspect_candidate('Economic Growth Forecasts', 'B12:AG19')

# %% [markdown]
# ### Gross State Product — Step Change
#
# Candidate source block: `B21:AG28` (8 rows × 32 columns).

# %%
inspect_candidate('Economic Growth Forecasts', 'B21:AG28')

# %% [markdown]
# ### Gross State Product — Accelerated Transition
#
# Candidate source block: `B30:AG37` (8 rows × 32 columns).

# %%
inspect_candidate('Economic Growth Forecasts', 'B30:AG37')

# %% [markdown]
# ### Household Disposable Income — Slower Growth
#
# Candidate source block: `B41:AG48` (8 rows × 32 columns).

# %%
inspect_candidate('Economic Growth Forecasts', 'B41:AG48')

# %% [markdown]
# ### Household Disposable Income — Step Change
#
# Candidate source block: `B50:AG57` (8 rows × 32 columns).

# %%
inspect_candidate('Economic Growth Forecasts', 'B50:AG57')

# %% [markdown]
# ### Household Disposable Income — Accelerated Transition
#
# Candidate source block: `B59:AG66` (8 rows × 32 columns).

# %%
inspect_candidate('Economic Growth Forecasts', 'B59:AG66')

# %% [markdown]
# ## Demand and Energy Forecasts
#
# Points readers to AEMO demand and energy forecasts; it does not embed a forecast data table.

# %%
show_sheet_discovery('Demand and Energy Forecasts')

# %% [markdown]
#
# No embedded semantic data table was identified. The populated cells are explanatory text and forecast-portal links.

# %% [markdown]
# ## End use fuel consumption data
#
# Data for end-use fuel consumption by scenario across the NEM chart, identified by multi-sectoral modelling conducted by CSIRO (Figure 1 in IASR).

# %%
show_sheet_discovery('End use fuel consumption data')

# %% [markdown]
# ### End-use fuel consumption — Slower Growth
#
# Candidate source block: `B6:AF15` (10 rows × 31 columns).

# %%
inspect_candidate('End use fuel consumption data', 'B6:AF15')

# %% [markdown]
# ### End-use fuel consumption — Step Change
#
# Candidate source block: `B17:AF26` (10 rows × 31 columns).

# %%
inspect_candidate('End use fuel consumption data', 'B17:AF26')

# %% [markdown]
# ### End-use fuel consumption — Accelerated Transition
#
# Candidate source block: `B28:AF37` (10 rows × 31 columns).

# %%
inspect_candidate('End use fuel consumption data', 'B28:AF37')

# %% [markdown]
# ## Appliance Uptake Forecasts
#
# Residential appliance uptake forecasts, impacts relative to base year.

# %%
show_sheet_discovery('Appliance Uptake Forecasts')

# %% [markdown]
# ### Residential appliance uptake — Slower Growth
#
# Candidate source block: `B13:AG20` (8 rows × 32 columns).

# %%
inspect_candidate('Appliance Uptake Forecasts', 'B13:AG20')

# %% [markdown]
# ### Residential appliance uptake — Step Change
#
# Candidate source block: `B22:AG29` (8 rows × 32 columns).

# %%
inspect_candidate('Appliance Uptake Forecasts', 'B22:AG29')

# %% [markdown]
# ### Residential appliance uptake — Accelerated Transition
#
# Candidate source block: `B31:AG38` (8 rows × 32 columns).

# %%
inspect_candidate('Appliance Uptake Forecasts', 'B31:AG38')

# %% [markdown]
# ## Elec. Retail Price Indices
#
# Provides residential electricity retail price indices relative to the base year.

# %%
show_sheet_discovery('Elec. Retail Price Indices')

# %% [markdown]
# ### NEM residential electricity price index
#
# Provides residential retail price indices by ISP scenario.
#
# Candidate source block: `B8:AG12` (5 rows × 32 columns).

# %%
inspect_candidate('Elec. Retail Price Indices', 'B8:AG12')

# %% [markdown]
# ## Connections Forecasts
#
# Residential connections forecasts.

# %%
show_sheet_discovery('Connections Forecasts')

# %% [markdown]
# ### Consultant forecast mapping
#
# Candidate source block: `B8:E11` (4 rows × 4 columns).

# %%
inspect_candidate('Connections Forecasts', 'B8:E11')

# %% [markdown]
# ### Residential connections — Slower Growth
#
# Candidate source block: `B15:AH22` (8 rows × 33 columns).

# %%
inspect_candidate('Connections Forecasts', 'B15:AH22')

# %% [markdown]
# ### Residential connections — Step Change
#
# Candidate source block: `B24:AH31` (8 rows × 33 columns).

# %%
inspect_candidate('Connections Forecasts', 'B24:AH31')

# %% [markdown]
# ### Residential connections — Accelerated Transition
#
# Candidate source block: `B33:AH40` (8 rows × 33 columns).

# %%
inspect_candidate('Connections Forecasts', 'B33:AH40')

# %% [markdown]
# ## Energy Efficiency
#
# Forecast of energy efficiency impact on annual consumption forecasts.

# %%
show_sheet_discovery('Energy Efficiency')

# %% [markdown]
# ### Consultant forecast mapping
#
# Candidate source block: `B8:G10` (3 rows × 6 columns).

# %%
inspect_candidate('Energy Efficiency', 'B8:G10')

# %% [markdown]
# ### Residential energy efficiency — Slower Growth
#
# Candidate source block: `B14:AG30` (17 rows × 32 columns).

# %%
inspect_candidate('Energy Efficiency', 'B14:AG30')

# %% [markdown]
# ### Residential energy efficiency — Step Change
#
# Candidate source block: `B32:AG48` (17 rows × 32 columns).

# %%
inspect_candidate('Energy Efficiency', 'B32:AG48')

# %% [markdown]
# ### Residential energy efficiency — Accelerated Transition
#
# Candidate source block: `B50:AG66` (17 rows × 32 columns).

# %%
inspect_candidate('Energy Efficiency', 'B50:AG66')

# %% [markdown]
# ### Residential energy efficiency — lower sensitivity
#
# Candidate source block: `B68:AG84` (17 rows × 32 columns).

# %%
inspect_candidate('Energy Efficiency', 'B68:AG84')

# %% [markdown]
# ### Residential energy efficiency — higher sensitivity
#
# Candidate source block: `B86:AG102` (17 rows × 32 columns).

# %%
inspect_candidate('Energy Efficiency', 'B86:AG102')

# %% [markdown]
# ### Business energy efficiency — Slower Growth
#
# Candidate source block: `B106:AG122` (17 rows × 32 columns).

# %%
inspect_candidate('Energy Efficiency', 'B106:AG122')

# %% [markdown]
# ### Business energy efficiency — Step Change
#
# Candidate source block: `B124:AG140` (17 rows × 32 columns).

# %%
inspect_candidate('Energy Efficiency', 'B124:AG140')

# %% [markdown]
# ### Business energy efficiency — Accelerated Transition
#
# Candidate source block: `B142:AG158` (17 rows × 32 columns).

# %%
inspect_candidate('Energy Efficiency', 'B142:AG158')

# %% [markdown]
# ### Business energy efficiency — lower sensitivity
#
# Candidate source block: `B160:AG176` (17 rows × 32 columns).

# %%
inspect_candidate('Energy Efficiency', 'B160:AG176')

# %% [markdown]
# ### Business energy efficiency — higher sensitivity
#
# Candidate source block: `B178:AG194` (17 rows × 32 columns).

# %%
inspect_candidate('Energy Efficiency', 'B178:AG194')

# %% [markdown]
# ## Rooftop PV
#
# Rooftop PV capacity and generation forecast.

# %%
show_sheet_discovery('Rooftop PV')

# %% [markdown]
# ### Consultant forecast mapping
#
# Candidate source block: `B8:E10` (3 rows × 4 columns).

# %%
inspect_candidate('Rooftop PV', 'B8:E10')

# %% [markdown]
# ### Rooftop PV capacity
#
# Candidate source block: `B12:AH63` (52 rows × 33 columns).

# %%
inspect_candidate('Rooftop PV', 'B12:AH63')

# %% [markdown]
# ### Rooftop PV energy
#
# Candidate source block: `B65:AH116` (52 rows × 33 columns).

# %%
inspect_candidate('Rooftop PV', 'B65:AH116')

# %% [markdown]
# ## PVNSG
#
# PV non-scheduled generation (PVNSG) capacity and generation forecast.

# %%
show_sheet_discovery('PVNSG')

# %% [markdown]
# ### Consultant forecast mapping
#
# Candidate source block: `B8:E10` (3 rows × 4 columns).

# %%
inspect_candidate('PVNSG', 'B8:E10')

# %% [markdown]
# ### PVNSG capacity
#
# Candidate source block: `B12:AH63` (52 rows × 33 columns).

# %%
inspect_candidate('PVNSG', 'B12:AH63')

# %% [markdown]
# ### PVNSG energy
#
# Candidate source block: `B65:AH116` (52 rows × 33 columns).

# %%
inspect_candidate('PVNSG', 'B65:AH116')

# %% [markdown]
# ## ONSG
#
# Other non-scheduled generation (ONSG) capacity forecast.

# %%
show_sheet_discovery('ONSG')

# %% [markdown]
# ### Sub-regional ONSG capacity
#
# Candidate source block: `B8:AH55` (48 rows × 33 columns).

# %%
inspect_candidate('ONSG', 'B8:AH55')

# %% [markdown]
# ### Regional ONSG capacity
#
# Candidate source block: `B57:AG74` (18 rows × 32 columns).

# %%
inspect_candidate('ONSG', 'B57:AG74')

# %% [markdown]
# ## Battery & Plug-in EVs
#
# Battery and plug-in electric vehicles uptake and energy consumption for driving purposes.

# %%
show_sheet_discovery('Battery & Plug-in EVs')

# %% [markdown]
# ### Consultant forecast mapping
#
# Candidate source block: `B7:E9` (3 rows × 4 columns).

# %%
inspect_candidate('Battery & Plug-in EVs', 'B7:E9')

# %% [markdown]
# ### BEV and PHEV energy
#
# Candidate source block: `B11:AH62` (52 rows × 33 columns).

# %%
inspect_candidate('Battery & Plug-in EVs', 'B11:AH62')

# %% [markdown]
# ### BEV and PHEV uptake
#
# Candidate source block: `B64:AH115` (52 rows × 33 columns).

# %%
inspect_candidate('Battery & Plug-in EVs', 'B64:AH115')

# %% [markdown]
# ## Fuel cell EVs
#
# Fuel cell electric vehicles uptake.

# %%
show_sheet_discovery('Fuel cell EVs')

# %% [markdown]
# ### Consultant forecast mapping
#
# Candidate source block: `B7:E9` (3 rows × 4 columns).

# %%
inspect_candidate('Fuel cell EVs', 'B7:E9')

# %% [markdown]
# ### Fuel-cell EV uptake
#
# Candidate source block: `B11:AH62` (52 rows × 33 columns).

# %%
inspect_candidate('Fuel cell EVs', 'B11:AH62')

# %% [markdown]
# ## EV V2G
#
# Vehicle to Grid battery characteristics.

# %%
show_sheet_discovery('EV V2G')

# %% [markdown]
# ### Consultant forecast mapping
#
# Candidate source block: `B8:E10` (3 rows × 4 columns).

# %%
inspect_candidate('EV V2G', 'B8:E10')

# %% [markdown]
# ### Vehicle-to-grid capacity
#
# Candidate source block: `B12:AH62` (51 rows × 33 columns).

# %%
inspect_candidate('EV V2G', 'B12:AH62')

# %% [markdown]
# ### Vehicle-to-grid depth
#
# Candidate source block: `B64:AH115` (52 rows × 33 columns).

# %%
inspect_candidate('EV V2G', 'B64:AH115')

# %% [markdown]
# ## Data Centre Forecasts
#
# Forecast of electricity consumption from data centre growth.

# %%
show_sheet_discovery('Data Centre Forecasts')

# %% [markdown]
# ### Consultant forecast mapping
#
# Candidate source block: `B6:E8` (3 rows × 4 columns).

# %%
inspect_candidate('Data Centre Forecasts', 'B6:E8')

# %% [markdown]
# ### Data-centre demand — Slower Growth
#
# Candidate source block: `B10:AF16` (7 rows × 31 columns).

# %%
inspect_candidate('Data Centre Forecasts', 'B10:AF16')

# %% [markdown]
# ### Data-centre demand — Step Change
#
# Candidate source block: `B18:AF24` (7 rows × 31 columns).

# %%
inspect_candidate('Data Centre Forecasts', 'B18:AF24')

# %% [markdown]
# ### Data-centre demand — Accelerated Transition
#
# Candidate source block: `B26:AF32` (7 rows × 31 columns).

# %%
inspect_candidate('Data Centre Forecasts', 'B26:AF32')

# %% [markdown]
# ## DSP
#
# Demand side participation forecast.

# %%
show_sheet_discovery('DSP')

# %% [markdown]
# ### Summer demand-side participation
#
# Candidate source block: `B7:AI84` (78 rows × 34 columns).

# %%
inspect_candidate('DSP', 'B7:AI84')

# %% [markdown]
# ### Winter demand-side participation
#
# Candidate source block: `B87:AI164` (78 rows × 34 columns).

# %%
inspect_candidate('DSP', 'B87:AI164')

# %% [markdown]
# ## Electrification
#
# Electrification in all sectors excluding road transportation.

# %%
show_sheet_discovery('Electrification')

# %% [markdown]
# ### Consultant forecast mapping
#
# Candidate source block: `B7:E10` (4 rows × 4 columns).

# %%
inspect_candidate('Electrification', 'B7:E10')

# %% [markdown]
# ### Electrification — Slower Growth
#
# Candidate source block: `B12:AF19` (8 rows × 31 columns).

# %%
inspect_candidate('Electrification', 'B12:AF19')

# %% [markdown]
# ### Electrification — Step Change
#
# Candidate source block: `B21:AF28` (8 rows × 31 columns).

# %%
inspect_candidate('Electrification', 'B21:AF28')

# %% [markdown]
# ### Electrification — Accelerated Transition
#
# Candidate source block: `B30:AF37` (8 rows × 31 columns).

# %%
inspect_candidate('Electrification', 'B30:AF37')

# %% [markdown]
# ## Embedded energy storages
#
# Embedded consumer energy storage (battery) forecast.

# %%
show_sheet_discovery('Embedded energy storages')

# %% [markdown]
# ### Forecast mapping
#
# Candidate source block: `B7:E9` (3 rows × 4 columns).

# %%
inspect_candidate('Embedded energy storages', 'B7:E9')

# %% [markdown]
# ### Embedded energy storage capacity
#
# Candidate source block: `B11:AH62` (52 rows × 33 columns).

# %%
inspect_candidate('Embedded energy storages', 'B11:AH62')

# %% [markdown]
# ### Embedded energy storage degraded energy
#
# Candidate source block: `B65:AH116` (52 rows × 33 columns).

# %%
inspect_candidate('Embedded energy storages', 'B65:AH116')

# %% [markdown]
# ## Aggregated energy storages
#
# The 'aggregated' share of embedded energy storages that is modelled like a Virtual Power Plant (VPP).

# %%
show_sheet_discovery('Aggregated energy storages')

# %% [markdown]
# ### Forecast mapping
#
# Candidate source block: `B7:E9` (3 rows × 4 columns).

# %%
inspect_candidate('Aggregated energy storages', 'B7:E9')

# %% [markdown]
# ### Aggregated energy storage capacity
#
# Candidate source block: `B11:AH62` (52 rows × 33 columns).

# %%
inspect_candidate('Aggregated energy storages', 'B11:AH62')

# %% [markdown]
# ### Aggregated energy storage degraded energy
#
# Candidate source block: `B65:AH116` (52 rows × 33 columns).

# %%
inspect_candidate('Aggregated energy storages', 'B65:AH116')

# %% [markdown]
# ## Network representation
#
# Description of how the network is modelled in the capacity expansion models.

# %%
show_sheet_discovery('Network representation')

# %% [markdown]
# ### Sub-regional flow-path representation
#
# Candidate source block: `B2:E22` (21 rows × 4 columns).

# %%
inspect_candidate('Network representation', 'B2:E22')

# %% [markdown]
# ### Sub-regional reference nodes
#
# Candidate source block: `B24:D42` (19 rows × 3 columns).

# %%
inspect_candidate('Network representation', 'B24:D42')

# %% [markdown]
# ### Regional topology representation
#
# Candidate source block: `B44:D53` (10 rows × 3 columns).

# %%
inspect_candidate('Network representation', 'B44:D53')

# %% [markdown]
# ### Regional reference nodes
#
# Candidate source block: `B55:D63` (9 rows × 3 columns).

# %%
inspect_candidate('Network representation', 'B55:D63')

# %% [markdown]
# ### Sub-regional load and generation representation
#
# Candidate source block: `B65:D82` (18 rows × 3 columns).

# %%
inspect_candidate('Network representation', 'B65:D82')

# %% [markdown]
# ## Renewable energy zones
#
# Renewable energy zones.

# %%
show_sheet_discovery('Renewable energy zones')

# %% [markdown]
# ### Candidate renewable energy zones
#
# Lists candidate REZ identifiers, names, NEM regions, and ISP sub-regions.
#
# Candidate source block: `B6:E53` (48 rows × 4 columns).

# %%
inspect_candidate('Renewable energy zones', 'B6:E53')

# %% [markdown]
# ## Network capability
#
# Maximum forward and reverse flow path capability for capacity expansion modelling.

# %%
show_sheet_discovery('Network capability')

# %% [markdown]
# ### Flow-path transfer capability
#
# Contains the 18 verified flow-path capability data rows; workbook headers are in rows 6–7.
#
# Candidate source block: `B8:K25` (18 rows × 10 columns).

# %%
inspect_candidate('Network capability', 'B8:K25')

# %% [markdown]
# ### Interconnector transfer capability
#
# Candidate source block: `B34:K42` (9 rows × 10 columns).

# %%
inspect_candidate('Network capability', 'B34:K42')

# %% [markdown]
# ### Committed-project transfer capability uplift
#
# Candidate source block: `B51:N60` (10 rows × 13 columns).

# %%
inspect_candidate('Network capability', 'B51:N60')

# %% [markdown]
# ### Sydney Ring generator coefficients
#
# Candidate source block: `B75:V84` (10 rows × 21 columns).

# %%
inspect_candidate('Network capability', 'B75:V84')

# %% [markdown]
# ### Reference temperatures
#
# Candidate source block: `B89:E94` (6 rows × 4 columns).

# %%
inspect_candidate('Network capability', 'B89:E94')

# %% [markdown]
# ### Murraylink dynamic temperature-dependent transfer capability
#
# Candidate source block: `B99:D115` (17 rows × 3 columns).

# %%
inspect_candidate('Network capability', 'B99:D115')

# %% [markdown]
# ### Basslink static daily energy throughput limit
#
# Candidate source block: `B122:C134` (13 rows × 2 columns).

# %%
inspect_candidate('Network capability', 'B122:C134')

# %% [markdown]
# ### Committed and anticipated project timing
#
# Candidate source block: `B139:C148` (10 rows × 2 columns).

# %%
inspect_candidate('Network capability', 'B139:C148')

# %% [markdown]
# ## Network losses
#
# Proportion of interconnector losses applied to regional reference nodes and loss equations.

# %%
show_sheet_discovery('Network losses')

# %% [markdown]
# ### Existing flow-path loss equations
#
# Candidate source block: `B5:J28` (24 rows × 9 columns).

# %%
inspect_candidate('Network losses', 'B5:J28')

# %% [markdown]
# ### Committed and anticipated project loss equations
#
# Candidate source block: `B30:J34` (5 rows × 9 columns).

# %%
inspect_candidate('Network losses', 'B30:J34')

# %% [markdown]
# ### Development-option loss equations
#
# Candidate source block: `B36:J88` (53 rows × 9 columns).

# %%
inspect_candidate('Network losses', 'B36:J88')

# %% [markdown]
# ## Transmission Reliability
#
# Defines the outage rates modelled to key flowpaths in the ESOO.

# %%
show_sheet_discovery('Transmission Reliability')

# %% [markdown]
# ### Transmission unplanned outage rates
#
# Contains the verified header row 7 and data rows 8–13.
#
# Candidate source block: `B7:E13` (7 rows × 4 columns).

# %%
inspect_candidate('Transmission Reliability', 'B7:E13')

# %% [markdown]
# ## Distribution network
#
# Inputs used to model distribution network opportunities to facilitate aggregate operation of consumer energy resources and other distributed resources.

# %%
show_sheet_discovery('Distribution network')

# %% [markdown]
# ### Mid-scale generation and storage build limits
#
# Candidate source block: `B11:G38` (28 rows × 6 columns).

# %%
inspect_candidate('Distribution network', 'B11:G38')

# %% [markdown]
# ### Distribution CER augmentation tranche costs
#
# Candidate source block: `B40:H57` (18 rows × 7 columns).

# %%
inspect_candidate('Distribution network', 'B40:H57')

# %% [markdown]
# ### Average CER generation-limit time-of-day profile
#
# Candidate source block: `B59:AZ1433` (1375 rows × 51 columns).

# %%
inspect_candidate('Distribution network', 'B59:AZ1433')

# %% [markdown]
# ## Connection cost
#
# Cost to connect different generation technologies.

# %%
show_sheet_discovery('Connection cost')

# %% [markdown]
# ### Wind and solar connection costs
#
# Candidate source block: `B6:J61` (56 rows × 9 columns).

# %%
inspect_candidate('Connection cost', 'B6:J61')

# %% [markdown]
# ### Other-generation regional connection costs
#
# Candidate source block: `B62:R73` (12 rows × 17 columns).

# %%
inspect_candidate('Connection cost', 'B62:R73')

# %% [markdown]
# ## Connection cost forecasts
#
# Forecast of transmission connection costs.

# %%
show_sheet_discovery('Connection cost forecasts')

# %% [markdown]
# ### Wind and solar connection-cost forecasts
#
# Candidate source block: `B8:AJ144` (137 rows × 35 columns).

# %%
inspect_candidate('Connection cost forecasts', 'B8:AJ144')

# %% [markdown]
# ### Other-generation connection-cost forecasts
#
# Candidate source block: `B147:AJ388` (242 rows × 35 columns).

# %%
inspect_candidate('Connection cost forecasts', 'B147:AJ388')

# %% [markdown]
# ## Flow path augmentation options
#
# Capability, cost and timing for flow path augmentation options.

# %%
show_sheet_discovery('Flow path augmentation options')

# %% [markdown]
# ### Flow-path augmentation options
#
# Combines repeated physical flow-path sections into one logical option dataset.
#
# Candidate source block: `B11:Q127` (117 rows × 16 columns).

# %%
inspect_candidate('Flow path augmentation options', 'B11:Q127')

# %% [markdown]
# ## Flow path cost forecasts
#
# Forecast of flow path augmentation costs.

# %%
show_sheet_discovery('Flow path cost forecasts')

# %% [markdown]
# ### Flow-path cost forecast — Slower Growth
#
# Candidate source block: `B10:AI111` (102 rows × 34 columns).

# %%
inspect_candidate('Flow path cost forecasts', 'B10:AI111')

# %% [markdown]
# ### Flow-path cost forecast — Step Change
#
# Candidate source block: `B115:AI216` (102 rows × 34 columns).

# %%
inspect_candidate('Flow path cost forecasts', 'B115:AI216')

# %% [markdown]
# ### Flow-path cost forecast — Accelerated Transition
#
# Candidate source block: `B220:AI321` (102 rows × 34 columns).

# %%
inspect_candidate('Flow path cost forecasts', 'B220:AI321')

# %% [markdown]
# ## REZ augmentations options
#
# Capability, cost and timing for REZ augmentation options.

# %%
show_sheet_discovery('REZ augmentations options')

# %% [markdown]
# ### Queensland REZ augmentation options
#
# Candidate source block: `B10:O37` (28 rows × 14 columns).

# %%
inspect_candidate('REZ augmentations options', 'B10:O37')

# %% [markdown]
# ### New South Wales REZ augmentation options
#
# Candidate source block: `B39:O77` (39 rows × 14 columns).

# %%
inspect_candidate('REZ augmentations options', 'B39:O77')

# %% [markdown]
# ### South Australia REZ augmentation options
#
# Candidate source block: `B79:O96` (18 rows × 14 columns).

# %%
inspect_candidate('REZ augmentations options', 'B79:O96')

# %% [markdown]
# ### Tasmania REZ augmentation options
#
# Candidate source block: `B98:O110` (13 rows × 14 columns).

# %%
inspect_candidate('REZ augmentations options', 'B98:O110')

# %% [markdown]
# ### Victoria REZ augmentation options
#
# Candidate source block: `B112:O137` (26 rows × 14 columns).

# %%
inspect_candidate('REZ augmentations options', 'B112:O137')

# %% [markdown]
# ## REZ cost forecasts
#
# Forecast of REZ augmentation costs.

# %%
show_sheet_discovery('REZ cost forecasts')

# %% [markdown]
# ### REZ cost forecast — Slower Growth
#
# Candidate source block: `B11:AJ117` (107 rows × 35 columns).

# %%
inspect_candidate('REZ cost forecasts', 'B11:AJ117')

# %% [markdown]
# ### REZ cost forecast — Step Change
#
# Candidate source block: `B118:AJ224` (107 rows × 35 columns).

# %%
inspect_candidate('REZ cost forecasts', 'B118:AJ224')

# %% [markdown]
# ### REZ cost forecast — Accelerated Transition
#
# Candidate source block: `B225:AJ331` (107 rows × 35 columns).

# %%
inspect_candidate('REZ cost forecasts', 'B225:AJ331')

# %% [markdown]
# ## Distribution cost forecasts
#
# Forecast of the distribution augmentation tranche costs.

# %%
show_sheet_discovery('Distribution cost forecasts')

# %% [markdown]
# ### Distribution CER augmentation cost forecast
#
# Candidate source block: `B5:AJ84` (80 rows × 35 columns).

# %%
inspect_candidate('Distribution cost forecasts', 'B5:AJ84')

# %% [markdown]
# ## Maximum capacity
#
# Installed capacity of existing, committed and anticipated generators.

# %%
show_sheet_discovery('Maximum capacity')

# %% [markdown]
# ### Existing, committed, anticipated, and additional generator capacity
#
# Candidate source block: `B9:J750` (742 rows × 9 columns).

# %%
inspect_candidate('Maximum capacity', 'B9:J750')

# %% [markdown]
# ### New generation technology capacity
#
# Candidate source block: `L9:O31` (23 rows × 4 columns).

# %%
inspect_candidate('Maximum capacity', 'L9:O31')

# %% [markdown]
# ## Hybrid site limits
#
# Interval-level charging/dispatch limits for sites with a combination of VRE and battery storage at one connection point.

# %%
show_sheet_discovery('Hybrid site limits')

# %% [markdown]
# ### Hybrid-site limits
#
# Contains the verified hybrid-site limit source table.
#
# Candidate source block: `B9:G67` (59 rows × 6 columns).

# %%
inspect_candidate('Hybrid site limits', 'B9:G67')

# %% [markdown]
# ## Seasonal ratings
#
# Winter and summer ratings for existing, committed, advanced and new entrant generators.

# %%
show_sheet_discovery('Seasonal ratings')

# %% [markdown]
# ### New generation technology seasonal ratings
#
# Candidate source block: `B9:E36` (28 rows × 4 columns).

# %%
inspect_candidate('Seasonal ratings', 'B9:E36')

# %% [markdown]
# ### Existing, committed, anticipated, and additional generator seasonal ratings
#
# Candidate source block: `B42:AI770` (729 rows × 34 columns).

# %%
inspect_candidate('Seasonal ratings', 'B42:AI770')

# %% [markdown]
# ## Generator Reliability Settings
#
# Defines a generators' unplanned outage rate, mean time to repair after an outage, and the derating experienced during a partial outage.

# %%
show_sheet_discovery('Generator Reliability Settings')

# %% [markdown]
# ### Existing generator long-duration outages
#
# Candidate source block: `B9:M16` (8 rows × 12 columns).

# %%
inspect_candidate('Generator Reliability Settings', 'B9:M16')

# %% [markdown]
# ### Existing generator outage rates and MTTR
#
# Candidate source block: `B21:M60` (40 rows × 12 columns).

# %%
inspect_candidate('Generator Reliability Settings', 'B21:M60')

# %% [markdown]
# ### New entrant reliability settings
#
# Candidate source block: `B62:H90` (29 rows × 7 columns).

# %%
inspect_candidate('Generator Reliability Settings', 'B62:H90')

# %% [markdown]
# ## Maintenance
#
# The percentage of time per year that a generator is expected to be out of service for maintenance. De-rating applied to generators under maintenance, staged construction or to track age-related degradation.

# %%
show_sheet_discovery('Maintenance')

# %% [markdown]
# ### Existing generator maintenance rates
#
# Candidate source block: `B5:D29` (25 rows × 3 columns).

# %%
inspect_candidate('Maintenance', 'B5:D29')

# %% [markdown]
# ### New entrant maintenance rates
#
# Candidate source block: `G5:I32` (28 rows × 3 columns).

# %%
inspect_candidate('Maintenance', 'G5:I32')

# %% [markdown]
# ## Retirement
#
# Announced and end-of-technical-life generator retirement.

# %%
show_sheet_discovery('Retirement')

# %% [markdown]
# ### Expected generator closure years
#
# Candidate source block: `B8:F738` (731 rows × 5 columns).

# %%
inspect_candidate('Retirement', 'B8:F738')

# %% [markdown]
# ### Generator retirement costs
#
# Candidate source block: `H8:I50` (43 rows × 2 columns).

# %%
inspect_candidate('Retirement', 'H8:I50')

# %% [markdown]
# ## Hydro Scheme Inflows
#
# Monthly aggregated inflow trends for reference years.

# %%
show_sheet_discovery('Hydro Scheme Inflows')

# %% [markdown]
# ### Secondary hydro scheme releases and outflows
#
# Candidate source block: `B4:T79` (76 rows × 19 columns).

# %%
inspect_candidate('Hydro Scheme Inflows', 'B4:T79')

# %% [markdown]
# ### Run-of-river hydro outflows
#
# Candidate source block: `B81:T121` (41 rows × 19 columns).

# %%
inspect_candidate('Hydro Scheme Inflows', 'B81:T121')

# %% [markdown]
# ### Hydro Tasmania scheme
#
# Candidate source block: `B123:T141` (19 rows × 19 columns).

# %%
inspect_candidate('Hydro Scheme Inflows', 'B123:T141')

# %% [markdown]
# ### Snowy Hydro weather-variability representation
#
# Candidate source block: `B143:T162` (20 rows × 19 columns).

# %%
inspect_candidate('Hydro Scheme Inflows', 'B143:T162')

# %% [markdown]
# ## Capacity Factors
#
# Capacity factors for renewable generators in the renewable energy zones.

# %%
show_sheet_discovery('Capacity Factors ')

# %% [markdown]
# ### New large-scale renewable capacity factors
#
# Candidate source block: `B2:V214` (213 rows × 21 columns).

# %%
inspect_candidate('Capacity Factors ', 'B2:V214')

# %% [markdown]
# ## Heat rates
#
# Efficiency of conversion of fuel to output for thermal generators.

# %%
show_sheet_discovery('Heat rates')

# %% [markdown]
# ### Existing generator heat rates
#
# Candidate source block: `B7:E740` (734 rows × 4 columns).

# %%
inspect_candidate('Heat rates', 'B7:E740')

# %% [markdown]
# ### New entrant heat rates
#
# Candidate source block: `H7:I31` (25 rows × 2 columns).

# %%
inspect_candidate('Heat rates', 'H7:I31')

# %% [markdown]
# ## Auxiliary
#
# Auxiliary (self) load for each generator or generator class.

# %%
show_sheet_discovery('Auxiliary')

# %% [markdown]
# ### Existing generator auxiliary load
#
# Candidate source block: `B5:E736` (732 rows × 4 columns).

# %%
inspect_candidate('Auxiliary', 'B5:E736')

# %% [markdown]
# ### New entrant auxiliary load
#
# Candidate source block: `G5:H29` (25 rows × 2 columns).

# %%
inspect_candidate('Auxiliary', 'G5:H29')

# %% [markdown]
# ## Storage properties
#
# Battery storage to power ratio and round-trip efficiency.

# %%
show_sheet_discovery('Storage properties')

# %% [markdown]
# ### Battery properties
#
# Candidate source block: `B2:J19` (18 rows × 9 columns).

# %%
inspect_candidate('Storage properties', 'B2:J19')

# %% [markdown]
# ### Existing pumped-hydro properties
#
# Candidate source block: `B21:E35` (15 rows × 4 columns).

# %%
inspect_candidate('Storage properties', 'B21:E35')

# %% [markdown]
# ### New entrant pumped-hydro properties
#
# Candidate source block: `G21:J27` (7 rows × 4 columns).

# %%
inspect_candidate('Storage properties', 'G21:J27')

# %% [markdown]
# ### On-site diesel storage
#
# Candidate source block: `B38:C45` (8 rows × 2 columns).

# %%
inspect_candidate('Storage properties', 'B38:C45')

# %% [markdown]
# ## Emissions intensity
#
# Emissions production per MWh of output for each generator or generator class.

# %%
show_sheet_discovery('Emissions intensity')

# %% [markdown]
# ### Existing generator emissions intensity
#
# Candidate source block: `B4:E744` (741 rows × 4 columns).

# %%
inspect_candidate('Emissions intensity', 'B4:E744')

# %% [markdown]
# ### New entrant emissions intensity
#
# Candidate source block: `G4:H29` (26 rows × 2 columns).

# %%
inspect_candidate('Emissions intensity', 'G4:H29')

# %% [markdown]
# ## Build costs
#
# Capital Cost projections for new entrant generators.

# %%
show_sheet_discovery('Build costs')

# %% [markdown]
# ### New entrant capital-cost projections
#
# Candidate source block: `B2:AJ77` (76 rows × 35 columns).

# %%
inspect_candidate('Build costs', 'B2:AJ77')

# %% [markdown]
# ## Fixed OPEX
#
# Fixed operating cost regardless of output for each generator or generator class.

# %%
show_sheet_discovery('Fixed OPEX')

# %% [markdown]
# ### Existing generator fixed OPEX
#
# Candidate source block: `B5:E739` (735 rows × 4 columns).

# %%
inspect_candidate('Fixed OPEX', 'B5:E739')

# %% [markdown]
# ### New entrant fixed OPEX
#
# Candidate source block: `G5:I32` (28 rows × 3 columns).

# %%
inspect_candidate('Fixed OPEX', 'G5:I32')

# %% [markdown]
# ## Variable OPEX
#
# Variable operating cost per MWh of output for each generator or generator class.

# %%
show_sheet_discovery('Variable OPEX')

# %% [markdown]
# ### Existing generator variable OPEX
#
# Candidate source block: `B5:E738` (734 rows × 4 columns).

# %%
inspect_candidate('Variable OPEX', 'B5:E738')

# %% [markdown]
# ### New entrant variable OPEX
#
# Candidate source block: `G5:H32` (28 rows × 2 columns).

# %%
inspect_candidate('Variable OPEX', 'G5:H32')

# %% [markdown]
# ## Marginal Loss Factors
#
# Marginal loss factors for each generator.

# %%
show_sheet_discovery('Marginal Loss Factors')

# %% [markdown]
# ### Existing generator marginal loss factors
#
# Candidate source block: `B10:F748` (739 rows × 5 columns).

# %%
inspect_candidate('Marginal Loss Factors', 'B10:F748')

# %% [markdown]
# ### New entrant generator marginal loss factors
#
# Candidate source block: `I10:M536` (527 rows × 5 columns).

# %%
inspect_candidate('Marginal Loss Factors', 'I10:M536')

# %% [markdown]
# ### New entrant electrolyser marginal loss factors
#
# Candidate source block: `O10:S161` (152 rows × 5 columns).

# %%
inspect_candidate('Marginal Loss Factors', 'O10:S161')

# %% [markdown]
# ## Locational Cost Factors
#
# Locational cost factors provide an indication of the variation in new entrants generators cost based on the shift in labour, equipment and shipping/delivery cost between regions.

# %%
show_sheet_discovery('Locational Cost Factors')

# %% [markdown]
# ### Non-pumped-hydro locational cost factors
#
# Candidate source block: `B9:H80` (72 rows × 7 columns).

# %%
inspect_candidate('Locational Cost Factors', 'B9:H80')

# %% [markdown]
# ### Pumped-hydro locational cost factors
#
# Candidate source block: `B83:I132` (50 rows × 8 columns).

# %%
inspect_candidate('Locational Cost Factors', 'B83:I132')

# %% [markdown]
# ### Technology cost breakdown ratios
#
# Candidate source block: `B134:G158` (25 rows × 6 columns).

# %%
inspect_candidate('Locational Cost Factors', 'B134:G158')

# %% [markdown]
# ### Technology-specific locational cost factors
#
# Candidate source block: `B161:X227` (67 rows × 23 columns).

# %%
inspect_candidate('Locational Cost Factors', 'B161:X227')

# %% [markdown]
# ## Build limits - REZs
#
# Modelled limitations for REZs impacting build constraints within the expansion modelling.

# %%
show_sheet_discovery('Build limits - REZs')

# %% [markdown]
# ### Initial REZ resource limits
#
# Candidate source block: `B2:Q62` (61 rows × 16 columns).

# %%
inspect_candidate('Build limits - REZs', 'B2:Q62')

# %% [markdown]
# ### Initial REZ transmission limits
#
# Candidate source block: `B64:N119` (56 rows × 13 columns).

# %%
inspect_candidate('Build limits - REZs', 'B64:N119')

# %% [markdown]
# ### REZ transmission modifiers
#
# Candidate source block: `B121:F132` (12 rows × 5 columns).

# %%
inspect_candidate('Build limits - REZs', 'B121:F132')

# %% [markdown]
# ### REZ group constraints
#
# Candidate source block: `B136:K265` (130 rows × 10 columns).

# %%
inspect_candidate('Build limits - REZs', 'B136:K265')

# %% [markdown]
# ### REZ transmission limit constraints
#
# Candidate source block: `B267:K317` (51 rows × 10 columns).

# %%
inspect_candidate('Build limits - REZs', 'B267:K317')

# %% [markdown]
# ### REZ secondary transmission limits
#
# Candidate source block: `B319:K335` (17 rows × 10 columns).

# %%
inspect_candidate('Build limits - REZs', 'B319:K335')

# %% [markdown]
# ### Non-REZ connections pipeline build limits
#
# Candidate source block: `B337:E356` (20 rows × 4 columns).

# %%
inspect_candidate('Build limits - REZs', 'B337:E356')

# %% [markdown]
# ### REZ technology-specific access-right limits
#
# Candidate source block: `B358:G368` (11 rows × 6 columns).

# %%
inspect_candidate('Build limits - REZs', 'B358:G368')

# %% [markdown]
# ## Build limits - PHES
#
# Modelled limitations for PHES impacting build constraints within the expansion modelling.

# %%
show_sheet_discovery('Build limits - PHES')

# %% [markdown]
# ### Pumped-hydro build limits
#
# Candidate source block: `B2:W27` (26 rows × 22 columns).

# %%
inspect_candidate('Build limits - PHES', 'B2:W27')

# %% [markdown]
# ## First-of-a-kind premium
#
# Premiums applied to emerging generation technologies to reflect the tendency of first-of-a-kind installations to exceed estimated costs.

# %%
show_sheet_discovery('First-of-a-kind premium')

# %% [markdown]
# ### First-of-a-kind premium factors
#
# Candidate source block: `B2:D11` (10 rows × 3 columns).

# %%
inspect_candidate('First-of-a-kind premium', 'B2:D11')

# %% [markdown]
# ## Lead time and project life
#
# Modelled limitations impacting build timings within the expansion modelling.

# %%
show_sheet_discovery('Lead time and project life')

# %% [markdown]
# ### Lead times and project lives
#
# Candidate source block: `B2:H35` (34 rows × 7 columns).

# %%
inspect_candidate('Lead time and project life', 'B2:H35')

# %% [markdown]
# ## Financial parameters
#
# Financial parameters (discount rate, weighted average cost of capital, value of customer reliability, and value of emissions reductions) used during cost benefit analysis.

# %%
show_sheet_discovery('Financial parameters')

# %% [markdown]
# ### Discount rate
#
# Candidate source block: `B2:F7` (6 rows × 5 columns).

# %%
inspect_candidate('Financial parameters', 'B2:F7')

# %% [markdown]
# ### Weighted Average Cost of Capital
#
# Candidate source block: `B10:F41` (32 rows × 5 columns).

# %%
inspect_candidate('Financial parameters', 'B10:F41')

# %% [markdown]
# ### Value of Customer Reliability
#
# Candidate source block: `B43:G51` (9 rows × 6 columns).

# %%
inspect_candidate('Financial parameters', 'B43:G51')

# %% [markdown]
# ### Value of emissions reduction
#
# Candidate source block: `B54:C90` (37 rows × 2 columns).

# %%
inspect_candidate('Financial parameters', 'B54:C90')

# %% [markdown]
# ## Affine Heat rates
#
# Heat rate curves for large thermal units.

# %%
show_sheet_discovery('Affine Heat rates')

# %% [markdown]
# ### Existing generator affine heat rates
#
# Candidate source block: `B6:F192` (187 rows × 5 columns).

# %%
inspect_candidate('Affine Heat rates', 'B6:F192')

# %% [markdown]
# ### New entrant affine heat rates
#
# Candidate source block: `H6:K29` (24 rows × 4 columns).

# %%
inspect_candidate('Affine Heat rates', 'H6:K29')

# %% [markdown]
# ## Max Ramp Rates
#
# Maximum rates of change for thermal unit output up and down.

# %%
show_sheet_discovery('Max Ramp Rates')

# %% [markdown]
# ### Existing thermal generator maximum ramp rates
#
# Candidate source block: `B7:F191` (185 rows × 5 columns).

# %%
inspect_candidate('Max Ramp Rates', 'B7:F191')

# %% [markdown]
# ### New entrant maximum ramp rates
#
# Candidate source block: `H7:J30` (24 rows × 3 columns).

# %%
inspect_candidate('Max Ramp Rates', 'H7:J30')

# %% [markdown]
# ## Coal Min Stable Level
#
# Coal generator minimum stable levels.

# %%
show_sheet_discovery('Coal Min Stable Level')

# %% [markdown]
# ### Coal generator minimum stable levels
#
# Candidate source block: `B2:G63` (62 rows × 6 columns).

# %%
inspect_candidate('Coal Min Stable Level', 'B2:G63')

# %% [markdown]
# ## GPG Min Stable Level
#
# Minimum operating levels for large GPG units.

# %%
show_sheet_discovery('GPG Min Stable Level')

# %% [markdown]
# ### Existing GPG minimum stable levels
#
# Candidate source block: `B10:E150` (141 rows × 4 columns).

# %%
inspect_candidate('GPG Min Stable Level', 'B10:E150')

# %% [markdown]
# ### New entrant GPG minimum stable levels
#
# Candidate source block: `G10:H35` (26 rows × 2 columns).

# %%
inspect_candidate('GPG Min Stable Level', 'G10:H35')

# %% [markdown]
# ## Coal and Biomass price
#
# Coal fuel price for each coal generator.

# %%
show_sheet_discovery('Coal and Biomass price')

# %% [markdown]
# ### Coal fuel prices
#
# Candidate source block: `B8:AG54` (47 rows × 32 columns).

# %%
inspect_candidate('Coal and Biomass price', 'B8:AG54')

# %% [markdown]
# ### Biomass fuel prices
#
# Candidate source block: `B57:AG61` (5 rows × 32 columns).

# %%
inspect_candidate('Coal and Biomass price', 'B57:AG61')

# %% [markdown]
# ## Gas, Liquid fuel, H2 price
#
# Fuel price for each gas and liquid fuel generator.

# %%
show_sheet_discovery('Gas, Liquid fuel, H2 price')

# %% [markdown]
# ### Existing GPG fuel costs
#
# Candidate source block: `B7:AG129` (123 rows × 32 columns).

# %%
inspect_candidate('Gas, Liquid fuel, H2 price', 'B7:AG129')

# %% [markdown]
# ### New entrant GPG fuel costs
#
# Candidate source block: `B132:AG224` (93 rows × 32 columns).

# %%
inspect_candidate('Gas, Liquid fuel, H2 price', 'B132:AG224')

# %% [markdown]
# ### Industrial fuel costs
#
# Candidate source block: `B228:AG249` (22 rows × 32 columns).

# %%
inspect_candidate('Gas, Liquid fuel, H2 price', 'B228:AG249')

# %% [markdown]
# ### Residential and commercial fuel costs
#
# Candidate source block: `B253:AG274` (22 rows × 32 columns).

# %%
inspect_candidate('Gas, Liquid fuel, H2 price', 'B253:AG274')

# %% [markdown]
# ### Liquid fuel prices
#
# Candidate source block: `B278:AG302` (25 rows × 32 columns).

# %%
inspect_candidate('Gas, Liquid fuel, H2 price', 'B278:AG302')

# %% [markdown]
# ### GPG secondary liquid-fuel prices
#
# Candidate source block: `B305:AG429` (125 rows × 32 columns).

# %%
inspect_candidate('Gas, Liquid fuel, H2 price', 'B305:AG429')

# %% [markdown]
# ### Hydrogen prices
#
# Candidate source block: `B433:AG438` (6 rows × 32 columns).

# %%
inspect_candidate('Gas, Liquid fuel, H2 price', 'B433:AG438')

# %% [markdown]
# ### Biomethane prices
#
# Candidate source block: `B440:AG452` (13 rows × 32 columns).

# %%
inspect_candidate('Gas, Liquid fuel, H2 price', 'B440:AG452')

# %% [markdown]
# ## Gas System Properties
#
# Key properties of Natural Gas Pipelines, Processing Facilities, Storage Facilities, Reserves and Resources, Pipeline Transmission Tariffs, and Production Costs.

# %%
show_sheet_discovery('Gas System Properties')

# %% [markdown]
# ### Gas pipelines
#
# Candidate source block: `B7:F49` (43 rows × 5 columns).

# %%
inspect_candidate('Gas System Properties', 'B7:F49')

# %% [markdown]
# ### Gas processing facilities
#
# Candidate source block: `B51:G105` (55 rows × 6 columns).

# %%
inspect_candidate('Gas System Properties', 'B51:G105')

# %% [markdown]
# ### Gas storage facilities
#
# Candidate source block: `B108:H122` (15 rows × 7 columns).

# %%
inspect_candidate('Gas System Properties', 'B108:H122')

# %% [markdown]
# ### Gas reserves and resources
#
# Candidate source block: `B124:F144` (21 rows × 5 columns).

# %%
inspect_candidate('Gas System Properties', 'B124:F144')

# %% [markdown]
# ### Pipeline transmission tariffs
#
# Candidate source block: `B146:E169` (24 rows × 4 columns).

# %%
inspect_candidate('Gas System Properties', 'B146:E169')

# %% [markdown]
# ### Gas production costs
#
# Candidate source block: `B171:E185` (15 rows × 4 columns).

# %%
inspect_candidate('Gas System Properties', 'B171:E185')

# %% [markdown]
# ## GPG emissions reduction - BioM
#
# Factors projecting emissions reduction from GPG due to blending of biomethane into fuel gas.

# %%
show_sheet_discovery('GPG emissions reduction - BioM')

# %% [markdown]
# ### GPG biomethane-blending emissions factor
#
# Candidate source block: `B2:AF12` (11 rows × 31 columns).

# %%
inspect_candidate('GPG emissions reduction - BioM', 'B2:AF12')

# %% [markdown]
# ## Power System Security
#
# Reflect power system constraints to reflect secure operating limits.

# %%
show_sheet_discovery('Power System Security')

# %% [markdown]
# ### Coal-retirement minimum-fault-level costs
#
# Candidate source block: `B4:D49` (46 rows × 3 columns).

# %%
inspect_candidate('Power System Security', 'B4:D49')

# %% [markdown]
# ### Efficient system-strength costs
#
# Candidate source block: `B52:AE56` (5 rows × 30 columns).

# %%
inspect_candidate('Power System Security', 'B52:AE56')

# %% [markdown]
# ### Synchronous unit commitment — standard scenarios
#
# Candidate source block: `B58:G72` (15 rows × 6 columns).

# %%
inspect_candidate('Power System Security', 'B58:G72')

# %% [markdown]
# ### Synchronous unit commitment — Accelerated Transition
#
# Candidate source block: `B74:G94` (21 rows × 6 columns).

# %%
inspect_candidate('Power System Security', 'B74:G94')

# %% [markdown]
# ## Reserves
#
# Minimum reserve levels for reliable regional supply.

# %%
show_sheet_discovery('Reserves')

# %% [markdown]
# ### Initial regional reserves
#
# Candidate source block: `B2:C14` (13 rows × 2 columns).

# %%
inspect_candidate('Reserves', 'B2:C14')

# %% [markdown]
# ## Hydrogen demand - Domestic
#
# Projections of domestic hydrogen demand.

# %%
show_sheet_discovery('Hydrogen demand - Domestic')

# %% [markdown]
# ### Domestic hydrogen demand
#
# Candidate source block: `B2:AH53` (52 rows × 33 columns).

# %%
inspect_candidate('Hydrogen demand - Domestic', 'B2:AH53')

# %% [markdown]
# ## Hydrogen monthly profiles
#
# Projections of monthly profiles of export/commodities hydrogen demand.

# %%
show_sheet_discovery('Hydrogen monthly profiles')

# %% [markdown]
# ### Hydrogen monthly consumption profiles
#
# Candidate source block: `B2:AG44` (43 rows × 32 columns).

# %%
inspect_candidate('Hydrogen monthly profiles', 'B2:AG44')

# %% [markdown]
# ## Hydrogen demand-Export&Commod
#
# Projections of demand for export hydrogen and green commodities production.

# %%
show_sheet_discovery('Hydrogen demand-Export&Commod')

# %% [markdown]
# ### Hydrogen export demand
#
# Candidate source block: `B2:AH52` (51 rows × 33 columns).

# %%
inspect_candidate('Hydrogen demand-Export&Commod', 'B2:AH52')

# %% [markdown]
# ### Hydrogen demand for green commodities
#
# Candidate source block: `B54:AH105` (52 rows × 33 columns).

# %%
inspect_candidate('Hydrogen demand-Export&Commod', 'B54:AH105')

# %% [markdown]
# ### Electricity demand for green steel
#
# Candidate source block: `B107:AH156` (50 rows × 33 columns).

# %%
inspect_candidate('Hydrogen demand-Export&Commod', 'B107:AH156')

# %% [markdown]
# ## Hydrogen consumption locations
#
# Location of hydrogen consumption and candidate hydrogen hubs and ports.

# %%
show_sheet_discovery('Hydrogen consumption locations')

# %% [markdown]
# ### Hydrogen consumption locations
#
# Candidate source block: `B5:F40` (36 rows × 5 columns).

# %%
inspect_candidate('Hydrogen consumption locations', 'B5:F40')

# %% [markdown]
# ### Hydrogen hubs
#
# Candidate source block: `B42:B44` (3 rows × 1 columns).

# %%
inspect_candidate('Hydrogen consumption locations', 'B42:B44')

# %% [markdown]
# ### Hydrogen and commodity export ports
#
# Candidate source block: `B46:D57` (12 rows × 3 columns).

# %%
inspect_candidate('Hydrogen consumption locations', 'B46:D57')

# %% [markdown]
# ## Water for Hydrogen
#
# Projections of treated water required for electrolytic hydrogen production.

# %%
show_sheet_discovery('Water for Hydrogen')

# %% [markdown]
# ### Water required for hydrogen production
#
# Candidate source block: `B2:AH52` (51 rows × 33 columns).

# %%
inspect_candidate('Water for Hydrogen', 'B2:AH52')

# %% [markdown]
# ## Desalination demand for H2
#
# Projections of electricity required for water treatment associated with electrolytic hydrogen production.

# %%
show_sheet_discovery('Desalination demand for H2')

# %% [markdown]
# ### Desalination electricity demand for hydrogen
#
# Candidate source block: `B2:AH52` (51 rows × 33 columns).

# %%
inspect_candidate('Desalination demand for H2', 'B2:AH52')

# %% [markdown]
# ## H2 as fuel for GPG Limit
#
# Maximum fuel limit applied to hydrogen for GPG.

# %%
show_sheet_discovery('H2 as fuel for GPG Limit')

# %% [markdown]
# ### Hydrogen-as-GPG-fuel limit
#
# Candidate source block: `B2:AG21` (20 rows × 32 columns).

# %%
inspect_candidate('H2 as fuel for GPG Limit', 'B2:AG21')

# %% [markdown]
# ## Build Cost - Hydrogen pipeline
#
# Projections of cost to build hydrogen pipelines in a given REZ.

# %%
show_sheet_discovery('Build Cost - Hydrogen pipeline')

# %% [markdown]
# ### Hydrogen pipeline build costs
#
# Candidate source block: `B2:AJ156` (155 rows × 35 columns).

# %%
inspect_candidate('Build Cost - Hydrogen pipeline', 'B2:AJ156')

# %% [markdown]
# ## Other hydrogen assumptions
#
# Various hydrogen-related assumptions.

# %%
show_sheet_discovery('Other hydrogen assumptions')

# %% [markdown]
# ### Hydrogen heat content
#
# Candidate source block: `B2:C5` (4 rows × 2 columns).

# %%
inspect_candidate('Other hydrogen assumptions', 'B2:C5')

# %% [markdown]
# ### Electrolyser electricity consumption rate
#
# Candidate source block: `B7:AF11` (5 rows × 31 columns).

# %%
inspect_candidate('Other hydrogen assumptions', 'B7:AF11')

# %% [markdown]
# ### Water cost for hydrogen production
#
# Candidate source block: `B13:AF17` (5 rows × 31 columns).

# %%
inspect_candidate('Other hydrogen assumptions', 'B13:AF17')

# %% [markdown]
# ### Ammonia-conversion electricity adjustment
#
# Candidate source block: `B19:AF23` (5 rows × 31 columns).

# %%
inspect_candidate('Other hydrogen assumptions', 'B19:AF23')

# %% [markdown]
# ### Minimum annual electrolyser utilisation factor
#
# Candidate source block: `B26:AF30` (5 rows × 31 columns).

# %%
inspect_candidate('Other hydrogen assumptions', 'B26:AF30')

# %% [markdown]
# ### Electrolyser balance of plant
#
# Candidate source block: `B32:C35` (4 rows × 2 columns).

# %%
inspect_candidate('Other hydrogen assumptions', 'B32:C35')

# %% [markdown]
# ## Summary Mapping
#
# Master look-up table for generator, storage, and electrolyser assets.

# %%
show_sheet_discovery('Summary Mapping')

# %% [markdown]
# ### Existing, committed, and anticipated asset mapping
#
# Candidate source block: `C2:AF733` (732 rows × 30 columns).

# %%
inspect_candidate('Summary Mapping', 'C2:AF733')

# %% [markdown]
# ### Consumer energy resource mapping
#
# Candidate source block: `C734:AF786` (53 rows × 30 columns).

# %%
inspect_candidate('Summary Mapping', 'C734:AF786')

# %% [markdown]
# ### New entrant asset mapping
#
# Candidate source block: `C790:AF1316` (527 rows × 30 columns).

# %%
inspect_candidate('Summary Mapping', 'C790:AF1316')

# %% [markdown]
# ### New entrant electrolyser mapping
#
# Candidate source block: `C1319:AF1381` (63 rows × 30 columns).

# %%
inspect_candidate('Summary Mapping', 'C1319:AF1381')

# %% [markdown]
# ## Coverage check
#
# The catalogue must account for every worksheet in workbook order. A worksheet can legitimately have no
# embedded semantic data table; those cases remain explicit rather than being silently omitted.

# %%
assert list(SHEET_RANGES) == workbook_formula.sheetnames[:len(SHEET_RANGES)]
sum(len(ranges) for ranges in SHEET_RANGES.values())

