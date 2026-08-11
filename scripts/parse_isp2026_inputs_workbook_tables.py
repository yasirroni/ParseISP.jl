# %% [markdown]
# # 2026 ISP inputs and assumptions workbook — semantic tables
#
# This notebook reads the semantic source tables identified by the discovery notebook.
# Source ranges are explicit Excel coordinates. Complex multi-row headers remain in the source block,
# and genuinely missing cells remain missing unless a table-specific workbook rule says otherwise.

# %%
from __future__ import annotations

import os
import warnings
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries

try:
    from IPython import get_ipython
    from IPython.display import display
except ImportError:
    def get_ipython():
        return None
    display = print

_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_NS_REL_DOC = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_NS_REL_PKG = 'http://schemas.openxmlformats.org/package/2006/relationships'
warnings.filterwarnings('ignore', message='Data Validation extension is not supported and will be removed', category=UserWarning, module='openpyxl.worksheet._reader')

def _repo_root() -> Path:
    try:
        return Path(__file__).resolve().parent.parent
    except NameError:
        cwd = Path.cwd().resolve()
        for candidate in (cwd, *cwd.parents):
            if (candidate / 'Project.toml').is_file() and (candidate / 'scripts').is_dir():
                return candidate
        raise RuntimeError('Could not locate the ParseISP.jl repository root.')

_PROJECT = _repo_root()
_DEFAULT_WORKBOOK = _PROJECT / 'scrapped' / 'ISP_reference_files' / '2026-isp-inputs-and-assumptions-workbook.xlsm'
WORKBOOK_PATH = Path(os.environ.get('PARSEISP_2026_INPUTS_WORKBOOK', _DEFAULT_WORKBOOK)).expanduser().resolve()
if not WORKBOOK_PATH.is_file():
    raise FileNotFoundError(f'ISP 2026 inputs workbook not found: {WORKBOOK_PATH}')

SHEET_RANGES = {'Disclaimer': [],
 'Change Log': ['B2:E657'],
 'Assumptions Summary': ['B6:D35', 'B37:E119', 'B121:C147'],
 'Scenarios': ['B5:E29'],
 'Existing Gen Data Summary': ['B10:AT738'],
 'New Entrant Data Summary': ['B9:BB535'],
 'New Electrolyser Data Summary': ['B5:AQ67'],
 'Fuel Price Summary': ['B7:S9', 'B11:AK738', 'B743:AK1268'],
 'Regional Build Costs Summary': ['B7:C10', 'B12:AV75'],
 'Energy Policy Targets': ['C15:E30',
                           'C32:F62',
                           'C65:G93',
                           'C96:E110',
                           'C114:N154',
                           'C157:E159',
                           'C161:I185',
                           'C187:F193',
                           'C195:E197',
                           'C200:E202',
                           'C206:E209',
                           'C213:E220',
                           'C223:E235',
                           'C239:F258',
                           'C262:E267',
                           'C271:E289',
                           'D291:E295',
                           'C297:E310',
                           'C313:E327'],
 'Carbon Budgets': ['B5:E9', 'B15:G21', 'B25:D31'],
 'Economic Growth Forecasts': ['B5:E8', 'B12:AG19', 'B21:AG28', 'B30:AG37', 'B41:AG48', 'B50:AG57', 'B59:AG66'],
 'Demand and Energy Forecasts': [],
 'End use fuel consumption data': ['B6:AF15', 'B17:AF26', 'B28:AF37'],
 'Appliance Uptake Forecasts': ['B13:AG20', 'B22:AG29', 'B31:AG38'],
 'Elec. Retail Price Indices': ['B8:AG12'],
 'Connections Forecasts': ['B8:E11', 'B15:AH22', 'B24:AH31', 'B33:AH40'],
 'Energy Efficiency': ['B8:G10',
                       'B14:AG30',
                       'B32:AG48',
                       'B50:AG66',
                       'B68:AG84',
                       'B86:AG102',
                       'B106:AG122',
                       'B124:AG140',
                       'B142:AG158',
                       'B160:AG176',
                       'B178:AG194'],
 'Rooftop PV': ['B8:E10', 'B12:AH63', 'B65:AH116'],
 'PVNSG': ['B8:E10', 'B12:AH63', 'B65:AH116'],
 'ONSG': ['B8:AH55', 'B57:AG74'],
 'Battery & Plug-in EVs': ['B7:E9', 'B11:AH62', 'B64:AH115'],
 'Fuel cell EVs': ['B7:E9', 'B11:AH62'],
 'EV V2G': ['B8:E10', 'B12:AH62', 'B64:AH115'],
 'Data Centre Forecasts': ['B6:E8', 'B10:AF16', 'B18:AF24', 'B26:AF32'],
 'DSP': ['B7:AI84', 'B87:AI164'],
 'Electrification': ['B7:E10', 'B12:AF19', 'B21:AF28', 'B30:AF37'],
 'Embedded energy storages': ['B7:E9', 'B11:AH62', 'B65:AH116'],
 'Aggregated energy storages': ['B7:E9', 'B11:AH62', 'B65:AH116'],
 'Network representation': ['B2:E22', 'B24:D42', 'B44:D53', 'B55:D63', 'B65:D82'],
 'Renewable energy zones': ['B6:E53'],
 'Network capability': ['B8:K25', 'B34:K42', 'B51:N60', 'B75:V84', 'B89:E94', 'B99:D115', 'B122:C134', 'B139:C148'],
 'Network losses': ['B5:J28', 'B30:J34', 'B36:J88'],
 'Transmission Reliability': ['B7:E13'],
 'Distribution network': ['B11:G38', 'B40:H57', 'B59:AZ1433'],
 'Connection cost': ['B6:J61', 'B62:R73'],
 'Connection cost forecasts': ['B8:AJ144', 'B147:AJ388'],
 'Flow path augmentation options': ['B11:Q127'],
 'Flow path cost forecasts': ['B10:AI111', 'B115:AI216', 'B220:AI321'],
 'REZ augmentations options': ['B10:O37', 'B39:O77', 'B79:O96', 'B98:O110', 'B112:O137'],
 'REZ cost forecasts': ['B11:AJ117', 'B118:AJ224', 'B225:AJ331'],
 'Distribution cost forecasts': ['B5:AJ84'],
 'Maximum capacity': ['B9:J750', 'L9:O31'],
 'Hybrid site limits': ['B9:G67'],
 'Seasonal ratings': ['B9:E36', 'B42:AI770'],
 'Generator Reliability Settings': ['B9:M16', 'B21:M60', 'B62:H90'],
 'Maintenance': ['B5:D29', 'G5:I32'],
 'Retirement': ['B8:F738', 'H8:I50'],
 'Hydro Scheme Inflows': ['B4:T79', 'B81:T121', 'B123:T141', 'B143:T162'],
 'Capacity Factors ': ['B2:V214'],
 'Heat rates': ['B7:E740', 'H7:I31'],
 'Auxiliary': ['B5:E736', 'G5:H29'],
 'Storage properties': ['B2:J19', 'B21:E35', 'G21:J27', 'B38:C45'],
 'Emissions intensity': ['B4:E744', 'G4:H29'],
 'Build costs': ['B2:AJ77'],
 'Fixed OPEX': ['B5:E739', 'G5:I32'],
 'Variable OPEX': ['B5:E738', 'G5:H32'],
 'Marginal Loss Factors': ['B10:F748', 'I10:M536', 'O10:S161'],
 'Locational Cost Factors': ['B9:H80', 'B83:I132', 'B134:G158', 'B161:X227'],
 'Build limits - REZs': ['B2:Q62',
                         'B64:N119',
                         'B121:F132',
                         'B136:K265',
                         'B267:K317',
                         'B319:K335',
                         'B337:E356',
                         'B358:G368'],
 'Build limits - PHES': ['B2:W27'],
 'First-of-a-kind premium': ['B2:D11'],
 'Lead time and project life': ['B2:H35'],
 'Financial parameters': ['B2:F7', 'B10:F41', 'B43:G51', 'B54:C90'],
 'Affine Heat rates': ['B6:F192', 'H6:K29'],
 'Max Ramp Rates': ['B7:F191', 'H7:J30'],
 'Coal Min Stable Level': ['B2:G63'],
 'GPG Min Stable Level': ['B10:E150', 'G10:H35'],
 'Coal and Biomass price': ['B8:AG54', 'B57:AG61'],
 'Gas, Liquid fuel, H2 price': ['B7:AG129',
                                'B132:AG224',
                                'B228:AG249',
                                'B253:AG274',
                                'B278:AG302',
                                'B305:AG429',
                                'B433:AG438',
                                'B440:AG452'],
 'Gas System Properties': ['B7:F49', 'B51:G105', 'B108:H122', 'B124:F144', 'B146:E169', 'B171:E185'],
 'GPG emissions reduction - BioM': ['B2:AF12'],
 'Power System Security': ['B4:D49', 'B52:AE56', 'B58:G72', 'B74:G94'],
 'Reserves': ['B2:C14'],
 'Hydrogen demand - Domestic': ['B2:AH53'],
 'Hydrogen monthly profiles': ['B2:AG44'],
 'Hydrogen demand-Export&Commod': ['B2:AH52', 'B54:AH105', 'B107:AH156'],
 'Hydrogen consumption locations': ['B5:F40', 'B42:B44', 'B46:D57'],
 'Water for Hydrogen': ['B2:AH52'],
 'Desalination demand for H2': ['B2:AH52'],
 'H2 as fuel for GPG Limit': ['B2:AG21'],
 'Build Cost - Hydrogen pipeline': ['B2:AJ156'],
 'Other hydrogen assumptions': ['B2:C5', 'B7:AF11', 'B13:AF17', 'B19:AF23', 'B26:AF30', 'B32:C35'],
 'Summary Mapping': ['C2:AF733', 'C734:AF786', 'C790:AF1316', 'C1319:AF1381']}

workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True, keep_links=False)

_SHEET_CACHE: dict[str, pd.DataFrame] = {}

def _sheet_source_frame(sheet_name: str) -> pd.DataFrame:
    if sheet_name in _SHEET_CACHE:
        return _SHEET_CACHE[sheet_name]
    ranges = SHEET_RANGES[sheet_name]
    if not ranges:
        return pd.DataFrame()
    bounds = [range_boundaries(cell_range) for cell_range in ranges]
    min_col = min(bound[0] for bound in bounds); min_row = min(bound[1] for bound in bounds)
    max_col = max(bound[2] for bound in bounds); max_row = max(bound[3] for bound in bounds)
    ws = workbook[sheet_name]
    values = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    frame = pd.DataFrame(values, index=pd.Index(range(min_row, max_row + 1), name="excel_row"), columns=[get_column_letter(c) for c in range(min_col, max_col + 1)])
    _SHEET_CACHE[sheet_name] = frame
    return frame

def read_source_range(sheet_name: str, cell_range: str) -> pd.DataFrame:
    """Read one exact Excel source range without inventing or filling values."""
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    columns = [get_column_letter(c) for c in range(min_col, max_col + 1)]
    return _sheet_source_frame(sheet_name).loc[min_row:max_row, columns].copy()

def _worksheet_xml_path(path: Path, sheet_name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        book = ET.fromstring(archive.read('xl/workbook.xml'))
        rels = ET.fromstring(archive.read('xl/_rels/workbook.xml.rels'))
        rel_map = {rel.attrib['Id']: rel.attrib['Target'] for rel in rels.findall(f'{{{_NS_REL_PKG}}}Relationship')}
        for sheet in book.find(f'{{{_NS}}}sheets'):
            if sheet.attrib['name'] == sheet_name:
                target = rel_map[sheet.attrib[f'{{{_NS_REL_DOC}}}id']].lstrip('/')
                return target if target.startswith('xl/') else 'xl/' + target
    raise KeyError(sheet_name)

def _merged_ranges(path: Path, sheet_name: str) -> list[str]:
    xml_path = _worksheet_xml_path(path, sheet_name)
    result = []
    with zipfile.ZipFile(path) as archive, archive.open(xml_path) as handle:
        for event, elem in ET.iterparse(handle, events=('end',)):
            if elem.tag == f'{{{_NS}}}mergeCell':
                result.append(elem.attrib['ref'])
            elem.clear()
    return result

def parse_flow_path_augmentation_options() -> pd.DataFrame:
    """Return the 62 semantic flow-path option rows while preserving genuine nulls."""
    frame = read_source_range('Flow path augmentation options', 'B11:Q127')
    option_rows = frame.index[frame.get('E').notna() & (frame.get('E') != 'Option name')]
    result = frame.loc[option_rows].copy()
    for merged in _merged_ranges(WORKBOOK_PATH, 'Flow path augmentation options'):
        min_col, min_row, max_col, max_row = range_boundaries(merged)
        if min_col <= 2 <= max_col:
            anchor = frame.at[min_row, 'B'] if min_row in frame.index and 'B' in frame.columns else None
            if anchor is not None:
                for row in result.index:
                    if min_row <= row <= max_row and pd.isna(result.at[row, 'B']):
                        result.at[row, 'B'] = anchor
    assert len(result) == 62, f'Expected 62 semantic option rows, found {len(result)}'
    assert result.loc[[59, 60, 61], 'H'].isna().all(), 'Rows 59–61 must retain missing direction values.'
    return result

def validate_table(frame: pd.DataFrame, spec: dict) -> None:
    if 'expected_rows' in spec:
        assert len(frame) == spec['expected_rows'], (spec['name'], len(frame), spec['expected_rows'])
    if 'expected_cols' in spec:
        assert len(frame.columns) == spec['expected_cols'], (spec['name'], len(frame.columns), spec['expected_cols'])
    if 'expected_semantic_rows' in spec:
        assert len(frame) == spec['expected_semantic_rows'], (spec['name'], len(frame), spec['expected_semantic_rows'])

def parse_spec(sheet_name: str, spec: dict) -> pd.DataFrame:
    frame = parse_flow_path_augmentation_options() if spec.get('parser') == 'flow_path_options' else read_source_range(sheet_name, spec['range'])
    validate_table(frame, spec)
    return frame

def show_table(frame: pd.DataFrame) -> None:
    """Display a table in notebooks and print only its shape in script mode."""
    if get_ipython() is None:
        print(f"{frame.shape[0]} rows × {frame.shape[1]} columns")
    else:
        display(frame)

# %% [markdown]
# ## Disclaimer
#
# States the workbook purpose, limitations, and conditions of use.
#
# The worksheet contains notice, purpose, and disclaimer text rather than a semantic data table.

# %% [markdown]
# ## Change Log
#
# Records changes made across workbook releases.

# %% [markdown]
# ### Workbook change log
#
# Records workbook versions, dates, changes, and supporting detail.
#
# Source block: `Change Log!B2:E657` (656 rows × 4 columns).

# %%
change_log_workbook_change_log = parse_spec('Change Log', {'name': 'Workbook change log', 'range': 'B2:E657'})
show_table(change_log_workbook_change_log)

# %% [markdown]
# ## Assumptions Summary
#
# Provides workbook metadata, worksheet descriptions, and supporting materials.

# %% [markdown]
# ### Version history
#
# Records workbook version numbers, dates, and descriptions.
#
# Source block: `Assumptions Summary!B6:D35` (30 rows × 3 columns).

# %%
assumptions_summary_version_history = parse_spec('Assumptions Summary', {'name': 'Version history', 'range': 'B6:D35'})
show_table(assumptions_summary_version_history)

# %% [markdown]
# ### Worksheet descriptions
#
# Maps worksheets to assumption groups, descriptions, and sources.
#
# Source block: `Assumptions Summary!B37:E119` (83 rows × 4 columns).

# %%
assumptions_summary_worksheet_descriptions = parse_spec('Assumptions Summary', {'name': 'Worksheet descriptions', 'range': 'B37:E119'})
show_table(assumptions_summary_worksheet_descriptions)

# %% [markdown]
# ### Supporting materials
#
# Lists supporting source materials referenced by the workbook.
#
# Source block: `Assumptions Summary!B121:C147` (27 rows × 2 columns).

# %%
assumptions_summary_supporting_materials = parse_spec('Assumptions Summary', {'name': 'Supporting materials', 'range': 'B121:C147'})
show_table(assumptions_summary_supporting_materials)

# %% [markdown]
# ## Scenarios
#
# Summary of scenario dimensions and parameters.

# %% [markdown]
# ### Scenario parameters
#
# Compares the main parameter settings across the three ISP scenarios.
#
# Source block: `Scenarios!B5:E29` (25 rows × 4 columns).

# %%
scenarios_scenario_parameters = parse_spec('Scenarios', {'name': 'Scenario parameters', 'range': 'B5:E29'})
show_table(scenarios_scenario_parameters)

# %% [markdown]
# ## Existing Gen Data Summary
#
# Summary (calculated) of the generator technical data.

# %% [markdown]
# ### Existing generation data summary
#
# Summarises technical data for existing, committed, anticipated, and additional generators.
#
# Source block: `Existing Gen Data Summary!B10:AT738` (729 rows × 45 columns).

# %%
existing_gen_data_summary_existing_generation_data_summary = parse_spec('Existing Gen Data Summary', {'name': 'Existing generation data summary', 'range': 'B10:AT738'})
show_table(existing_gen_data_summary_existing_generation_data_summary)

# %% [markdown]
# ## New Entrant Data Summary
#
# Summary (calculated) of the new entrant generation and storage technical data.

# %% [markdown]
# ### New entrant data summary
#
# Summarises technical and cost data for new entrant generation and storage technologies.
#
# Source block: `New Entrant Data Summary!B9:BB535` (527 rows × 53 columns).

# %%
new_entrant_data_summary_new_entrant_data_summary = parse_spec('New Entrant Data Summary', {'name': 'New entrant data summary', 'range': 'B9:BB535'})
show_table(new_entrant_data_summary_new_entrant_data_summary)

# %% [markdown]
# ## New Electrolyser Data Summary
#
# Summary (calculated) of the new entrant electrolyser technical data.

# %% [markdown]
# ### New electrolyser data summary
#
# Summarises technical and cost data for new entrant electrolysers.
#
# Source block: `New Electrolyser Data Summary!B5:AQ67` (63 rows × 42 columns).

# %%
new_electrolyser_data_summary_new_electrolyser_data_summary = parse_spec('New Electrolyser Data Summary', {'name': 'New electrolyser data summary', 'range': 'B5:AQ67'})
show_table(new_electrolyser_data_summary_new_electrolyser_data_summary)

# %% [markdown]
# ## Fuel Price Summary
#
# Summary (calculated) of generator fuel costs.

# %% [markdown]
# ### Fuel-price scenario selection
#
# Maps the selected ISP scenario to the fuel-price summary calculations.
#
# Source block: `Fuel Price Summary!B7:S9` (3 rows × 18 columns).

# %%
fuel_price_summary_fuel_price_scenario_selection = parse_spec('Fuel Price Summary', {'name': 'Fuel-price scenario selection', 'range': 'B7:S9'})
show_table(fuel_price_summary_fuel_price_scenario_selection)

# %% [markdown]
# ### Existing generator fuel prices
#
# Summarises fuel prices for existing, committed, anticipated, and additional generators.
#
# Source block: `Fuel Price Summary!B11:AK738` (728 rows × 36 columns).

# %%
fuel_price_summary_existing_generator_fuel_prices = parse_spec('Fuel Price Summary', {'name': 'Existing generator fuel prices', 'range': 'B11:AK738'})
show_table(fuel_price_summary_existing_generator_fuel_prices)

# %% [markdown]
# ### New entrant fuel prices
#
# Summarises fuel prices for new entrant generation technologies.
#
# Source block: `Fuel Price Summary!B743:AK1268` (526 rows × 36 columns).

# %%
fuel_price_summary_new_entrant_fuel_prices = parse_spec('Fuel Price Summary', {'name': 'New entrant fuel prices', 'range': 'B743:AK1268'})
show_table(fuel_price_summary_new_entrant_fuel_prices)

# %% [markdown]
# ## Regional Build Costs Summary
#
# Summary (calculated) of regional build costs for a selectable scenario / technology.

# %% [markdown]
# ### Build-cost selection
#
# Records the scenario and technology controls used by the regional build-cost summary.
#
# Source block: `Regional Build Costs Summary!B7:C10` (4 rows × 2 columns).

# %%
regional_build_costs_summary_build_cost_selection = parse_spec('Regional Build Costs Summary', {'name': 'Build-cost selection', 'range': 'B7:C10'})
show_table(regional_build_costs_summary_build_cost_selection)

# %% [markdown]
# ### Regional build costs
#
# Summarises regional build costs after locational cost factors are applied.
#
# Source block: `Regional Build Costs Summary!B12:AV75` (64 rows × 47 columns).

# %%
regional_build_costs_summary_regional_build_costs = parse_spec('Regional Build Costs Summary', {'name': 'Regional build costs', 'range': 'B12:AV75'})
show_table(regional_build_costs_summary_regional_build_costs)

# %% [markdown]
# ## Energy Policy Targets
#
# Target renewable settings for NEM-wide, Queensland, New South Wales, Victoria, South Australia and Tasmania energy policy targets.

# %% [markdown]
# ### Powering Australia Plan 2030 target
#
# Source block: `Energy Policy Targets!C15:E30` (16 rows × 3 columns).

# %%
energy_policy_targets_powering_australia_plan_2030_target = parse_spec('Energy Policy Targets', {'name': 'Powering Australia Plan 2030 target', 'range': 'C15:E30'})
show_table(energy_policy_targets_powering_australia_plan_2030_target)

# %% [markdown]
# ### Capacity Investment Scheme generation target
#
# Source block: `Energy Policy Targets!C32:F62` (31 rows × 4 columns).

# %%
energy_policy_targets_capacity_investment_scheme_generation_target = parse_spec('Energy Policy Targets', {'name': 'Capacity Investment Scheme generation target', 'range': 'C32:F62'})
show_table(energy_policy_targets_capacity_investment_scheme_generation_target)

# %% [markdown]
# ### Capacity Investment Scheme clean dispatchable capacity target
#
# Source block: `Energy Policy Targets!C65:G93` (29 rows × 5 columns).

# %%
energy_policy_targets_capacity_investment_scheme_clean_dispatchable_capacity_target = parse_spec('Energy Policy Targets', {'name': 'Capacity Investment Scheme clean dispatchable capacity target', 'range': 'C65:G93'})
show_table(energy_policy_targets_capacity_investment_scheme_clean_dispatchable_capacity_target)

# %% [markdown]
# ### Large-scale Renewable Energy Target
#
# Source block: `Energy Policy Targets!C96:E110` (15 rows × 3 columns).

# %%
energy_policy_targets_large_scale_renewable_energy_target = parse_spec('Energy Policy Targets', {'name': 'Large-scale Renewable Energy Target', 'range': 'C96:E110'})
show_table(energy_policy_targets_large_scale_renewable_energy_target)

# %% [markdown]
# ### NSW Electricity Infrastructure Roadmap
#
# Source block: `Energy Policy Targets!C114:N154` (41 rows × 12 columns).

# %%
energy_policy_targets_nsw_electricity_infrastructure_roadmap = parse_spec('Energy Policy Targets', {'name': 'NSW Electricity Infrastructure Roadmap', 'range': 'C114:N154'})
show_table(energy_policy_targets_nsw_electricity_infrastructure_roadmap)

# %% [markdown]
# ### Long-term energy services agreements
#
# Source block: `Energy Policy Targets!C157:E159` (3 rows × 3 columns).

# %%
energy_policy_targets_long_term_energy_services_agreements = parse_spec('Energy Policy Targets', {'name': 'Long-term energy services agreements', 'range': 'C157:E159'})
show_table(energy_policy_targets_long_term_energy_services_agreements)

# %% [markdown]
# ### REZ Access Scheme
#
# Source block: `Energy Policy Targets!C161:I185` (25 rows × 7 columns).

# %%
energy_policy_targets_rez_access_scheme = parse_spec('Energy Policy Targets', {'name': 'REZ Access Scheme', 'range': 'C161:I185'})
show_table(energy_policy_targets_rez_access_scheme)

# %% [markdown]
# ### NSW Roadmap Tender 7 firming
#
# Source block: `Energy Policy Targets!C187:F193` (7 rows × 4 columns).

# %%
energy_policy_targets_nsw_roadmap_tender_7_firming = parse_spec('Energy Policy Targets', {'name': 'NSW Roadmap Tender 7 firming', 'range': 'C187:F193'})
show_table(energy_policy_targets_nsw_roadmap_tender_7_firming)

# %% [markdown]
# ### NSW Renewable Fuels Scheme
#
# Source block: `Energy Policy Targets!C195:E197` (3 rows × 3 columns).

# %%
energy_policy_targets_nsw_renewable_fuels_scheme = parse_spec('Energy Policy Targets', {'name': 'NSW Renewable Fuels Scheme', 'range': 'C195:E197'})
show_table(energy_policy_targets_nsw_renewable_fuels_scheme)

# %% [markdown]
# ### NSW electricity landholder payment scheme
#
# Source block: `Energy Policy Targets!C200:E202` (3 rows × 3 columns).

# %%
energy_policy_targets_nsw_electricity_landholder_payment_scheme = parse_spec('Energy Policy Targets', {'name': 'NSW electricity landholder payment scheme', 'range': 'C200:E202'})
show_table(energy_policy_targets_nsw_electricity_landholder_payment_scheme)

# %% [markdown]
# ### Queensland landholder payment scheme
#
# Source block: `Energy Policy Targets!C206:E209` (4 rows × 3 columns).

# %%
energy_policy_targets_queensland_landholder_payment_scheme = parse_spec('Energy Policy Targets', {'name': 'Queensland landholder payment scheme', 'range': 'C206:E209'})
show_table(energy_policy_targets_queensland_landholder_payment_scheme)

# %% [markdown]
# ### South Australia net renewable energy generation target
#
# Source block: `Energy Policy Targets!C213:E220` (8 rows × 3 columns).

# %%
energy_policy_targets_south_australia_net_renewable_energy_generation_target = parse_spec('Energy Policy Targets', {'name': 'South Australia net renewable energy generation target', 'range': 'C213:E220'})
show_table(energy_policy_targets_south_australia_net_renewable_energy_generation_target)

# %% [markdown]
# ### Firm Energy Reliability Mechanism
#
# Source block: `Energy Policy Targets!C223:E235` (13 rows × 3 columns).

# %%
energy_policy_targets_firm_energy_reliability_mechanism = parse_spec('Energy Policy Targets', {'name': 'Firm Energy Reliability Mechanism', 'range': 'C223:E235'})
show_table(energy_policy_targets_firm_energy_reliability_mechanism)

# %% [markdown]
# ### Tasmania Renewable Energy Target
#
# Source block: `Energy Policy Targets!C239:F258` (20 rows × 4 columns).

# %%
energy_policy_targets_tasmania_renewable_energy_target = parse_spec('Energy Policy Targets', {'name': 'Tasmania Renewable Energy Target', 'range': 'C239:F258'})
show_table(energy_policy_targets_tasmania_renewable_energy_target)

# %% [markdown]
# ### Tasmanian landholder payment scheme
#
# Source block: `Energy Policy Targets!C262:E267` (6 rows × 3 columns).

# %%
energy_policy_targets_tasmanian_landholder_payment_scheme = parse_spec('Energy Policy Targets', {'name': 'Tasmanian landholder payment scheme', 'range': 'C262:E267'})
show_table(energy_policy_targets_tasmanian_landholder_payment_scheme)

# %% [markdown]
# ### Victorian Renewable Energy Target
#
# Source block: `Energy Policy Targets!C271:E289` (19 rows × 3 columns).

# %%
energy_policy_targets_victorian_renewable_energy_target = parse_spec('Energy Policy Targets', {'name': 'Victorian Renewable Energy Target', 'range': 'C271:E289'})
show_table(energy_policy_targets_victorian_renewable_energy_target)

# %% [markdown]
# ### VRET auctions
#
# Source block: `Energy Policy Targets!D291:E295` (5 rows × 2 columns).

# %%
energy_policy_targets_vret_auctions = parse_spec('Energy Policy Targets', {'name': 'VRET auctions', 'range': 'D291:E295'})
show_table(energy_policy_targets_vret_auctions)

# %% [markdown]
# ### Victorian Energy Storage Target
#
# Source block: `Energy Policy Targets!C297:E310` (14 rows × 3 columns).

# %%
energy_policy_targets_victorian_energy_storage_target = parse_spec('Energy Policy Targets', {'name': 'Victorian Energy Storage Target', 'range': 'C297:E310'})
show_table(energy_policy_targets_victorian_energy_storage_target)

# %% [markdown]
# ### Victorian Offshore Wind Target
#
# Source block: `Energy Policy Targets!C313:E327` (15 rows × 3 columns).

# %%
energy_policy_targets_victorian_offshore_wind_target = parse_spec('Energy Policy Targets', {'name': 'Victorian Offshore Wind Target', 'range': 'C313:E327'})
show_table(energy_policy_targets_victorian_offshore_wind_target)

# %% [markdown]
# ## Carbon Budgets
#
# Global mean temperature increases by 2100 aligned with each scenario, and cumulative carbon budgets over the period to 2050.

# %% [markdown]
# ### NEM-wide carbon budgets
#
# Sets cumulative NEM-wide carbon budgets by scenario.
#
# Source block: `Carbon Budgets!B5:E9` (5 rows × 4 columns).

# %%
carbon_budgets_nem_wide_carbon_budgets = parse_spec('Carbon Budgets', {'name': 'NEM-wide carbon budgets', 'range': 'B5:E9'})
show_table(carbon_budgets_nem_wide_carbon_budgets)

# %% [markdown]
# ### State carbon targets
#
# Records jurisdictional carbon targets used by the workbook.
#
# Source block: `Carbon Budgets!B15:G21` (7 rows × 6 columns).

# %%
carbon_budgets_state_carbon_targets = parse_spec('Carbon Budgets', {'name': 'State carbon targets', 'range': 'B15:G21'})
show_table(carbon_budgets_state_carbon_targets)

# %% [markdown]
# ### Converted state carbon budgets
#
# Expresses state carbon targets in the workbook carbon-budget form.
#
# Source block: `Carbon Budgets!B25:D31` (7 rows × 3 columns).

# %%
carbon_budgets_converted_state_carbon_budgets = parse_spec('Carbon Budgets', {'name': 'Converted state carbon budgets', 'range': 'B25:D31'})
show_table(carbon_budgets_converted_state_carbon_budgets)

# %% [markdown]
# ## Economic Growth Forecasts
#
# Forecasts of Gross State Product (GSP) and Household Disposable Income (HDI).

# %% [markdown]
# ### Consultant forecast mapping
#
# Maps ISP scenarios to the consultant economic-growth scenarios.
#
# Source block: `Economic Growth Forecasts!B5:E8` (4 rows × 4 columns).

# %%
economic_growth_forecasts_consultant_forecast_mapping = parse_spec('Economic Growth Forecasts', {'name': 'Consultant forecast mapping', 'range': 'B5:E8'})
show_table(economic_growth_forecasts_consultant_forecast_mapping)

# %% [markdown]
# ### Gross State Product — Slower Growth
#
# Source block: `Economic Growth Forecasts!B12:AG19` (8 rows × 32 columns).

# %%
economic_growth_forecasts_gross_state_product_slower_growth = parse_spec('Economic Growth Forecasts', {'name': 'Gross State Product — Slower Growth', 'range': 'B12:AG19'})
show_table(economic_growth_forecasts_gross_state_product_slower_growth)

# %% [markdown]
# ### Gross State Product — Step Change
#
# Source block: `Economic Growth Forecasts!B21:AG28` (8 rows × 32 columns).

# %%
economic_growth_forecasts_gross_state_product_step_change = parse_spec('Economic Growth Forecasts', {'name': 'Gross State Product — Step Change', 'range': 'B21:AG28'})
show_table(economic_growth_forecasts_gross_state_product_step_change)

# %% [markdown]
# ### Gross State Product — Accelerated Transition
#
# Source block: `Economic Growth Forecasts!B30:AG37` (8 rows × 32 columns).

# %%
economic_growth_forecasts_gross_state_product_accelerated_transition = parse_spec('Economic Growth Forecasts', {'name': 'Gross State Product — Accelerated Transition', 'range': 'B30:AG37'})
show_table(economic_growth_forecasts_gross_state_product_accelerated_transition)

# %% [markdown]
# ### Household Disposable Income — Slower Growth
#
# Source block: `Economic Growth Forecasts!B41:AG48` (8 rows × 32 columns).

# %%
economic_growth_forecasts_household_disposable_income_slower_growth = parse_spec('Economic Growth Forecasts', {'name': 'Household Disposable Income — Slower Growth', 'range': 'B41:AG48'})
show_table(economic_growth_forecasts_household_disposable_income_slower_growth)

# %% [markdown]
# ### Household Disposable Income — Step Change
#
# Source block: `Economic Growth Forecasts!B50:AG57` (8 rows × 32 columns).

# %%
economic_growth_forecasts_household_disposable_income_step_change = parse_spec('Economic Growth Forecasts', {'name': 'Household Disposable Income — Step Change', 'range': 'B50:AG57'})
show_table(economic_growth_forecasts_household_disposable_income_step_change)

# %% [markdown]
# ### Household Disposable Income — Accelerated Transition
#
# Source block: `Economic Growth Forecasts!B59:AG66` (8 rows × 32 columns).

# %%
economic_growth_forecasts_household_disposable_income_accelerated_transition = parse_spec('Economic Growth Forecasts', {'name': 'Household Disposable Income — Accelerated Transition', 'range': 'B59:AG66'})
show_table(economic_growth_forecasts_household_disposable_income_accelerated_transition)

# %% [markdown]
# ## Demand and Energy Forecasts
#
# Points readers to AEMO demand and energy forecasts; it does not embed a forecast data table.
#
# The worksheet contains explanatory text and links to AEMO forecasts rather than an embedded forecast data table.

# %% [markdown]
# ## End use fuel consumption data
#
# Data for end-use fuel consumption by scenario across the NEM chart, identified by multi-sectoral modelling conducted by CSIRO (Figure 1 in IASR).

# %% [markdown]
# ### End-use fuel consumption — Slower Growth
#
# Source block: `End use fuel consumption data!B6:AF15` (10 rows × 31 columns).

# %%
end_use_fuel_consumption_data_end_use_fuel_consumption_slower_growth = parse_spec('End use fuel consumption data', {'name': 'End-use fuel consumption — Slower Growth', 'range': 'B6:AF15'})
show_table(end_use_fuel_consumption_data_end_use_fuel_consumption_slower_growth)

# %% [markdown]
# ### End-use fuel consumption — Step Change
#
# Source block: `End use fuel consumption data!B17:AF26` (10 rows × 31 columns).

# %%
end_use_fuel_consumption_data_end_use_fuel_consumption_step_change = parse_spec('End use fuel consumption data', {'name': 'End-use fuel consumption — Step Change', 'range': 'B17:AF26'})
show_table(end_use_fuel_consumption_data_end_use_fuel_consumption_step_change)

# %% [markdown]
# ### End-use fuel consumption — Accelerated Transition
#
# Source block: `End use fuel consumption data!B28:AF37` (10 rows × 31 columns).

# %%
end_use_fuel_consumption_data_end_use_fuel_consumption_accelerated_transition = parse_spec('End use fuel consumption data', {'name': 'End-use fuel consumption — Accelerated Transition', 'range': 'B28:AF37'})
show_table(end_use_fuel_consumption_data_end_use_fuel_consumption_accelerated_transition)

# %% [markdown]
# ## Appliance Uptake Forecasts
#
# Residential appliance uptake forecasts, impacts relative to base year.

# %% [markdown]
# ### Residential appliance uptake — Slower Growth
#
# Source block: `Appliance Uptake Forecasts!B13:AG20` (8 rows × 32 columns).

# %%
appliance_uptake_forecasts_residential_appliance_uptake_slower_growth = parse_spec('Appliance Uptake Forecasts', {'name': 'Residential appliance uptake — Slower Growth', 'range': 'B13:AG20'})
show_table(appliance_uptake_forecasts_residential_appliance_uptake_slower_growth)

# %% [markdown]
# ### Residential appliance uptake — Step Change
#
# Source block: `Appliance Uptake Forecasts!B22:AG29` (8 rows × 32 columns).

# %%
appliance_uptake_forecasts_residential_appliance_uptake_step_change = parse_spec('Appliance Uptake Forecasts', {'name': 'Residential appliance uptake — Step Change', 'range': 'B22:AG29'})
show_table(appliance_uptake_forecasts_residential_appliance_uptake_step_change)

# %% [markdown]
# ### Residential appliance uptake — Accelerated Transition
#
# Source block: `Appliance Uptake Forecasts!B31:AG38` (8 rows × 32 columns).

# %%
appliance_uptake_forecasts_residential_appliance_uptake_accelerated_transition = parse_spec('Appliance Uptake Forecasts', {'name': 'Residential appliance uptake — Accelerated Transition', 'range': 'B31:AG38'})
show_table(appliance_uptake_forecasts_residential_appliance_uptake_accelerated_transition)

# %% [markdown]
# ## Elec. Retail Price Indices
#
# Provides residential electricity retail price indices relative to the base year.

# %% [markdown]
# ### NEM residential electricity price index
#
# Provides residential retail price indices by ISP scenario.
#
# Source block: `Elec. Retail Price Indices!B8:AG12` (5 rows × 32 columns).

# %%
elec_retail_price_indices_nem_residential_electricity_price_index = parse_spec('Elec. Retail Price Indices', {'name': 'NEM residential electricity price index', 'range': 'B8:AG12'})
show_table(elec_retail_price_indices_nem_residential_electricity_price_index)

# %% [markdown]
# ## Connections Forecasts
#
# Residential connections forecasts.

# %% [markdown]
# ### Consultant forecast mapping
#
# Source block: `Connections Forecasts!B8:E11` (4 rows × 4 columns).

# %%
connections_forecasts_consultant_forecast_mapping = parse_spec('Connections Forecasts', {'name': 'Consultant forecast mapping', 'range': 'B8:E11'})
show_table(connections_forecasts_consultant_forecast_mapping)

# %% [markdown]
# ### Residential connections — Slower Growth
#
# Source block: `Connections Forecasts!B15:AH22` (8 rows × 33 columns).

# %%
connections_forecasts_residential_connections_slower_growth = parse_spec('Connections Forecasts', {'name': 'Residential connections — Slower Growth', 'range': 'B15:AH22'})
show_table(connections_forecasts_residential_connections_slower_growth)

# %% [markdown]
# ### Residential connections — Step Change
#
# Source block: `Connections Forecasts!B24:AH31` (8 rows × 33 columns).

# %%
connections_forecasts_residential_connections_step_change = parse_spec('Connections Forecasts', {'name': 'Residential connections — Step Change', 'range': 'B24:AH31'})
show_table(connections_forecasts_residential_connections_step_change)

# %% [markdown]
# ### Residential connections — Accelerated Transition
#
# Source block: `Connections Forecasts!B33:AH40` (8 rows × 33 columns).

# %%
connections_forecasts_residential_connections_accelerated_transition = parse_spec('Connections Forecasts', {'name': 'Residential connections — Accelerated Transition', 'range': 'B33:AH40'})
show_table(connections_forecasts_residential_connections_accelerated_transition)

# %% [markdown]
# ## Energy Efficiency
#
# Forecast of energy efficiency impact on annual consumption forecasts.

# %% [markdown]
# ### Consultant forecast mapping
#
# Source block: `Energy Efficiency!B8:G10` (3 rows × 6 columns).

# %%
energy_efficiency_consultant_forecast_mapping = parse_spec('Energy Efficiency', {'name': 'Consultant forecast mapping', 'range': 'B8:G10'})
show_table(energy_efficiency_consultant_forecast_mapping)

# %% [markdown]
# ### Residential energy efficiency — Slower Growth
#
# Source block: `Energy Efficiency!B14:AG30` (17 rows × 32 columns).

# %%
energy_efficiency_residential_energy_efficiency_slower_growth = parse_spec('Energy Efficiency', {'name': 'Residential energy efficiency — Slower Growth', 'range': 'B14:AG30'})
show_table(energy_efficiency_residential_energy_efficiency_slower_growth)

# %% [markdown]
# ### Residential energy efficiency — Step Change
#
# Source block: `Energy Efficiency!B32:AG48` (17 rows × 32 columns).

# %%
energy_efficiency_residential_energy_efficiency_step_change = parse_spec('Energy Efficiency', {'name': 'Residential energy efficiency — Step Change', 'range': 'B32:AG48'})
show_table(energy_efficiency_residential_energy_efficiency_step_change)

# %% [markdown]
# ### Residential energy efficiency — Accelerated Transition
#
# Source block: `Energy Efficiency!B50:AG66` (17 rows × 32 columns).

# %%
energy_efficiency_residential_energy_efficiency_accelerated_transition = parse_spec('Energy Efficiency', {'name': 'Residential energy efficiency — Accelerated Transition', 'range': 'B50:AG66'})
show_table(energy_efficiency_residential_energy_efficiency_accelerated_transition)

# %% [markdown]
# ### Residential energy efficiency — lower sensitivity
#
# Source block: `Energy Efficiency!B68:AG84` (17 rows × 32 columns).

# %%
energy_efficiency_residential_energy_efficiency_lower_sensitivity = parse_spec('Energy Efficiency', {'name': 'Residential energy efficiency — lower sensitivity', 'range': 'B68:AG84'})
show_table(energy_efficiency_residential_energy_efficiency_lower_sensitivity)

# %% [markdown]
# ### Residential energy efficiency — higher sensitivity
#
# Source block: `Energy Efficiency!B86:AG102` (17 rows × 32 columns).

# %%
energy_efficiency_residential_energy_efficiency_higher_sensitivity = parse_spec('Energy Efficiency', {'name': 'Residential energy efficiency — higher sensitivity', 'range': 'B86:AG102'})
show_table(energy_efficiency_residential_energy_efficiency_higher_sensitivity)

# %% [markdown]
# ### Business energy efficiency — Slower Growth
#
# Source block: `Energy Efficiency!B106:AG122` (17 rows × 32 columns).

# %%
energy_efficiency_business_energy_efficiency_slower_growth = parse_spec('Energy Efficiency', {'name': 'Business energy efficiency — Slower Growth', 'range': 'B106:AG122'})
show_table(energy_efficiency_business_energy_efficiency_slower_growth)

# %% [markdown]
# ### Business energy efficiency — Step Change
#
# Source block: `Energy Efficiency!B124:AG140` (17 rows × 32 columns).

# %%
energy_efficiency_business_energy_efficiency_step_change = parse_spec('Energy Efficiency', {'name': 'Business energy efficiency — Step Change', 'range': 'B124:AG140'})
show_table(energy_efficiency_business_energy_efficiency_step_change)

# %% [markdown]
# ### Business energy efficiency — Accelerated Transition
#
# Source block: `Energy Efficiency!B142:AG158` (17 rows × 32 columns).

# %%
energy_efficiency_business_energy_efficiency_accelerated_transition = parse_spec('Energy Efficiency', {'name': 'Business energy efficiency — Accelerated Transition', 'range': 'B142:AG158'})
show_table(energy_efficiency_business_energy_efficiency_accelerated_transition)

# %% [markdown]
# ### Business energy efficiency — lower sensitivity
#
# Source block: `Energy Efficiency!B160:AG176` (17 rows × 32 columns).

# %%
energy_efficiency_business_energy_efficiency_lower_sensitivity = parse_spec('Energy Efficiency', {'name': 'Business energy efficiency — lower sensitivity', 'range': 'B160:AG176'})
show_table(energy_efficiency_business_energy_efficiency_lower_sensitivity)

# %% [markdown]
# ### Business energy efficiency — higher sensitivity
#
# Source block: `Energy Efficiency!B178:AG194` (17 rows × 32 columns).

# %%
energy_efficiency_business_energy_efficiency_higher_sensitivity = parse_spec('Energy Efficiency', {'name': 'Business energy efficiency — higher sensitivity', 'range': 'B178:AG194'})
show_table(energy_efficiency_business_energy_efficiency_higher_sensitivity)

# %% [markdown]
# ## Rooftop PV
#
# Rooftop PV capacity and generation forecast.

# %% [markdown]
# ### Consultant forecast mapping
#
# Source block: `Rooftop PV!B8:E10` (3 rows × 4 columns).

# %%
rooftop_pv_consultant_forecast_mapping = parse_spec('Rooftop PV', {'name': 'Consultant forecast mapping', 'range': 'B8:E10'})
show_table(rooftop_pv_consultant_forecast_mapping)

# %% [markdown]
# ### Rooftop PV capacity
#
# Source block: `Rooftop PV!B12:AH63` (52 rows × 33 columns).

# %%
rooftop_pv_rooftop_pv_capacity = parse_spec('Rooftop PV', {'name': 'Rooftop PV capacity', 'range': 'B12:AH63'})
show_table(rooftop_pv_rooftop_pv_capacity)

# %% [markdown]
# ### Rooftop PV energy
#
# Source block: `Rooftop PV!B65:AH116` (52 rows × 33 columns).

# %%
rooftop_pv_rooftop_pv_energy = parse_spec('Rooftop PV', {'name': 'Rooftop PV energy', 'range': 'B65:AH116'})
show_table(rooftop_pv_rooftop_pv_energy)

# %% [markdown]
# ## PVNSG
#
# PV non-scheduled generation (PVNSG) capacity and generation forecast.

# %% [markdown]
# ### Consultant forecast mapping
#
# Source block: `PVNSG!B8:E10` (3 rows × 4 columns).

# %%
pvnsg_consultant_forecast_mapping = parse_spec('PVNSG', {'name': 'Consultant forecast mapping', 'range': 'B8:E10'})
show_table(pvnsg_consultant_forecast_mapping)

# %% [markdown]
# ### PVNSG capacity
#
# Source block: `PVNSG!B12:AH63` (52 rows × 33 columns).

# %%
pvnsg_pvnsg_capacity = parse_spec('PVNSG', {'name': 'PVNSG capacity', 'range': 'B12:AH63'})
show_table(pvnsg_pvnsg_capacity)

# %% [markdown]
# ### PVNSG energy
#
# Source block: `PVNSG!B65:AH116` (52 rows × 33 columns).

# %%
pvnsg_pvnsg_energy = parse_spec('PVNSG', {'name': 'PVNSG energy', 'range': 'B65:AH116'})
show_table(pvnsg_pvnsg_energy)

# %% [markdown]
# ## ONSG
#
# Other non-scheduled generation (ONSG) capacity forecast.

# %% [markdown]
# ### Sub-regional ONSG capacity
#
# Source block: `ONSG!B8:AH55` (48 rows × 33 columns).

# %%
onsg_sub_regional_onsg_capacity = parse_spec('ONSG', {'name': 'Sub-regional ONSG capacity', 'range': 'B8:AH55'})
show_table(onsg_sub_regional_onsg_capacity)

# %% [markdown]
# ### Regional ONSG capacity
#
# Source block: `ONSG!B57:AG74` (18 rows × 32 columns).

# %%
onsg_regional_onsg_capacity = parse_spec('ONSG', {'name': 'Regional ONSG capacity', 'range': 'B57:AG74'})
show_table(onsg_regional_onsg_capacity)

# %% [markdown]
# ## Battery & Plug-in EVs
#
# Battery and plug-in electric vehicles uptake and energy consumption for driving purposes.

# %% [markdown]
# ### Consultant forecast mapping
#
# Source block: `Battery & Plug-in EVs!B7:E9` (3 rows × 4 columns).

# %%
battery_and_plug_in_evs_consultant_forecast_mapping = parse_spec('Battery & Plug-in EVs', {'name': 'Consultant forecast mapping', 'range': 'B7:E9'})
show_table(battery_and_plug_in_evs_consultant_forecast_mapping)

# %% [markdown]
# ### BEV and PHEV energy
#
# Source block: `Battery & Plug-in EVs!B11:AH62` (52 rows × 33 columns).

# %%
battery_and_plug_in_evs_bev_and_phev_energy = parse_spec('Battery & Plug-in EVs', {'name': 'BEV and PHEV energy', 'range': 'B11:AH62'})
show_table(battery_and_plug_in_evs_bev_and_phev_energy)

# %% [markdown]
# ### BEV and PHEV uptake
#
# Source block: `Battery & Plug-in EVs!B64:AH115` (52 rows × 33 columns).

# %%
battery_and_plug_in_evs_bev_and_phev_uptake = parse_spec('Battery & Plug-in EVs', {'name': 'BEV and PHEV uptake', 'range': 'B64:AH115'})
show_table(battery_and_plug_in_evs_bev_and_phev_uptake)

# %% [markdown]
# ## Fuel cell EVs
#
# Fuel cell electric vehicles uptake.

# %% [markdown]
# ### Consultant forecast mapping
#
# Source block: `Fuel cell EVs!B7:E9` (3 rows × 4 columns).

# %%
fuel_cell_evs_consultant_forecast_mapping = parse_spec('Fuel cell EVs', {'name': 'Consultant forecast mapping', 'range': 'B7:E9'})
show_table(fuel_cell_evs_consultant_forecast_mapping)

# %% [markdown]
# ### Fuel-cell EV uptake
#
# Source block: `Fuel cell EVs!B11:AH62` (52 rows × 33 columns).

# %%
fuel_cell_evs_fuel_cell_ev_uptake = parse_spec('Fuel cell EVs', {'name': 'Fuel-cell EV uptake', 'range': 'B11:AH62'})
show_table(fuel_cell_evs_fuel_cell_ev_uptake)

# %% [markdown]
# ## EV V2G
#
# Vehicle to Grid battery characteristics.

# %% [markdown]
# ### Consultant forecast mapping
#
# Source block: `EV V2G!B8:E10` (3 rows × 4 columns).

# %%
ev_v2g_consultant_forecast_mapping = parse_spec('EV V2G', {'name': 'Consultant forecast mapping', 'range': 'B8:E10'})
show_table(ev_v2g_consultant_forecast_mapping)

# %% [markdown]
# ### Vehicle-to-grid capacity
#
# Source block: `EV V2G!B12:AH62` (51 rows × 33 columns).

# %%
ev_v2g_vehicle_to_grid_capacity = parse_spec('EV V2G', {'name': 'Vehicle-to-grid capacity', 'range': 'B12:AH62'})
show_table(ev_v2g_vehicle_to_grid_capacity)

# %% [markdown]
# ### Vehicle-to-grid depth
#
# Source block: `EV V2G!B64:AH115` (52 rows × 33 columns).

# %%
ev_v2g_vehicle_to_grid_depth = parse_spec('EV V2G', {'name': 'Vehicle-to-grid depth', 'range': 'B64:AH115'})
show_table(ev_v2g_vehicle_to_grid_depth)

# %% [markdown]
# ## Data Centre Forecasts
#
# Forecast of electricity consumption from data centre growth.

# %% [markdown]
# ### Consultant forecast mapping
#
# Source block: `Data Centre Forecasts!B6:E8` (3 rows × 4 columns).

# %%
data_centre_forecasts_consultant_forecast_mapping = parse_spec('Data Centre Forecasts', {'name': 'Consultant forecast mapping', 'range': 'B6:E8'})
show_table(data_centre_forecasts_consultant_forecast_mapping)

# %% [markdown]
# ### Data-centre demand — Slower Growth
#
# Source block: `Data Centre Forecasts!B10:AF16` (7 rows × 31 columns).

# %%
data_centre_forecasts_data_centre_demand_slower_growth = parse_spec('Data Centre Forecasts', {'name': 'Data-centre demand — Slower Growth', 'range': 'B10:AF16'})
show_table(data_centre_forecasts_data_centre_demand_slower_growth)

# %% [markdown]
# ### Data-centre demand — Step Change
#
# Source block: `Data Centre Forecasts!B18:AF24` (7 rows × 31 columns).

# %%
data_centre_forecasts_data_centre_demand_step_change = parse_spec('Data Centre Forecasts', {'name': 'Data-centre demand — Step Change', 'range': 'B18:AF24'})
show_table(data_centre_forecasts_data_centre_demand_step_change)

# %% [markdown]
# ### Data-centre demand — Accelerated Transition
#
# Source block: `Data Centre Forecasts!B26:AF32` (7 rows × 31 columns).

# %%
data_centre_forecasts_data_centre_demand_accelerated_transition = parse_spec('Data Centre Forecasts', {'name': 'Data-centre demand — Accelerated Transition', 'range': 'B26:AF32'})
show_table(data_centre_forecasts_data_centre_demand_accelerated_transition)

# %% [markdown]
# ## DSP
#
# Demand side participation forecast.

# %% [markdown]
# ### Summer demand-side participation
#
# Source block: `DSP!B7:AI84` (78 rows × 34 columns).

# %%
dsp_summer_demand_side_participation = parse_spec('DSP', {'name': 'Summer demand-side participation', 'range': 'B7:AI84'})
show_table(dsp_summer_demand_side_participation)

# %% [markdown]
# ### Winter demand-side participation
#
# Source block: `DSP!B87:AI164` (78 rows × 34 columns).

# %%
dsp_winter_demand_side_participation = parse_spec('DSP', {'name': 'Winter demand-side participation', 'range': 'B87:AI164'})
show_table(dsp_winter_demand_side_participation)

# %% [markdown]
# ## Electrification
#
# Electrification in all sectors excluding road transportation.

# %% [markdown]
# ### Consultant forecast mapping
#
# Source block: `Electrification!B7:E10` (4 rows × 4 columns).

# %%
electrification_consultant_forecast_mapping = parse_spec('Electrification', {'name': 'Consultant forecast mapping', 'range': 'B7:E10'})
show_table(electrification_consultant_forecast_mapping)

# %% [markdown]
# ### Electrification — Slower Growth
#
# Source block: `Electrification!B12:AF19` (8 rows × 31 columns).

# %%
electrification_electrification_slower_growth = parse_spec('Electrification', {'name': 'Electrification — Slower Growth', 'range': 'B12:AF19'})
show_table(electrification_electrification_slower_growth)

# %% [markdown]
# ### Electrification — Step Change
#
# Source block: `Electrification!B21:AF28` (8 rows × 31 columns).

# %%
electrification_electrification_step_change = parse_spec('Electrification', {'name': 'Electrification — Step Change', 'range': 'B21:AF28'})
show_table(electrification_electrification_step_change)

# %% [markdown]
# ### Electrification — Accelerated Transition
#
# Source block: `Electrification!B30:AF37` (8 rows × 31 columns).

# %%
electrification_electrification_accelerated_transition = parse_spec('Electrification', {'name': 'Electrification — Accelerated Transition', 'range': 'B30:AF37'})
show_table(electrification_electrification_accelerated_transition)

# %% [markdown]
# ## Embedded energy storages
#
# Embedded consumer energy storage (battery) forecast.

# %% [markdown]
# ### Forecast mapping
#
# Source block: `Embedded energy storages!B7:E9` (3 rows × 4 columns).

# %%
embedded_energy_storages_forecast_mapping = parse_spec('Embedded energy storages', {'name': 'Forecast mapping', 'range': 'B7:E9'})
show_table(embedded_energy_storages_forecast_mapping)

# %% [markdown]
# ### Embedded energy storage capacity
#
# Source block: `Embedded energy storages!B11:AH62` (52 rows × 33 columns).

# %%
embedded_energy_storages_embedded_energy_storage_capacity = parse_spec('Embedded energy storages', {'name': 'Embedded energy storage capacity', 'range': 'B11:AH62'})
show_table(embedded_energy_storages_embedded_energy_storage_capacity)

# %% [markdown]
# ### Embedded energy storage degraded energy
#
# Source block: `Embedded energy storages!B65:AH116` (52 rows × 33 columns).

# %%
embedded_energy_storages_embedded_energy_storage_degraded_energy = parse_spec('Embedded energy storages', {'name': 'Embedded energy storage degraded energy', 'range': 'B65:AH116'})
show_table(embedded_energy_storages_embedded_energy_storage_degraded_energy)

# %% [markdown]
# ## Aggregated energy storages
#
# The 'aggregated' share of embedded energy storages that is modelled like a Virtual Power Plant (VPP).

# %% [markdown]
# ### Forecast mapping
#
# Source block: `Aggregated energy storages!B7:E9` (3 rows × 4 columns).

# %%
aggregated_energy_storages_forecast_mapping = parse_spec('Aggregated energy storages', {'name': 'Forecast mapping', 'range': 'B7:E9'})
show_table(aggregated_energy_storages_forecast_mapping)

# %% [markdown]
# ### Aggregated energy storage capacity
#
# Source block: `Aggregated energy storages!B11:AH62` (52 rows × 33 columns).

# %%
aggregated_energy_storages_aggregated_energy_storage_capacity = parse_spec('Aggregated energy storages', {'name': 'Aggregated energy storage capacity', 'range': 'B11:AH62'})
show_table(aggregated_energy_storages_aggregated_energy_storage_capacity)

# %% [markdown]
# ### Aggregated energy storage degraded energy
#
# Source block: `Aggregated energy storages!B65:AH116` (52 rows × 33 columns).

# %%
aggregated_energy_storages_aggregated_energy_storage_degraded_energy = parse_spec('Aggregated energy storages', {'name': 'Aggregated energy storage degraded energy', 'range': 'B65:AH116'})
show_table(aggregated_energy_storages_aggregated_energy_storage_degraded_energy)

# %% [markdown]
# ## Network representation
#
# Description of how the network is modelled in the capacity expansion models.

# %% [markdown]
# ### Sub-regional flow-path representation
#
# Source block: `Network representation!B2:E22` (21 rows × 4 columns).

# %%
network_representation_sub_regional_flow_path_representation = parse_spec('Network representation', {'name': 'Sub-regional flow-path representation', 'range': 'B2:E22'})
show_table(network_representation_sub_regional_flow_path_representation)

# %% [markdown]
# ### Sub-regional reference nodes
#
# Source block: `Network representation!B24:D42` (19 rows × 3 columns).

# %%
network_representation_sub_regional_reference_nodes = parse_spec('Network representation', {'name': 'Sub-regional reference nodes', 'range': 'B24:D42'})
show_table(network_representation_sub_regional_reference_nodes)

# %% [markdown]
# ### Regional topology representation
#
# Source block: `Network representation!B44:D53` (10 rows × 3 columns).

# %%
network_representation_regional_topology_representation = parse_spec('Network representation', {'name': 'Regional topology representation', 'range': 'B44:D53'})
show_table(network_representation_regional_topology_representation)

# %% [markdown]
# ### Regional reference nodes
#
# Source block: `Network representation!B55:D63` (9 rows × 3 columns).

# %%
network_representation_regional_reference_nodes = parse_spec('Network representation', {'name': 'Regional reference nodes', 'range': 'B55:D63'})
show_table(network_representation_regional_reference_nodes)

# %% [markdown]
# ### Sub-regional load and generation representation
#
# Source block: `Network representation!B65:D82` (18 rows × 3 columns).

# %%
network_representation_sub_regional_load_and_generation_representation = parse_spec('Network representation', {'name': 'Sub-regional load and generation representation', 'range': 'B65:D82'})
show_table(network_representation_sub_regional_load_and_generation_representation)

# %% [markdown]
# ## Renewable energy zones
#
# Renewable energy zones.

# %% [markdown]
# ### Candidate renewable energy zones
#
# Lists candidate REZ identifiers, names, NEM regions, and ISP sub-regions.
#
# Source block: `Renewable energy zones!B6:E53` (48 rows × 4 columns).

# %%
renewable_energy_zones_candidate_renewable_energy_zones = parse_spec('Renewable energy zones', {'name': 'Candidate renewable energy zones', 'range': 'B6:E53'})
show_table(renewable_energy_zones_candidate_renewable_energy_zones)

# %% [markdown]
# ## Network capability
#
# Maximum forward and reverse flow path capability for capacity expansion modelling.

# %% [markdown]
# ### Flow-path transfer capability
#
# Contains the 18 verified flow-path capability data rows; workbook headers are in rows 6–7.
#
# Source block: `Network capability!B8:K25` (18 rows × 10 columns).

# %%
network_capability_flow_path_transfer_capability = parse_spec('Network capability', {'name': 'Flow-path transfer capability', 'range': 'B8:K25', 'expected_rows': 18, 'expected_cols': 10})
show_table(network_capability_flow_path_transfer_capability)

# %% [markdown]
# ### Interconnector transfer capability
#
# Source block: `Network capability!B34:K42` (9 rows × 10 columns).

# %%
network_capability_interconnector_transfer_capability = parse_spec('Network capability', {'name': 'Interconnector transfer capability', 'range': 'B34:K42'})
show_table(network_capability_interconnector_transfer_capability)

# %% [markdown]
# ### Committed-project transfer capability uplift
#
# Source block: `Network capability!B51:N60` (10 rows × 13 columns).

# %%
network_capability_committed_project_transfer_capability_uplift = parse_spec('Network capability', {'name': 'Committed-project transfer capability uplift', 'range': 'B51:N60'})
show_table(network_capability_committed_project_transfer_capability_uplift)

# %% [markdown]
# ### Sydney Ring generator coefficients
#
# Source block: `Network capability!B75:V84` (10 rows × 21 columns).

# %%
network_capability_sydney_ring_generator_coefficients = parse_spec('Network capability', {'name': 'Sydney Ring generator coefficients', 'range': 'B75:V84'})
show_table(network_capability_sydney_ring_generator_coefficients)

# %% [markdown]
# ### Reference temperatures
#
# Source block: `Network capability!B89:E94` (6 rows × 4 columns).

# %%
network_capability_reference_temperatures = parse_spec('Network capability', {'name': 'Reference temperatures', 'range': 'B89:E94'})
show_table(network_capability_reference_temperatures)

# %% [markdown]
# ### Murraylink dynamic temperature-dependent transfer capability
#
# Source block: `Network capability!B99:D115` (17 rows × 3 columns).

# %%
network_capability_murraylink_dynamic_temperature_dependent_transfer_capability = parse_spec('Network capability', {'name': 'Murraylink dynamic temperature-dependent transfer capability', 'range': 'B99:D115'})
show_table(network_capability_murraylink_dynamic_temperature_dependent_transfer_capability)

# %% [markdown]
# ### Basslink static daily energy throughput limit
#
# Source block: `Network capability!B122:C134` (13 rows × 2 columns).

# %%
network_capability_basslink_static_daily_energy_throughput_limit = parse_spec('Network capability', {'name': 'Basslink static daily energy throughput limit', 'range': 'B122:C134'})
show_table(network_capability_basslink_static_daily_energy_throughput_limit)

# %% [markdown]
# ### Committed and anticipated project timing
#
# Source block: `Network capability!B139:C148` (10 rows × 2 columns).

# %%
network_capability_committed_and_anticipated_project_timing = parse_spec('Network capability', {'name': 'Committed and anticipated project timing', 'range': 'B139:C148'})
show_table(network_capability_committed_and_anticipated_project_timing)

# %% [markdown]
# ## Network losses
#
# Proportion of interconnector losses applied to regional reference nodes and loss equations.

# %% [markdown]
# ### Existing flow-path loss equations
#
# Source block: `Network losses!B5:J28` (24 rows × 9 columns).

# %%
network_losses_existing_flow_path_loss_equations = parse_spec('Network losses', {'name': 'Existing flow-path loss equations', 'range': 'B5:J28'})
show_table(network_losses_existing_flow_path_loss_equations)

# %% [markdown]
# ### Committed and anticipated project loss equations
#
# Source block: `Network losses!B30:J34` (5 rows × 9 columns).

# %%
network_losses_committed_and_anticipated_project_loss_equations = parse_spec('Network losses', {'name': 'Committed and anticipated project loss equations', 'range': 'B30:J34'})
show_table(network_losses_committed_and_anticipated_project_loss_equations)

# %% [markdown]
# ### Development-option loss equations
#
# Source block: `Network losses!B36:J88` (53 rows × 9 columns).

# %%
network_losses_development_option_loss_equations = parse_spec('Network losses', {'name': 'Development-option loss equations', 'range': 'B36:J88'})
show_table(network_losses_development_option_loss_equations)

# %% [markdown]
# ## Transmission Reliability
#
# Defines the outage rates modelled to key flowpaths in the ESOO.

# %% [markdown]
# ### Transmission unplanned outage rates
#
# Contains the verified header row 7 and data rows 8–13.
#
# Source block: `Transmission Reliability!B7:E13` (7 rows × 4 columns).

# %%
transmission_reliability_transmission_unplanned_outage_rates = parse_spec('Transmission Reliability', {'name': 'Transmission unplanned outage rates', 'range': 'B7:E13', 'expected_rows': 7, 'expected_cols': 4})
show_table(transmission_reliability_transmission_unplanned_outage_rates)

# %% [markdown]
# ## Distribution network
#
# Inputs used to model distribution network opportunities to facilitate aggregate operation of consumer energy resources and other distributed resources.

# %% [markdown]
# ### Mid-scale generation and storage build limits
#
# Source block: `Distribution network!B11:G38` (28 rows × 6 columns).

# %%
distribution_network_mid_scale_generation_and_storage_build_limits = parse_spec('Distribution network', {'name': 'Mid-scale generation and storage build limits', 'range': 'B11:G38'})
show_table(distribution_network_mid_scale_generation_and_storage_build_limits)

# %% [markdown]
# ### Distribution CER augmentation tranche costs
#
# Source block: `Distribution network!B40:H57` (18 rows × 7 columns).

# %%
distribution_network_distribution_cer_augmentation_tranche_costs = parse_spec('Distribution network', {'name': 'Distribution CER augmentation tranche costs', 'range': 'B40:H57'})
show_table(distribution_network_distribution_cer_augmentation_tranche_costs)

# %% [markdown]
# ### Average CER generation-limit time-of-day profile
#
# Source block: `Distribution network!B59:AZ1433` (1375 rows × 51 columns).

# %%
distribution_network_average_cer_generation_limit_time_of_day_profile = parse_spec('Distribution network', {'name': 'Average CER generation-limit time-of-day profile', 'range': 'B59:AZ1433'})
show_table(distribution_network_average_cer_generation_limit_time_of_day_profile)

# %% [markdown]
# ## Connection cost
#
# Cost to connect different generation technologies.

# %% [markdown]
# ### Wind and solar connection costs
#
# Source block: `Connection cost!B6:J61` (56 rows × 9 columns).

# %%
connection_cost_wind_and_solar_connection_costs = parse_spec('Connection cost', {'name': 'Wind and solar connection costs', 'range': 'B6:J61'})
show_table(connection_cost_wind_and_solar_connection_costs)

# %% [markdown]
# ### Other-generation regional connection costs
#
# Source block: `Connection cost!B62:R73` (12 rows × 17 columns).

# %%
connection_cost_other_generation_regional_connection_costs = parse_spec('Connection cost', {'name': 'Other-generation regional connection costs', 'range': 'B62:R73'})
show_table(connection_cost_other_generation_regional_connection_costs)

# %% [markdown]
# ## Connection cost forecasts
#
# Forecast of transmission connection costs.

# %% [markdown]
# ### Wind and solar connection-cost forecasts
#
# Source block: `Connection cost forecasts!B8:AJ144` (137 rows × 35 columns).

# %%
connection_cost_forecasts_wind_and_solar_connection_cost_forecasts = parse_spec('Connection cost forecasts', {'name': 'Wind and solar connection-cost forecasts', 'range': 'B8:AJ144'})
show_table(connection_cost_forecasts_wind_and_solar_connection_cost_forecasts)

# %% [markdown]
# ### Other-generation connection-cost forecasts
#
# Source block: `Connection cost forecasts!B147:AJ388` (242 rows × 35 columns).

# %%
connection_cost_forecasts_other_generation_connection_cost_forecasts = parse_spec('Connection cost forecasts', {'name': 'Other-generation connection-cost forecasts', 'range': 'B147:AJ388'})
show_table(connection_cost_forecasts_other_generation_connection_cost_forecasts)

# %% [markdown]
# ## Flow path augmentation options
#
# Capability, cost and timing for flow path augmentation options.

# %% [markdown]
# ### Flow-path augmentation options
#
# Combines repeated physical flow-path sections into one logical option dataset.
#
# Source block: `Flow path augmentation options!B11:Q127` (117 rows × 16 columns).

# %%
flow_path_augmentation_options_flow_path_augmentation_options = parse_spec('Flow path augmentation options', {'name': 'Flow-path augmentation options',
 'range': 'B11:Q127',
 'parser': 'flow_path_options',
 'expected_semantic_rows': 62})
show_table(flow_path_augmentation_options_flow_path_augmentation_options)

# %% [markdown]
# ## Flow path cost forecasts
#
# Forecast of flow path augmentation costs.

# %% [markdown]
# ### Flow-path cost forecast — Slower Growth
#
# Source block: `Flow path cost forecasts!B10:AI111` (102 rows × 34 columns).

# %%
flow_path_cost_forecasts_flow_path_cost_forecast_slower_growth = parse_spec('Flow path cost forecasts', {'name': 'Flow-path cost forecast — Slower Growth', 'range': 'B10:AI111'})
show_table(flow_path_cost_forecasts_flow_path_cost_forecast_slower_growth)

# %% [markdown]
# ### Flow-path cost forecast — Step Change
#
# Source block: `Flow path cost forecasts!B115:AI216` (102 rows × 34 columns).

# %%
flow_path_cost_forecasts_flow_path_cost_forecast_step_change = parse_spec('Flow path cost forecasts', {'name': 'Flow-path cost forecast — Step Change', 'range': 'B115:AI216'})
show_table(flow_path_cost_forecasts_flow_path_cost_forecast_step_change)

# %% [markdown]
# ### Flow-path cost forecast — Accelerated Transition
#
# Source block: `Flow path cost forecasts!B220:AI321` (102 rows × 34 columns).

# %%
flow_path_cost_forecasts_flow_path_cost_forecast_accelerated_transition = parse_spec('Flow path cost forecasts', {'name': 'Flow-path cost forecast — Accelerated Transition', 'range': 'B220:AI321'})
show_table(flow_path_cost_forecasts_flow_path_cost_forecast_accelerated_transition)

# %% [markdown]
# ## REZ augmentations options
#
# Capability, cost and timing for REZ augmentation options.

# %% [markdown]
# ### Queensland REZ augmentation options
#
# Source block: `REZ augmentations options!B10:O37` (28 rows × 14 columns).

# %%
rez_augmentations_options_queensland_rez_augmentation_options = parse_spec('REZ augmentations options', {'name': 'Queensland REZ augmentation options', 'range': 'B10:O37'})
show_table(rez_augmentations_options_queensland_rez_augmentation_options)

# %% [markdown]
# ### New South Wales REZ augmentation options
#
# Source block: `REZ augmentations options!B39:O77` (39 rows × 14 columns).

# %%
rez_augmentations_options_new_south_wales_rez_augmentation_options = parse_spec('REZ augmentations options', {'name': 'New South Wales REZ augmentation options', 'range': 'B39:O77'})
show_table(rez_augmentations_options_new_south_wales_rez_augmentation_options)

# %% [markdown]
# ### South Australia REZ augmentation options
#
# Source block: `REZ augmentations options!B79:O96` (18 rows × 14 columns).

# %%
rez_augmentations_options_south_australia_rez_augmentation_options = parse_spec('REZ augmentations options', {'name': 'South Australia REZ augmentation options', 'range': 'B79:O96'})
show_table(rez_augmentations_options_south_australia_rez_augmentation_options)

# %% [markdown]
# ### Tasmania REZ augmentation options
#
# Source block: `REZ augmentations options!B98:O110` (13 rows × 14 columns).

# %%
rez_augmentations_options_tasmania_rez_augmentation_options = parse_spec('REZ augmentations options', {'name': 'Tasmania REZ augmentation options', 'range': 'B98:O110'})
show_table(rez_augmentations_options_tasmania_rez_augmentation_options)

# %% [markdown]
# ### Victoria REZ augmentation options
#
# Source block: `REZ augmentations options!B112:O137` (26 rows × 14 columns).

# %%
rez_augmentations_options_victoria_rez_augmentation_options = parse_spec('REZ augmentations options', {'name': 'Victoria REZ augmentation options', 'range': 'B112:O137'})
show_table(rez_augmentations_options_victoria_rez_augmentation_options)

# %% [markdown]
# ## REZ cost forecasts
#
# Forecast of REZ augmentation costs.

# %% [markdown]
# ### REZ cost forecast — Slower Growth
#
# Source block: `REZ cost forecasts!B11:AJ117` (107 rows × 35 columns).

# %%
rez_cost_forecasts_rez_cost_forecast_slower_growth = parse_spec('REZ cost forecasts', {'name': 'REZ cost forecast — Slower Growth', 'range': 'B11:AJ117'})
show_table(rez_cost_forecasts_rez_cost_forecast_slower_growth)

# %% [markdown]
# ### REZ cost forecast — Step Change
#
# Source block: `REZ cost forecasts!B118:AJ224` (107 rows × 35 columns).

# %%
rez_cost_forecasts_rez_cost_forecast_step_change = parse_spec('REZ cost forecasts', {'name': 'REZ cost forecast — Step Change', 'range': 'B118:AJ224'})
show_table(rez_cost_forecasts_rez_cost_forecast_step_change)

# %% [markdown]
# ### REZ cost forecast — Accelerated Transition
#
# Source block: `REZ cost forecasts!B225:AJ331` (107 rows × 35 columns).

# %%
rez_cost_forecasts_rez_cost_forecast_accelerated_transition = parse_spec('REZ cost forecasts', {'name': 'REZ cost forecast — Accelerated Transition', 'range': 'B225:AJ331'})
show_table(rez_cost_forecasts_rez_cost_forecast_accelerated_transition)

# %% [markdown]
# ## Distribution cost forecasts
#
# Forecast of the distribution augmentation tranche costs.

# %% [markdown]
# ### Distribution CER augmentation cost forecast
#
# Source block: `Distribution cost forecasts!B5:AJ84` (80 rows × 35 columns).

# %%
distribution_cost_forecasts_distribution_cer_augmentation_cost_forecast = parse_spec('Distribution cost forecasts', {'name': 'Distribution CER augmentation cost forecast', 'range': 'B5:AJ84'})
show_table(distribution_cost_forecasts_distribution_cer_augmentation_cost_forecast)

# %% [markdown]
# ## Maximum capacity
#
# Installed capacity of existing, committed and anticipated generators.

# %% [markdown]
# ### Existing, committed, anticipated, and additional generator capacity
#
# Source block: `Maximum capacity!B9:J750` (742 rows × 9 columns).

# %%
maximum_capacity_existing_committed_anticipated_and_additional_generator_capacity = parse_spec('Maximum capacity', {'name': 'Existing, committed, anticipated, and additional generator capacity', 'range': 'B9:J750'})
show_table(maximum_capacity_existing_committed_anticipated_and_additional_generator_capacity)

# %% [markdown]
# ### New generation technology capacity
#
# Source block: `Maximum capacity!L9:O31` (23 rows × 4 columns).

# %%
maximum_capacity_new_generation_technology_capacity = parse_spec('Maximum capacity', {'name': 'New generation technology capacity', 'range': 'L9:O31'})
show_table(maximum_capacity_new_generation_technology_capacity)

# %% [markdown]
# ## Hybrid site limits
#
# Interval-level charging/dispatch limits for sites with a combination of VRE and battery storage at one connection point.

# %% [markdown]
# ### Hybrid-site limits
#
# Contains the verified hybrid-site limit source table.
#
# Source block: `Hybrid site limits!B9:G67` (59 rows × 6 columns).

# %%
hybrid_site_limits_hybrid_site_limits = parse_spec('Hybrid site limits', {'name': 'Hybrid-site limits', 'range': 'B9:G67'})
show_table(hybrid_site_limits_hybrid_site_limits)

# %% [markdown]
# ## Seasonal ratings
#
# Winter and summer ratings for existing, committed, advanced and new entrant generators.

# %% [markdown]
# ### New generation technology seasonal ratings
#
# Source block: `Seasonal ratings!B9:E36` (28 rows × 4 columns).

# %%
seasonal_ratings_new_generation_technology_seasonal_ratings = parse_spec('Seasonal ratings', {'name': 'New generation technology seasonal ratings', 'range': 'B9:E36'})
show_table(seasonal_ratings_new_generation_technology_seasonal_ratings)

# %% [markdown]
# ### Existing, committed, anticipated, and additional generator seasonal ratings
#
# Source block: `Seasonal ratings!B42:AI770` (729 rows × 34 columns).

# %%
seasonal_ratings_existing_committed_anticipated_and_additional_generator_seasonal_ratings = parse_spec('Seasonal ratings', {'name': 'Existing, committed, anticipated, and additional generator seasonal ratings', 'range': 'B42:AI770'})
show_table(seasonal_ratings_existing_committed_anticipated_and_additional_generator_seasonal_ratings)

# %% [markdown]
# ## Generator Reliability Settings
#
# Defines a generators' unplanned outage rate, mean time to repair after an outage, and the derating experienced during a partial outage.

# %% [markdown]
# ### Existing generator long-duration outages
#
# Source block: `Generator Reliability Settings!B9:M16` (8 rows × 12 columns).

# %%
generator_reliability_settings_existing_generator_long_duration_outages = parse_spec('Generator Reliability Settings', {'name': 'Existing generator long-duration outages', 'range': 'B9:M16'})
show_table(generator_reliability_settings_existing_generator_long_duration_outages)

# %% [markdown]
# ### Existing generator outage rates and MTTR
#
# Source block: `Generator Reliability Settings!B21:M60` (40 rows × 12 columns).

# %%
generator_reliability_settings_existing_generator_outage_rates_and_mttr = parse_spec('Generator Reliability Settings', {'name': 'Existing generator outage rates and MTTR', 'range': 'B21:M60'})
show_table(generator_reliability_settings_existing_generator_outage_rates_and_mttr)

# %% [markdown]
# ### New entrant reliability settings
#
# Source block: `Generator Reliability Settings!B62:H90` (29 rows × 7 columns).

# %%
generator_reliability_settings_new_entrant_reliability_settings = parse_spec('Generator Reliability Settings', {'name': 'New entrant reliability settings', 'range': 'B62:H90'})
show_table(generator_reliability_settings_new_entrant_reliability_settings)

# %% [markdown]
# ## Maintenance
#
# The percentage of time per year that a generator is expected to be out of service for maintenance. De-rating applied to generators under maintenance, staged construction or to track age-related degradation.

# %% [markdown]
# ### Existing generator maintenance rates
#
# Source block: `Maintenance!B5:D29` (25 rows × 3 columns).

# %%
maintenance_existing_generator_maintenance_rates = parse_spec('Maintenance', {'name': 'Existing generator maintenance rates', 'range': 'B5:D29'})
show_table(maintenance_existing_generator_maintenance_rates)

# %% [markdown]
# ### New entrant maintenance rates
#
# Source block: `Maintenance!G5:I32` (28 rows × 3 columns).

# %%
maintenance_new_entrant_maintenance_rates = parse_spec('Maintenance', {'name': 'New entrant maintenance rates', 'range': 'G5:I32'})
show_table(maintenance_new_entrant_maintenance_rates)

# %% [markdown]
# ## Retirement
#
# Announced and end-of-technical-life generator retirement.

# %% [markdown]
# ### Expected generator closure years
#
# Source block: `Retirement!B8:F738` (731 rows × 5 columns).

# %%
retirement_expected_generator_closure_years = parse_spec('Retirement', {'name': 'Expected generator closure years', 'range': 'B8:F738'})
show_table(retirement_expected_generator_closure_years)

# %% [markdown]
# ### Generator retirement costs
#
# Source block: `Retirement!H8:I50` (43 rows × 2 columns).

# %%
retirement_generator_retirement_costs = parse_spec('Retirement', {'name': 'Generator retirement costs', 'range': 'H8:I50'})
show_table(retirement_generator_retirement_costs)

# %% [markdown]
# ## Hydro Scheme Inflows
#
# Monthly aggregated inflow trends for reference years.

# %% [markdown]
# ### Secondary hydro scheme releases and outflows
#
# Source block: `Hydro Scheme Inflows!B4:T79` (76 rows × 19 columns).

# %%
hydro_scheme_inflows_secondary_hydro_scheme_releases_and_outflows = parse_spec('Hydro Scheme Inflows', {'name': 'Secondary hydro scheme releases and outflows', 'range': 'B4:T79'})
show_table(hydro_scheme_inflows_secondary_hydro_scheme_releases_and_outflows)

# %% [markdown]
# ### Run-of-river hydro outflows
#
# Source block: `Hydro Scheme Inflows!B81:T121` (41 rows × 19 columns).

# %%
hydro_scheme_inflows_run_of_river_hydro_outflows = parse_spec('Hydro Scheme Inflows', {'name': 'Run-of-river hydro outflows', 'range': 'B81:T121'})
show_table(hydro_scheme_inflows_run_of_river_hydro_outflows)

# %% [markdown]
# ### Hydro Tasmania scheme
#
# Source block: `Hydro Scheme Inflows!B123:T141` (19 rows × 19 columns).

# %%
hydro_scheme_inflows_hydro_tasmania_scheme = parse_spec('Hydro Scheme Inflows', {'name': 'Hydro Tasmania scheme', 'range': 'B123:T141'})
show_table(hydro_scheme_inflows_hydro_tasmania_scheme)

# %% [markdown]
# ### Snowy Hydro weather-variability representation
#
# Source block: `Hydro Scheme Inflows!B143:T162` (20 rows × 19 columns).

# %%
hydro_scheme_inflows_snowy_hydro_weather_variability_representation = parse_spec('Hydro Scheme Inflows', {'name': 'Snowy Hydro weather-variability representation', 'range': 'B143:T162'})
show_table(hydro_scheme_inflows_snowy_hydro_weather_variability_representation)

# %% [markdown]
# ## Capacity Factors
#
# Capacity factors for renewable generators in the renewable energy zones.

# %% [markdown]
# ### New large-scale renewable capacity factors
#
# Source block: `Capacity Factors !B2:V214` (213 rows × 21 columns).

# %%
capacity_factors_new_large_scale_renewable_capacity_factors = parse_spec('Capacity Factors ', {'name': 'New large-scale renewable capacity factors', 'range': 'B2:V214'})
show_table(capacity_factors_new_large_scale_renewable_capacity_factors)

# %% [markdown]
# ## Heat rates
#
# Efficiency of conversion of fuel to output for thermal generators.

# %% [markdown]
# ### Existing generator heat rates
#
# Source block: `Heat rates!B7:E740` (734 rows × 4 columns).

# %%
heat_rates_existing_generator_heat_rates = parse_spec('Heat rates', {'name': 'Existing generator heat rates', 'range': 'B7:E740'})
show_table(heat_rates_existing_generator_heat_rates)

# %% [markdown]
# ### New entrant heat rates
#
# Source block: `Heat rates!H7:I31` (25 rows × 2 columns).

# %%
heat_rates_new_entrant_heat_rates = parse_spec('Heat rates', {'name': 'New entrant heat rates', 'range': 'H7:I31'})
show_table(heat_rates_new_entrant_heat_rates)

# %% [markdown]
# ## Auxiliary
#
# Auxiliary (self) load for each generator or generator class.

# %% [markdown]
# ### Existing generator auxiliary load
#
# Source block: `Auxiliary!B5:E736` (732 rows × 4 columns).

# %%
auxiliary_existing_generator_auxiliary_load = parse_spec('Auxiliary', {'name': 'Existing generator auxiliary load', 'range': 'B5:E736'})
show_table(auxiliary_existing_generator_auxiliary_load)

# %% [markdown]
# ### New entrant auxiliary load
#
# Source block: `Auxiliary!G5:H29` (25 rows × 2 columns).

# %%
auxiliary_new_entrant_auxiliary_load = parse_spec('Auxiliary', {'name': 'New entrant auxiliary load', 'range': 'G5:H29'})
show_table(auxiliary_new_entrant_auxiliary_load)

# %% [markdown]
# ## Storage properties
#
# Battery storage to power ratio and round-trip efficiency.

# %% [markdown]
# ### Battery properties
#
# Source block: `Storage properties!B2:J19` (18 rows × 9 columns).

# %%
storage_properties_battery_properties = parse_spec('Storage properties', {'name': 'Battery properties', 'range': 'B2:J19'})
show_table(storage_properties_battery_properties)

# %% [markdown]
# ### Existing pumped-hydro properties
#
# Source block: `Storage properties!B21:E35` (15 rows × 4 columns).

# %%
storage_properties_existing_pumped_hydro_properties = parse_spec('Storage properties', {'name': 'Existing pumped-hydro properties', 'range': 'B21:E35'})
show_table(storage_properties_existing_pumped_hydro_properties)

# %% [markdown]
# ### New entrant pumped-hydro properties
#
# Source block: `Storage properties!G21:J27` (7 rows × 4 columns).

# %%
storage_properties_new_entrant_pumped_hydro_properties = parse_spec('Storage properties', {'name': 'New entrant pumped-hydro properties', 'range': 'G21:J27'})
show_table(storage_properties_new_entrant_pumped_hydro_properties)

# %% [markdown]
# ### On-site diesel storage
#
# Source block: `Storage properties!B38:C45` (8 rows × 2 columns).

# %%
storage_properties_on_site_diesel_storage = parse_spec('Storage properties', {'name': 'On-site diesel storage', 'range': 'B38:C45'})
show_table(storage_properties_on_site_diesel_storage)

# %% [markdown]
# ## Emissions intensity
#
# Emissions production per MWh of output for each generator or generator class.

# %% [markdown]
# ### Existing generator emissions intensity
#
# Source block: `Emissions intensity!B4:E744` (741 rows × 4 columns).

# %%
emissions_intensity_existing_generator_emissions_intensity = parse_spec('Emissions intensity', {'name': 'Existing generator emissions intensity', 'range': 'B4:E744'})
show_table(emissions_intensity_existing_generator_emissions_intensity)

# %% [markdown]
# ### New entrant emissions intensity
#
# Source block: `Emissions intensity!G4:H29` (26 rows × 2 columns).

# %%
emissions_intensity_new_entrant_emissions_intensity = parse_spec('Emissions intensity', {'name': 'New entrant emissions intensity', 'range': 'G4:H29'})
show_table(emissions_intensity_new_entrant_emissions_intensity)

# %% [markdown]
# ## Build costs
#
# Capital Cost projections for new entrant generators.

# %% [markdown]
# ### New entrant capital-cost projections
#
# Source block: `Build costs!B2:AJ77` (76 rows × 35 columns).

# %%
build_costs_new_entrant_capital_cost_projections = parse_spec('Build costs', {'name': 'New entrant capital-cost projections', 'range': 'B2:AJ77'})
show_table(build_costs_new_entrant_capital_cost_projections)

# %% [markdown]
# ## Fixed OPEX
#
# Fixed operating cost regardless of output for each generator or generator class.

# %% [markdown]
# ### Existing generator fixed OPEX
#
# Source block: `Fixed OPEX!B5:E739` (735 rows × 4 columns).

# %%
fixed_opex_existing_generator_fixed_opex = parse_spec('Fixed OPEX', {'name': 'Existing generator fixed OPEX', 'range': 'B5:E739'})
show_table(fixed_opex_existing_generator_fixed_opex)

# %% [markdown]
# ### New entrant fixed OPEX
#
# Source block: `Fixed OPEX!G5:I32` (28 rows × 3 columns).

# %%
fixed_opex_new_entrant_fixed_opex = parse_spec('Fixed OPEX', {'name': 'New entrant fixed OPEX', 'range': 'G5:I32'})
show_table(fixed_opex_new_entrant_fixed_opex)

# %% [markdown]
# ## Variable OPEX
#
# Variable operating cost per MWh of output for each generator or generator class.

# %% [markdown]
# ### Existing generator variable OPEX
#
# Source block: `Variable OPEX!B5:E738` (734 rows × 4 columns).

# %%
variable_opex_existing_generator_variable_opex = parse_spec('Variable OPEX', {'name': 'Existing generator variable OPEX', 'range': 'B5:E738'})
show_table(variable_opex_existing_generator_variable_opex)

# %% [markdown]
# ### New entrant variable OPEX
#
# Source block: `Variable OPEX!G5:H32` (28 rows × 2 columns).

# %%
variable_opex_new_entrant_variable_opex = parse_spec('Variable OPEX', {'name': 'New entrant variable OPEX', 'range': 'G5:H32'})
show_table(variable_opex_new_entrant_variable_opex)

# %% [markdown]
# ## Marginal Loss Factors
#
# Marginal loss factors for each generator.

# %% [markdown]
# ### Existing generator marginal loss factors
#
# Source block: `Marginal Loss Factors!B10:F748` (739 rows × 5 columns).

# %%
marginal_loss_factors_existing_generator_marginal_loss_factors = parse_spec('Marginal Loss Factors', {'name': 'Existing generator marginal loss factors', 'range': 'B10:F748'})
show_table(marginal_loss_factors_existing_generator_marginal_loss_factors)

# %% [markdown]
# ### New entrant generator marginal loss factors
#
# Source block: `Marginal Loss Factors!I10:M536` (527 rows × 5 columns).

# %%
marginal_loss_factors_new_entrant_generator_marginal_loss_factors = parse_spec('Marginal Loss Factors', {'name': 'New entrant generator marginal loss factors', 'range': 'I10:M536'})
show_table(marginal_loss_factors_new_entrant_generator_marginal_loss_factors)

# %% [markdown]
# ### New entrant electrolyser marginal loss factors
#
# Source block: `Marginal Loss Factors!O10:S161` (152 rows × 5 columns).

# %%
marginal_loss_factors_new_entrant_electrolyser_marginal_loss_factors = parse_spec('Marginal Loss Factors', {'name': 'New entrant electrolyser marginal loss factors', 'range': 'O10:S161'})
show_table(marginal_loss_factors_new_entrant_electrolyser_marginal_loss_factors)

# %% [markdown]
# ## Locational Cost Factors
#
# Locational cost factors provide an indication of the variation in new entrants generators cost based on the shift in labour, equipment and shipping/delivery cost between regions.

# %% [markdown]
# ### Non-pumped-hydro locational cost factors
#
# Source block: `Locational Cost Factors!B9:H80` (72 rows × 7 columns).

# %%
locational_cost_factors_non_pumped_hydro_locational_cost_factors = parse_spec('Locational Cost Factors', {'name': 'Non-pumped-hydro locational cost factors', 'range': 'B9:H80'})
show_table(locational_cost_factors_non_pumped_hydro_locational_cost_factors)

# %% [markdown]
# ### Pumped-hydro locational cost factors
#
# Source block: `Locational Cost Factors!B83:I132` (50 rows × 8 columns).

# %%
locational_cost_factors_pumped_hydro_locational_cost_factors = parse_spec('Locational Cost Factors', {'name': 'Pumped-hydro locational cost factors', 'range': 'B83:I132'})
show_table(locational_cost_factors_pumped_hydro_locational_cost_factors)

# %% [markdown]
# ### Technology cost breakdown ratios
#
# Source block: `Locational Cost Factors!B134:G158` (25 rows × 6 columns).

# %%
locational_cost_factors_technology_cost_breakdown_ratios = parse_spec('Locational Cost Factors', {'name': 'Technology cost breakdown ratios', 'range': 'B134:G158'})
show_table(locational_cost_factors_technology_cost_breakdown_ratios)

# %% [markdown]
# ### Technology-specific locational cost factors
#
# Source block: `Locational Cost Factors!B161:X227` (67 rows × 23 columns).

# %%
locational_cost_factors_technology_specific_locational_cost_factors = parse_spec('Locational Cost Factors', {'name': 'Technology-specific locational cost factors', 'range': 'B161:X227'})
show_table(locational_cost_factors_technology_specific_locational_cost_factors)

# %% [markdown]
# ## Build limits - REZs
#
# Modelled limitations for REZs impacting build constraints within the expansion modelling.

# %% [markdown]
# ### Initial REZ resource limits
#
# Source block: `Build limits - REZs!B2:Q62` (61 rows × 16 columns).

# %%
build_limits_rezs_initial_rez_resource_limits = parse_spec('Build limits - REZs', {'name': 'Initial REZ resource limits', 'range': 'B2:Q62'})
show_table(build_limits_rezs_initial_rez_resource_limits)

# %% [markdown]
# ### Initial REZ transmission limits
#
# Source block: `Build limits - REZs!B64:N119` (56 rows × 13 columns).

# %%
build_limits_rezs_initial_rez_transmission_limits = parse_spec('Build limits - REZs', {'name': 'Initial REZ transmission limits', 'range': 'B64:N119'})
show_table(build_limits_rezs_initial_rez_transmission_limits)

# %% [markdown]
# ### REZ transmission modifiers
#
# Source block: `Build limits - REZs!B121:F132` (12 rows × 5 columns).

# %%
build_limits_rezs_rez_transmission_modifiers = parse_spec('Build limits - REZs', {'name': 'REZ transmission modifiers', 'range': 'B121:F132'})
show_table(build_limits_rezs_rez_transmission_modifiers)

# %% [markdown]
# ### REZ group constraints
#
# Source block: `Build limits - REZs!B136:K265` (130 rows × 10 columns).

# %%
build_limits_rezs_rez_group_constraints = parse_spec('Build limits - REZs', {'name': 'REZ group constraints', 'range': 'B136:K265'})
show_table(build_limits_rezs_rez_group_constraints)

# %% [markdown]
# ### REZ transmission limit constraints
#
# Source block: `Build limits - REZs!B267:K317` (51 rows × 10 columns).

# %%
build_limits_rezs_rez_transmission_limit_constraints = parse_spec('Build limits - REZs', {'name': 'REZ transmission limit constraints', 'range': 'B267:K317'})
show_table(build_limits_rezs_rez_transmission_limit_constraints)

# %% [markdown]
# ### REZ secondary transmission limits
#
# Source block: `Build limits - REZs!B319:K335` (17 rows × 10 columns).

# %%
build_limits_rezs_rez_secondary_transmission_limits = parse_spec('Build limits - REZs', {'name': 'REZ secondary transmission limits', 'range': 'B319:K335'})
show_table(build_limits_rezs_rez_secondary_transmission_limits)

# %% [markdown]
# ### Non-REZ connections pipeline build limits
#
# Source block: `Build limits - REZs!B337:E356` (20 rows × 4 columns).

# %%
build_limits_rezs_non_rez_connections_pipeline_build_limits = parse_spec('Build limits - REZs', {'name': 'Non-REZ connections pipeline build limits', 'range': 'B337:E356'})
show_table(build_limits_rezs_non_rez_connections_pipeline_build_limits)

# %% [markdown]
# ### REZ technology-specific access-right limits
#
# Source block: `Build limits - REZs!B358:G368` (11 rows × 6 columns).

# %%
build_limits_rezs_rez_technology_specific_access_right_limits = parse_spec('Build limits - REZs', {'name': 'REZ technology-specific access-right limits', 'range': 'B358:G368'})
show_table(build_limits_rezs_rez_technology_specific_access_right_limits)

# %% [markdown]
# ## Build limits - PHES
#
# Modelled limitations for PHES impacting build constraints within the expansion modelling.

# %% [markdown]
# ### Pumped-hydro build limits
#
# Source block: `Build limits - PHES!B2:W27` (26 rows × 22 columns).

# %%
build_limits_phes_pumped_hydro_build_limits = parse_spec('Build limits - PHES', {'name': 'Pumped-hydro build limits', 'range': 'B2:W27'})
show_table(build_limits_phes_pumped_hydro_build_limits)

# %% [markdown]
# ## First-of-a-kind premium
#
# Premiums applied to emerging generation technologies to reflect the tendency of first-of-a-kind installations to exceed estimated costs.

# %% [markdown]
# ### First-of-a-kind premium factors
#
# Source block: `First-of-a-kind premium!B2:D11` (10 rows × 3 columns).

# %%
first_of_a_kind_premium_first_of_a_kind_premium_factors = parse_spec('First-of-a-kind premium', {'name': 'First-of-a-kind premium factors', 'range': 'B2:D11'})
show_table(first_of_a_kind_premium_first_of_a_kind_premium_factors)

# %% [markdown]
# ## Lead time and project life
#
# Modelled limitations impacting build timings within the expansion modelling.

# %% [markdown]
# ### Lead times and project lives
#
# Source block: `Lead time and project life!B2:H35` (34 rows × 7 columns).

# %%
lead_time_and_project_life_lead_times_and_project_lives = parse_spec('Lead time and project life', {'name': 'Lead times and project lives', 'range': 'B2:H35'})
show_table(lead_time_and_project_life_lead_times_and_project_lives)

# %% [markdown]
# ## Financial parameters
#
# Financial parameters (discount rate, weighted average cost of capital, value of customer reliability, and value of emissions reductions) used during cost benefit analysis.

# %% [markdown]
# ### Discount rate
#
# Source block: `Financial parameters!B2:F7` (6 rows × 5 columns).

# %%
financial_parameters_discount_rate = parse_spec('Financial parameters', {'name': 'Discount rate', 'range': 'B2:F7'})
show_table(financial_parameters_discount_rate)

# %% [markdown]
# ### Weighted Average Cost of Capital
#
# Source block: `Financial parameters!B10:F41` (32 rows × 5 columns).

# %%
financial_parameters_weighted_average_cost_of_capital = parse_spec('Financial parameters', {'name': 'Weighted Average Cost of Capital', 'range': 'B10:F41'})
show_table(financial_parameters_weighted_average_cost_of_capital)

# %% [markdown]
# ### Value of Customer Reliability
#
# Source block: `Financial parameters!B43:G51` (9 rows × 6 columns).

# %%
financial_parameters_value_of_customer_reliability = parse_spec('Financial parameters', {'name': 'Value of Customer Reliability', 'range': 'B43:G51'})
show_table(financial_parameters_value_of_customer_reliability)

# %% [markdown]
# ### Value of emissions reduction
#
# Source block: `Financial parameters!B54:C90` (37 rows × 2 columns).

# %%
financial_parameters_value_of_emissions_reduction = parse_spec('Financial parameters', {'name': 'Value of emissions reduction', 'range': 'B54:C90'})
show_table(financial_parameters_value_of_emissions_reduction)

# %% [markdown]
# ## Affine Heat rates
#
# Heat rate curves for large thermal units.

# %% [markdown]
# ### Existing generator affine heat rates
#
# Source block: `Affine Heat rates!B6:F192` (187 rows × 5 columns).

# %%
affine_heat_rates_existing_generator_affine_heat_rates = parse_spec('Affine Heat rates', {'name': 'Existing generator affine heat rates', 'range': 'B6:F192'})
show_table(affine_heat_rates_existing_generator_affine_heat_rates)

# %% [markdown]
# ### New entrant affine heat rates
#
# Source block: `Affine Heat rates!H6:K29` (24 rows × 4 columns).

# %%
affine_heat_rates_new_entrant_affine_heat_rates = parse_spec('Affine Heat rates', {'name': 'New entrant affine heat rates', 'range': 'H6:K29'})
show_table(affine_heat_rates_new_entrant_affine_heat_rates)

# %% [markdown]
# ## Max Ramp Rates
#
# Maximum rates of change for thermal unit output up and down.

# %% [markdown]
# ### Existing thermal generator maximum ramp rates
#
# Source block: `Max Ramp Rates!B7:F191` (185 rows × 5 columns).

# %%
max_ramp_rates_existing_thermal_generator_maximum_ramp_rates = parse_spec('Max Ramp Rates', {'name': 'Existing thermal generator maximum ramp rates', 'range': 'B7:F191'})
show_table(max_ramp_rates_existing_thermal_generator_maximum_ramp_rates)

# %% [markdown]
# ### New entrant maximum ramp rates
#
# Source block: `Max Ramp Rates!H7:J30` (24 rows × 3 columns).

# %%
max_ramp_rates_new_entrant_maximum_ramp_rates = parse_spec('Max Ramp Rates', {'name': 'New entrant maximum ramp rates', 'range': 'H7:J30'})
show_table(max_ramp_rates_new_entrant_maximum_ramp_rates)

# %% [markdown]
# ## Coal Min Stable Level
#
# Coal generator minimum stable levels.

# %% [markdown]
# ### Coal generator minimum stable levels
#
# Source block: `Coal Min Stable Level!B2:G63` (62 rows × 6 columns).

# %%
coal_min_stable_level_coal_generator_minimum_stable_levels = parse_spec('Coal Min Stable Level', {'name': 'Coal generator minimum stable levels', 'range': 'B2:G63'})
show_table(coal_min_stable_level_coal_generator_minimum_stable_levels)

# %% [markdown]
# ## GPG Min Stable Level
#
# Minimum operating levels for large GPG units.

# %% [markdown]
# ### Existing GPG minimum stable levels
#
# Source block: `GPG Min Stable Level!B10:E150` (141 rows × 4 columns).

# %%
gpg_min_stable_level_existing_gpg_minimum_stable_levels = parse_spec('GPG Min Stable Level', {'name': 'Existing GPG minimum stable levels', 'range': 'B10:E150'})
show_table(gpg_min_stable_level_existing_gpg_minimum_stable_levels)

# %% [markdown]
# ### New entrant GPG minimum stable levels
#
# Source block: `GPG Min Stable Level!G10:H35` (26 rows × 2 columns).

# %%
gpg_min_stable_level_new_entrant_gpg_minimum_stable_levels = parse_spec('GPG Min Stable Level', {'name': 'New entrant GPG minimum stable levels', 'range': 'G10:H35'})
show_table(gpg_min_stable_level_new_entrant_gpg_minimum_stable_levels)

# %% [markdown]
# ## Coal and Biomass price
#
# Coal fuel price for each coal generator.

# %% [markdown]
# ### Coal fuel prices
#
# Source block: `Coal and Biomass price!B8:AG54` (47 rows × 32 columns).

# %%
coal_and_biomass_price_coal_fuel_prices = parse_spec('Coal and Biomass price', {'name': 'Coal fuel prices', 'range': 'B8:AG54'})
show_table(coal_and_biomass_price_coal_fuel_prices)

# %% [markdown]
# ### Biomass fuel prices
#
# Source block: `Coal and Biomass price!B57:AG61` (5 rows × 32 columns).

# %%
coal_and_biomass_price_biomass_fuel_prices = parse_spec('Coal and Biomass price', {'name': 'Biomass fuel prices', 'range': 'B57:AG61'})
show_table(coal_and_biomass_price_biomass_fuel_prices)

# %% [markdown]
# ## Gas, Liquid fuel, H2 price
#
# Fuel price for each gas and liquid fuel generator.

# %% [markdown]
# ### Existing GPG fuel costs
#
# Source block: `Gas, Liquid fuel, H2 price!B7:AG129` (123 rows × 32 columns).

# %%
gas_liquid_fuel_h2_price_existing_gpg_fuel_costs = parse_spec('Gas, Liquid fuel, H2 price', {'name': 'Existing GPG fuel costs', 'range': 'B7:AG129'})
show_table(gas_liquid_fuel_h2_price_existing_gpg_fuel_costs)

# %% [markdown]
# ### New entrant GPG fuel costs
#
# Source block: `Gas, Liquid fuel, H2 price!B132:AG224` (93 rows × 32 columns).

# %%
gas_liquid_fuel_h2_price_new_entrant_gpg_fuel_costs = parse_spec('Gas, Liquid fuel, H2 price', {'name': 'New entrant GPG fuel costs', 'range': 'B132:AG224'})
show_table(gas_liquid_fuel_h2_price_new_entrant_gpg_fuel_costs)

# %% [markdown]
# ### Industrial fuel costs
#
# Source block: `Gas, Liquid fuel, H2 price!B228:AG249` (22 rows × 32 columns).

# %%
gas_liquid_fuel_h2_price_industrial_fuel_costs = parse_spec('Gas, Liquid fuel, H2 price', {'name': 'Industrial fuel costs', 'range': 'B228:AG249'})
show_table(gas_liquid_fuel_h2_price_industrial_fuel_costs)

# %% [markdown]
# ### Residential and commercial fuel costs
#
# Source block: `Gas, Liquid fuel, H2 price!B253:AG274` (22 rows × 32 columns).

# %%
gas_liquid_fuel_h2_price_residential_and_commercial_fuel_costs = parse_spec('Gas, Liquid fuel, H2 price', {'name': 'Residential and commercial fuel costs', 'range': 'B253:AG274'})
show_table(gas_liquid_fuel_h2_price_residential_and_commercial_fuel_costs)

# %% [markdown]
# ### Liquid fuel prices
#
# Source block: `Gas, Liquid fuel, H2 price!B278:AG302` (25 rows × 32 columns).

# %%
gas_liquid_fuel_h2_price_liquid_fuel_prices = parse_spec('Gas, Liquid fuel, H2 price', {'name': 'Liquid fuel prices', 'range': 'B278:AG302'})
show_table(gas_liquid_fuel_h2_price_liquid_fuel_prices)

# %% [markdown]
# ### GPG secondary liquid-fuel prices
#
# Source block: `Gas, Liquid fuel, H2 price!B305:AG429` (125 rows × 32 columns).

# %%
gas_liquid_fuel_h2_price_gpg_secondary_liquid_fuel_prices = parse_spec('Gas, Liquid fuel, H2 price', {'name': 'GPG secondary liquid-fuel prices', 'range': 'B305:AG429'})
show_table(gas_liquid_fuel_h2_price_gpg_secondary_liquid_fuel_prices)

# %% [markdown]
# ### Hydrogen prices
#
# Source block: `Gas, Liquid fuel, H2 price!B433:AG438` (6 rows × 32 columns).

# %%
gas_liquid_fuel_h2_price_hydrogen_prices = parse_spec('Gas, Liquid fuel, H2 price', {'name': 'Hydrogen prices', 'range': 'B433:AG438'})
show_table(gas_liquid_fuel_h2_price_hydrogen_prices)

# %% [markdown]
# ### Biomethane prices
#
# Source block: `Gas, Liquid fuel, H2 price!B440:AG452` (13 rows × 32 columns).

# %%
gas_liquid_fuel_h2_price_biomethane_prices = parse_spec('Gas, Liquid fuel, H2 price', {'name': 'Biomethane prices', 'range': 'B440:AG452'})
show_table(gas_liquid_fuel_h2_price_biomethane_prices)

# %% [markdown]
# ## Gas System Properties
#
# Key properties of Natural Gas Pipelines, Processing Facilities, Storage Facilities, Reserves and Resources, Pipeline Transmission Tariffs, and Production Costs.

# %% [markdown]
# ### Gas pipelines
#
# Source block: `Gas System Properties!B7:F49` (43 rows × 5 columns).

# %%
gas_system_properties_gas_pipelines = parse_spec('Gas System Properties', {'name': 'Gas pipelines', 'range': 'B7:F49'})
show_table(gas_system_properties_gas_pipelines)

# %% [markdown]
# ### Gas processing facilities
#
# Source block: `Gas System Properties!B51:G105` (55 rows × 6 columns).

# %%
gas_system_properties_gas_processing_facilities = parse_spec('Gas System Properties', {'name': 'Gas processing facilities', 'range': 'B51:G105'})
show_table(gas_system_properties_gas_processing_facilities)

# %% [markdown]
# ### Gas storage facilities
#
# Source block: `Gas System Properties!B108:H122` (15 rows × 7 columns).

# %%
gas_system_properties_gas_storage_facilities = parse_spec('Gas System Properties', {'name': 'Gas storage facilities', 'range': 'B108:H122'})
show_table(gas_system_properties_gas_storage_facilities)

# %% [markdown]
# ### Gas reserves and resources
#
# Source block: `Gas System Properties!B124:F144` (21 rows × 5 columns).

# %%
gas_system_properties_gas_reserves_and_resources = parse_spec('Gas System Properties', {'name': 'Gas reserves and resources', 'range': 'B124:F144'})
show_table(gas_system_properties_gas_reserves_and_resources)

# %% [markdown]
# ### Pipeline transmission tariffs
#
# Source block: `Gas System Properties!B146:E169` (24 rows × 4 columns).

# %%
gas_system_properties_pipeline_transmission_tariffs = parse_spec('Gas System Properties', {'name': 'Pipeline transmission tariffs', 'range': 'B146:E169'})
show_table(gas_system_properties_pipeline_transmission_tariffs)

# %% [markdown]
# ### Gas production costs
#
# Source block: `Gas System Properties!B171:E185` (15 rows × 4 columns).

# %%
gas_system_properties_gas_production_costs = parse_spec('Gas System Properties', {'name': 'Gas production costs', 'range': 'B171:E185'})
show_table(gas_system_properties_gas_production_costs)

# %% [markdown]
# ## GPG emissions reduction - BioM
#
# Factors projecting emissions reduction from GPG due to blending of biomethane into fuel gas.

# %% [markdown]
# ### GPG biomethane-blending emissions factor
#
# Source block: `GPG emissions reduction - BioM!B2:AF12` (11 rows × 31 columns).

# %%
gpg_emissions_reduction_biom_gpg_biomethane_blending_emissions_factor = parse_spec('GPG emissions reduction - BioM', {'name': 'GPG biomethane-blending emissions factor', 'range': 'B2:AF12'})
show_table(gpg_emissions_reduction_biom_gpg_biomethane_blending_emissions_factor)

# %% [markdown]
# ## Power System Security
#
# Reflect power system constraints to reflect secure operating limits.

# %% [markdown]
# ### Coal-retirement minimum-fault-level costs
#
# Source block: `Power System Security!B4:D49` (46 rows × 3 columns).

# %%
power_system_security_coal_retirement_minimum_fault_level_costs = parse_spec('Power System Security', {'name': 'Coal-retirement minimum-fault-level costs', 'range': 'B4:D49'})
show_table(power_system_security_coal_retirement_minimum_fault_level_costs)

# %% [markdown]
# ### Efficient system-strength costs
#
# Source block: `Power System Security!B52:AE56` (5 rows × 30 columns).

# %%
power_system_security_efficient_system_strength_costs = parse_spec('Power System Security', {'name': 'Efficient system-strength costs', 'range': 'B52:AE56'})
show_table(power_system_security_efficient_system_strength_costs)

# %% [markdown]
# ### Synchronous unit commitment — standard scenarios
#
# Source block: `Power System Security!B58:G72` (15 rows × 6 columns).

# %%
power_system_security_synchronous_unit_commitment_standard_scenarios = parse_spec('Power System Security', {'name': 'Synchronous unit commitment — standard scenarios', 'range': 'B58:G72'})
show_table(power_system_security_synchronous_unit_commitment_standard_scenarios)

# %% [markdown]
# ### Synchronous unit commitment — Accelerated Transition
#
# Source block: `Power System Security!B74:G94` (21 rows × 6 columns).

# %%
power_system_security_synchronous_unit_commitment_accelerated_transition = parse_spec('Power System Security', {'name': 'Synchronous unit commitment — Accelerated Transition', 'range': 'B74:G94'})
show_table(power_system_security_synchronous_unit_commitment_accelerated_transition)

# %% [markdown]
# ## Reserves
#
# Minimum reserve levels for reliable regional supply.

# %% [markdown]
# ### Initial regional reserves
#
# Source block: `Reserves!B2:C14` (13 rows × 2 columns).

# %%
reserves_initial_regional_reserves = parse_spec('Reserves', {'name': 'Initial regional reserves', 'range': 'B2:C14'})
show_table(reserves_initial_regional_reserves)

# %% [markdown]
# ## Hydrogen demand - Domestic
#
# Projections of domestic hydrogen demand.

# %% [markdown]
# ### Domestic hydrogen demand
#
# Source block: `Hydrogen demand - Domestic!B2:AH53` (52 rows × 33 columns).

# %%
hydrogen_demand_domestic_domestic_hydrogen_demand = parse_spec('Hydrogen demand - Domestic', {'name': 'Domestic hydrogen demand', 'range': 'B2:AH53'})
show_table(hydrogen_demand_domestic_domestic_hydrogen_demand)

# %% [markdown]
# ## Hydrogen monthly profiles
#
# Projections of monthly profiles of export/commodities hydrogen demand.

# %% [markdown]
# ### Hydrogen monthly consumption profiles
#
# Source block: `Hydrogen monthly profiles!B2:AG44` (43 rows × 32 columns).

# %%
hydrogen_monthly_profiles_hydrogen_monthly_consumption_profiles = parse_spec('Hydrogen monthly profiles', {'name': 'Hydrogen monthly consumption profiles', 'range': 'B2:AG44'})
show_table(hydrogen_monthly_profiles_hydrogen_monthly_consumption_profiles)

# %% [markdown]
# ## Hydrogen demand-Export&Commod
#
# Projections of demand for export hydrogen and green commodities production.

# %% [markdown]
# ### Hydrogen export demand
#
# Source block: `Hydrogen demand-Export&Commod!B2:AH52` (51 rows × 33 columns).

# %%
hydrogen_demand_export_and_commod_hydrogen_export_demand = parse_spec('Hydrogen demand-Export&Commod', {'name': 'Hydrogen export demand', 'range': 'B2:AH52'})
show_table(hydrogen_demand_export_and_commod_hydrogen_export_demand)

# %% [markdown]
# ### Hydrogen demand for green commodities
#
# Source block: `Hydrogen demand-Export&Commod!B54:AH105` (52 rows × 33 columns).

# %%
hydrogen_demand_export_and_commod_hydrogen_demand_for_green_commodities = parse_spec('Hydrogen demand-Export&Commod', {'name': 'Hydrogen demand for green commodities', 'range': 'B54:AH105'})
show_table(hydrogen_demand_export_and_commod_hydrogen_demand_for_green_commodities)

# %% [markdown]
# ### Electricity demand for green steel
#
# Source block: `Hydrogen demand-Export&Commod!B107:AH156` (50 rows × 33 columns).

# %%
hydrogen_demand_export_and_commod_electricity_demand_for_green_steel = parse_spec('Hydrogen demand-Export&Commod', {'name': 'Electricity demand for green steel', 'range': 'B107:AH156'})
show_table(hydrogen_demand_export_and_commod_electricity_demand_for_green_steel)

# %% [markdown]
# ## Hydrogen consumption locations
#
# Location of hydrogen consumption and candidate hydrogen hubs and ports.

# %% [markdown]
# ### Hydrogen consumption locations
#
# Source block: `Hydrogen consumption locations!B5:F40` (36 rows × 5 columns).

# %%
hydrogen_consumption_locations_hydrogen_consumption_locations = parse_spec('Hydrogen consumption locations', {'name': 'Hydrogen consumption locations', 'range': 'B5:F40'})
show_table(hydrogen_consumption_locations_hydrogen_consumption_locations)

# %% [markdown]
# ### Hydrogen hubs
#
# Source block: `Hydrogen consumption locations!B42:B44` (3 rows × 1 columns).

# %%
hydrogen_consumption_locations_hydrogen_hubs = parse_spec('Hydrogen consumption locations', {'name': 'Hydrogen hubs', 'range': 'B42:B44'})
show_table(hydrogen_consumption_locations_hydrogen_hubs)

# %% [markdown]
# ### Hydrogen and commodity export ports
#
# Source block: `Hydrogen consumption locations!B46:D57` (12 rows × 3 columns).

# %%
hydrogen_consumption_locations_hydrogen_and_commodity_export_ports = parse_spec('Hydrogen consumption locations', {'name': 'Hydrogen and commodity export ports', 'range': 'B46:D57'})
show_table(hydrogen_consumption_locations_hydrogen_and_commodity_export_ports)

# %% [markdown]
# ## Water for Hydrogen
#
# Projections of treated water required for electrolytic hydrogen production.

# %% [markdown]
# ### Water required for hydrogen production
#
# Source block: `Water for Hydrogen!B2:AH52` (51 rows × 33 columns).

# %%
water_for_hydrogen_water_required_for_hydrogen_production = parse_spec('Water for Hydrogen', {'name': 'Water required for hydrogen production', 'range': 'B2:AH52'})
show_table(water_for_hydrogen_water_required_for_hydrogen_production)

# %% [markdown]
# ## Desalination demand for H2
#
# Projections of electricity required for water treatment associated with electrolytic hydrogen production.

# %% [markdown]
# ### Desalination electricity demand for hydrogen
#
# Source block: `Desalination demand for H2!B2:AH52` (51 rows × 33 columns).

# %%
desalination_demand_for_h2_desalination_electricity_demand_for_hydrogen = parse_spec('Desalination demand for H2', {'name': 'Desalination electricity demand for hydrogen', 'range': 'B2:AH52'})
show_table(desalination_demand_for_h2_desalination_electricity_demand_for_hydrogen)

# %% [markdown]
# ## H2 as fuel for GPG Limit
#
# Maximum fuel limit applied to hydrogen for GPG.

# %% [markdown]
# ### Hydrogen-as-GPG-fuel limit
#
# Source block: `H2 as fuel for GPG Limit!B2:AG21` (20 rows × 32 columns).

# %%
h2_as_fuel_for_gpg_limit_hydrogen_as_gpg_fuel_limit = parse_spec('H2 as fuel for GPG Limit', {'name': 'Hydrogen-as-GPG-fuel limit', 'range': 'B2:AG21'})
show_table(h2_as_fuel_for_gpg_limit_hydrogen_as_gpg_fuel_limit)

# %% [markdown]
# ## Build Cost - Hydrogen pipeline
#
# Projections of cost to build hydrogen pipelines in a given REZ.

# %% [markdown]
# ### Hydrogen pipeline build costs
#
# Source block: `Build Cost - Hydrogen pipeline!B2:AJ156` (155 rows × 35 columns).

# %%
build_cost_hydrogen_pipeline_hydrogen_pipeline_build_costs = parse_spec('Build Cost - Hydrogen pipeline', {'name': 'Hydrogen pipeline build costs', 'range': 'B2:AJ156'})
show_table(build_cost_hydrogen_pipeline_hydrogen_pipeline_build_costs)

# %% [markdown]
# ## Other hydrogen assumptions
#
# Various hydrogen-related assumptions.

# %% [markdown]
# ### Hydrogen heat content
#
# Source block: `Other hydrogen assumptions!B2:C5` (4 rows × 2 columns).

# %%
other_hydrogen_assumptions_hydrogen_heat_content = parse_spec('Other hydrogen assumptions', {'name': 'Hydrogen heat content', 'range': 'B2:C5'})
show_table(other_hydrogen_assumptions_hydrogen_heat_content)

# %% [markdown]
# ### Electrolyser electricity consumption rate
#
# Source block: `Other hydrogen assumptions!B7:AF11` (5 rows × 31 columns).

# %%
other_hydrogen_assumptions_electrolyser_electricity_consumption_rate = parse_spec('Other hydrogen assumptions', {'name': 'Electrolyser electricity consumption rate', 'range': 'B7:AF11'})
show_table(other_hydrogen_assumptions_electrolyser_electricity_consumption_rate)

# %% [markdown]
# ### Water cost for hydrogen production
#
# Source block: `Other hydrogen assumptions!B13:AF17` (5 rows × 31 columns).

# %%
other_hydrogen_assumptions_water_cost_for_hydrogen_production = parse_spec('Other hydrogen assumptions', {'name': 'Water cost for hydrogen production', 'range': 'B13:AF17'})
show_table(other_hydrogen_assumptions_water_cost_for_hydrogen_production)

# %% [markdown]
# ### Ammonia-conversion electricity adjustment
#
# Source block: `Other hydrogen assumptions!B19:AF23` (5 rows × 31 columns).

# %%
other_hydrogen_assumptions_ammonia_conversion_electricity_adjustment = parse_spec('Other hydrogen assumptions', {'name': 'Ammonia-conversion electricity adjustment', 'range': 'B19:AF23'})
show_table(other_hydrogen_assumptions_ammonia_conversion_electricity_adjustment)

# %% [markdown]
# ### Minimum annual electrolyser utilisation factor
#
# Source block: `Other hydrogen assumptions!B26:AF30` (5 rows × 31 columns).

# %%
other_hydrogen_assumptions_minimum_annual_electrolyser_utilisation_factor = parse_spec('Other hydrogen assumptions', {'name': 'Minimum annual electrolyser utilisation factor', 'range': 'B26:AF30'})
show_table(other_hydrogen_assumptions_minimum_annual_electrolyser_utilisation_factor)

# %% [markdown]
# ### Electrolyser balance of plant
#
# Source block: `Other hydrogen assumptions!B32:C35` (4 rows × 2 columns).

# %%
other_hydrogen_assumptions_electrolyser_balance_of_plant = parse_spec('Other hydrogen assumptions', {'name': 'Electrolyser balance of plant', 'range': 'B32:C35'})
show_table(other_hydrogen_assumptions_electrolyser_balance_of_plant)

# %% [markdown]
# ## Summary Mapping
#
# Master look-up table for generator, storage, and electrolyser assets.

# %% [markdown]
# ### Existing, committed, and anticipated asset mapping
#
# Source block: `Summary Mapping!C2:AF733` (732 rows × 30 columns).

# %%
summary_mapping_existing_committed_and_anticipated_asset_mapping = parse_spec('Summary Mapping', {'name': 'Existing, committed, and anticipated asset mapping', 'range': 'C2:AF733'})
show_table(summary_mapping_existing_committed_and_anticipated_asset_mapping)

# %% [markdown]
# ### Consumer energy resource mapping
#
# Source block: `Summary Mapping!C734:AF786` (53 rows × 30 columns).

# %%
summary_mapping_consumer_energy_resource_mapping = parse_spec('Summary Mapping', {'name': 'Consumer energy resource mapping', 'range': 'C734:AF786'})
show_table(summary_mapping_consumer_energy_resource_mapping)

# %% [markdown]
# ### New entrant asset mapping
#
# Source block: `Summary Mapping!C790:AF1316` (527 rows × 30 columns).

# %%
summary_mapping_new_entrant_asset_mapping = parse_spec('Summary Mapping', {'name': 'New entrant asset mapping', 'range': 'C790:AF1316'})
show_table(summary_mapping_new_entrant_asset_mapping)

# %% [markdown]
# ### New entrant electrolyser mapping
#
# Source block: `Summary Mapping!C1319:AF1381` (63 rows × 30 columns).

# %%
summary_mapping_new_entrant_electrolyser_mapping = parse_spec('Summary Mapping', {'name': 'New entrant electrolyser mapping', 'range': 'C1319:AF1381'})
show_table(summary_mapping_new_entrant_electrolyser_mapping)

# %% [markdown]
# ## Validation
#
# The checks below confirm worksheet coverage and the three source contracts already verified for ParseISP.

# %%
assert list(SHEET_RANGES) == workbook.sheetnames[:len(SHEET_RANGES)]
assert network_capability_flow_path_transfer_capability.shape == (18, 10) if len(SHEET_RANGES) >= 32 else True
assert transmission_reliability_transmission_unplanned_outage_rates.shape == (7, 4) if len(SHEET_RANGES) >= 34 else True
assert flow_path_augmentation_options_flow_path_augmentation_options.shape[0] == 62 if len(SHEET_RANGES) >= 38 else True

